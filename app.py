import base64
import hashlib
import hmac
import html
import json
import os
import secrets
import re
import time
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
import csv
import io
from datetime import datetime
from functools import wraps
from urllib.parse import parse_qs, urlparse

import mimetypes

import requests as req
from dotenv import load_dotenv
from flask import (
    Response,
    Flask,
    abort,
    jsonify,
    redirect,
    render_template,
    render_template_string,
    request,
    send_file,
    session,
    url_for,
)
from werkzeug.utils import secure_filename

import database as db
from services import analytics_importer, analytics_service, monitor_service, report_exporter

load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", secrets.token_hex(32))
app.config["SEND_FILE_MAX_AGE_DEFAULT"] = 0  # disable static file caching

# Two default profiles: the site editor (buraq) and the TV kiosk (sales).
# Extra accounts (second admin, guest viewers) exist ONLY when their env
# vars are set — no more default logins that nobody remembers creating.
ADMIN_LOGIN = os.getenv("ADMIN_LOGIN", "buraq")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "buraq1234")
ADMIN1_LOGIN = os.getenv("ADMIN1_LOGIN", "")
ADMIN1_PASSWORD = os.getenv("ADMIN1_PASSWORD", "")
GUEST_PASSWORD = os.getenv("GUEST_PASSWORD", "")
BOT_TOKEN = os.getenv("BOT_TOKEN", "").strip()
WEBHOOK_BASE_URL = os.getenv("WEBHOOK_BASE_URL", "").strip()
WEBHOOK_SECRET = os.getenv("WEBHOOK_SECRET", "").strip()
TELEGRAM_MAX_PHOTO_BYTES = 10 * 1024 * 1024
PORT = int(os.getenv("PORT", "5000"))
CHAT_ADMIN_CACHE_TTL = 300
CHAT_ADMIN_CACHE = {}
WELCOME_MEDIA_PENDING = set()
WELCOME_MEDIA_LOCK = threading.Lock()
WELCOME_MEDIA_SEMAPHORE = threading.Semaphore(2)
WELCOME_MEDIA_FILE_IDS = {"video": "", "voice": ""}
WELCOME_MEDIA_FILE_IDS_LOCK = threading.Lock()
TRACK_KEYBOARD_ANCHORS = {}
TRACK_KEYBOARD_ANCHORS_LOCK = threading.Lock()

ROLE_EDITOR = "editor"
ROLE_VIEWER = "viewer"
ROLE_KIOSK  = "kiosk"  # TV display profile — sees only Sales Monitor

# TV-display kiosk credentials (used for 50-inch monitor display in the office).
# These are intentionally hard-coded so the TV always has a working login.
KIOSK_LOGIN    = os.getenv("KIOSK_LOGIN", "sales")
KIOSK_PASSWORD = os.getenv("KIOSK_PASSWORD", "sales123")

ALLOWED_EXT = {"pdf", "png", "jpg", "jpeg", "xlsx", "xls", "xlsm", "doc", "docx", "zip"}
# Announcements additionally accept GIF and video attachments.
ANNOUNCEMENT_ALLOWED_EXT = ALLOWED_EXT | {"gif", "mp4", "mov", "m4v", "webp"}

TRACK_BUTTON = "Yuk holati"
TRACK_BUTTON_LABELS = {
    "uz_latn": "Yuk holati",
    "uz_cyrl": "Юк ҳолати",
    "ru": "Статус груза",
    "en": "Cargo status",
}
TRACK_BUTTON_TEXTS = set(TRACK_BUTTON_LABELS.values())
GROUP_REMOVE_COMMANDS = {"removebot", "leavebot", "botni_ochir", "hidekeyboard"}
MENU_RESTORE_COMMANDS = {"menu", "keyboard", "showmenu", "tugma", "knopka"}
AI_STATUS_COMMANDS = {"aistatus", "aidiag", "ai_status", "ai_diag"}
AI_TEST_COMMANDS = {"aitest", "ai_test"}
NO_ACTIVE_CARGO_MESSAGES = {
    "uz_latn": "Hozirgi vaqtda yo'lda kelayotgan yukingiz mavjud emas",
    "uz_cyrl": "Ҳозирги вақтда йўлда келаётган юкингиз мавжуд эмас",
    "ru": "В данный момент у вас нет груза в пути",
    "en": "You currently have no cargo in transit",
}
AI_ASK_BL_MESSAGES = {
    "uz_latn": "Kechirasiz, yuk holatini tekshirish uchun BL kodingizni yuboring.",
    "uz_cyrl": "Кечирасиз, юк ҳолатини текшириш учун BL кодингизни юборинг.",
    "ru": "Пожалуйста, отправьте BL-код, чтобы я мог проверить статус груза.",
    "en": "Please send your BL code so I can check the cargo status.",
}
AI_UNKNOWN_MESSAGES = {
    "uz_latn": "Kechirasiz, xabaringizni to'liq tushunmadim. Iltimos, BL kodingizni yuboring.",
    "uz_cyrl": "Кечирасиз, хабарингизни тўлиқ тушунмадим. Илтимос, BL кодингизни юборинг.",
    "ru": "Извините, я не до конца понял ваше сообщение. Пожалуйста, отправьте BL-код.",
    "en": "Sorry, I couldn't fully understand your message. Please send your BL code.",
}
TRACK_BUTTON_COOLDOWN_SECONDS = 60
TRACK_BUTTON_COOLDOWN_MESSAGES = {
    "uz_latn": "⏳ Iltimos, keyingi so'rov uchun <b>{seconds}</b> soniya kuting.",
    "uz_cyrl": "⏳ Илтимос, кейинги сўров учун <b>{seconds}</b> сония кутинг.",
    "ru": "⏳ Пожалуйста, подождите <b>{seconds}</b> сек. перед следующим запросом.",
    "en": "⏳ Please wait <b>{seconds}</b> sec. before your next request.",
}
CANCEL_BUTTON = "❌ Отмена"
STATE_WAITING_BL = "waiting_bl"
COMM_RATE_PREFIX = "comm_rate"
FILE_PREFIX = "file"
FILE_ALL_PREFIX = "fileall"  # callback_data = "fileall:<bl_id>"

# Localized label for the single "send all packing lists" button.
PACKING_LIST_DOWNLOAD_BUTTON_LABELS = {
    "uz_latn": "📦 Packing listlarni yuklab olish",
    "uz_cyrl": "📦 PACKING LIST юклаб олиш",
    "ru": "📦 Скачать PACKING LIST",
    "en": "📦 Download PACKING lists",
}


def packing_list_download_label(language: str | None) -> str:
    normalized = normalize_message_language(language) if language else "uz_latn"
    return PACKING_LIST_DOWNLOAD_BUTTON_LABELS.get(
        normalized,
        PACKING_LIST_DOWNLOAD_BUTTON_LABELS["uz_latn"],
    )

CANCEL_REPLY_MARKUP = {
    "keyboard": [[{"text": CANCEL_BUTTON}]],
    "resize_keyboard": True,
    "one_time_keyboard": True,
    "is_persistent": False,
}

REMOVE_REPLY_MARKUP = {
    "remove_keyboard": True,
    "selective": False,
}

UPLOAD_FOLDER = os.getenv("UPLOAD_FOLDER") or os.path.join(str(db.APP_DATA_DIR), "uploads")
WELCOME_VIDEO_PATH = os.path.join(os.path.dirname(__file__), "media", "welcome_guide.mp4")
WELCOME_VOICE_PATH = os.path.join(os.path.dirname(__file__), "media", "welcome_voice.ogg")
GOOGLE_SHEETS_URL_SETTING_KEY = "google_sheets_url"

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
db.init_db()


def _migrate_ombor_column_defaults() -> None:
    """One-time fix: plans created before the column-letter correction
    have the old defaults W/AA/AH stored. Real BURAQ sheet uses V/Z/AG.
    Update only rows that exactly match the old defaults (do not touch
    plans where user explicitly chose other columns)."""
    try:
        conn = db.get_conn()
        try:
            conn.execute(
                "UPDATE analytics_sales_plans "
                "SET ombor_cbm_col='V', ombor_date_col='Z', ombor_seller_col='AG' "
                "WHERE UPPER(COALESCE(ombor_cbm_col,''))='W' "
                "  AND UPPER(COALESCE(ombor_date_col,''))='AA' "
                "  AND UPPER(COALESCE(ombor_seller_col,''))='AH'"
            )
            conn.commit()
        finally:
            conn.close()
    except Exception as exc:
        app.logger.warning(f"_migrate_ombor_column_defaults skipped: {exc}")


_migrate_ombor_column_defaults()


def _normalize_sheet_cell(value: str | None) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("\u00a0", " ")
    text = re.sub(r"\s+", " ", text)
    return text


def _extract_sheet_date(value: str | None) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    dot_match = re.search(r"\b\d{2}\.\d{2}\.\d{4}\b", raw)
    if dot_match:
        return dot_match.group(0)
    iso_match = re.search(r"\b\d{4}-\d{2}-\d{2}\b", raw)
    if iso_match:
        year, month, day = iso_match.group(0).split("-")
        return f"{day}.{month}.{year}"
    return ""


def _sheet_float(value) -> float:
    raw = str(value or "").strip().replace("\u00a0", "").replace(" ", "").replace(",", ".")
    match = re.search(r"-?\d+(?:\.\d+)?", raw)
    if not match:
        return 0.0
    try:
        return float(match.group(0))
    except ValueError:
        return 0.0


def _sheet_int(value) -> int:
    return int(round(_sheet_float(value)))


def _google_sheets_export_url(sheet_url: str) -> str:
    url = (sheet_url or "").strip()
    if not url:
        raise ValueError("Ссылка на Google Sheets не указана")
    parsed = urlparse(url)
    if "docs.google.com" not in parsed.netloc:
        return url
    if "/export" in parsed.path and ("format=csv" in parsed.query or "output=csv" in parsed.query):
        return url
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", parsed.path)
    if not match:
        raise ValueError("Не удалось распознать ссылку Google Sheets")
    spreadsheet_id = match.group(1)
    query = parse_qs(parsed.query)
    gid = ""
    if query.get("gid"):
        gid = query["gid"][0]
    elif parsed.fragment:
        fragment_match = re.search(r"gid=(\d+)", parsed.fragment)
        if fragment_match:
            gid = fragment_match.group(1)
    export_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=csv"
    if gid:
        export_url += f"&gid={gid}"
    return export_url


def _fetch_google_sheet_matrix(sheet_url: str) -> list[list[str]]:
    export_url = _google_sheets_export_url(sheet_url)
    response = req.get(export_url, timeout=30)
    response.raise_for_status()
    text = response.content.decode("utf-8-sig", errors="replace")
    return list(csv.reader(io.StringIO(text)))


def _parse_google_sheet_rows(sheet_url: str) -> list[dict]:
    rows = _fetch_google_sheet_matrix(sheet_url)
    if not rows:
        return []

    parsed_rows: list[dict] = []
    aggregated_rows: dict[tuple[str, str], dict] = {}
    seen_ids: set[str] = set()

    def cell(row_index: int, col_index: int) -> str:
        if row_index < 0 or row_index >= len(rows):
            return ""
        row = rows[row_index]
        if col_index < 0 or col_index >= len(row):
            return ""
        return str(row[col_index] or "").strip()

    for header_row_index, row in enumerate(rows):
        for start_col_index, value in enumerate(row):
            if _normalize_sheet_cell(value) != "shipping mark":
                continue

            sheet_date = ""
            for up_index in range(header_row_index - 1, -1, -1):
                search_row = rows[up_index]
                left = max(0, start_col_index - 1)
                right = min(len(search_row), start_col_index + 6)
                for candidate in search_row[left:right]:
                    sheet_date = _extract_sheet_date(candidate)
                    if sheet_date:
                        break
                if sheet_date:
                    break

            blank_streak = 0
            data_row_index = header_row_index + 1
            while data_row_index < len(rows):
                code = cell(data_row_index, start_col_index)
                relevant = [cell(data_row_index, start_col_index + offset) for offset in range(4)]
                normalized_code = _normalize_sheet_cell(code)

                if normalized_code == "shipping mark":
                    break
                if normalized_code in {"total", "итого"}:
                    break

                if not any(relevant):
                    blank_streak += 1
                    if blank_streak >= 3:
                        break
                    data_row_index += 1
                    continue

                blank_streak = 0
                if code:
                    row_id = f"{header_row_index}:{data_row_index}:{start_col_index}"
                    if row_id not in seen_ids:
                        normalized_code = code.strip().upper()
                        quantity_places = _sheet_int(cell(data_row_index, start_col_index + 1))
                        volume_cbm = _sheet_float(cell(data_row_index, start_col_index + 2))
                        weight_kg = _sheet_float(cell(data_row_index, start_col_index + 3))
                        quantity_piece = str(quantity_places) if quantity_places else ""
                        if sheet_date:
                            aggregate_key = (sheet_date, normalized_code)
                            existing = aggregated_rows.get(aggregate_key)
                            if existing:
                                existing["quantity_places"] += quantity_places
                                existing["volume_cbm"] += volume_cbm
                                existing["weight_kg"] += weight_kg
                                if quantity_piece:
                                    existing["quantity_places_items"].append(quantity_piece)
                                existing["source_rows"].append(data_row_index + 1)
                                existing["merged_count"] += 1
                            else:
                                aggregated_rows[aggregate_key] = {
                                    "id": f"{sheet_date}:{normalized_code}",
                                    "sheet_date": sheet_date,
                                    "code": normalized_code,
                                    "quantity_places": quantity_places,
                                    "quantity_places_items": [quantity_piece] if quantity_piece else [],
                                    "volume_cbm": volume_cbm,
                                    "weight_kg": weight_kg,
                                    "source_row": data_row_index + 1,
                                    "source_rows": [data_row_index + 1],
                                    "merged_count": 1,
                                }
                        else:
                            parsed_rows.append(
                                {
                                    "id": row_id,
                                    "sheet_date": sheet_date,
                                    "code": normalized_code,
                                    "quantity_places": quantity_places,
                                    "quantity_places_items": [quantity_piece] if quantity_piece else [],
                                    "volume_cbm": volume_cbm,
                                    "weight_kg": weight_kg,
                                    "source_row": data_row_index + 1,
                                    "source_rows": [data_row_index + 1],
                                    "merged_count": 1,
                                }
                            )
                        seen_ids.add(row_id)
                data_row_index += 1

    parsed_rows.extend(aggregated_rows.values())
    for item in parsed_rows:
        item["quantity_places_display"] = " + ".join(
            [part for part in (item.get("quantity_places_items") or []) if str(part).strip()]
        )
    parsed_rows.sort(
        key=lambda item: (
            item.get("sheet_date") or "",
            item.get("code") or "",
            min(item.get("source_rows") or [item.get("source_row") or 0]),
        )
    )
    return parsed_rows


def normalize_message_language(language: str | None) -> str:
    language = (language or "").strip().lower()
    if language in TRACK_BUTTON_LABELS:
        return language
    return getattr(db, "DEFAULT_MESSAGE_LANGUAGE", "uz_latn")


def get_track_button_text(*, chat_id=None, language: str | None = None) -> str:
    normalized_language = normalize_message_language(language)
    if chat_id is not None and language is None:
        bl = db.find_latest_active_bl_by_chat(chat_id) or db.find_latest_bl_by_chat(chat_id)
        if bl:
            normalized_language = normalize_message_language(bl.get("message_language"))
    return TRACK_BUTTON_LABELS.get(normalized_language, TRACK_BUTTON)


def get_chat_message_language(chat_id) -> str:
    bl = db.find_latest_active_bl_by_chat(chat_id) or db.find_latest_bl_by_chat(chat_id)
    if bl:
        return normalize_message_language(bl.get("message_language"))
    return normalize_message_language(None)


def normalize_ai_language(language: str | None, chat_id=None) -> str:
    value = (language or "").strip().lower()
    if value == "uz_latin":
        return "uz_latn"
    if value == "uz_cyrillic":
        return "uz_cyrl"
    if value == "ru":
        return "ru"
    if value == "en" or value == "english":
        return "en"
    return get_chat_message_language(chat_id) if chat_id is not None else normalize_message_language(None)


def get_ai_ask_bl_text(language: str | None, chat_id=None) -> str:
    normalized = normalize_ai_language(language, chat_id=chat_id)
    return AI_ASK_BL_MESSAGES.get(normalized, AI_ASK_BL_MESSAGES["uz_latn"])


def get_ai_unknown_text(language: str | None, chat_id=None) -> str:
    normalized = normalize_ai_language(language, chat_id=chat_id)
    return AI_UNKNOWN_MESSAGES.get(normalized, AI_UNKNOWN_MESSAGES["uz_latn"])


def get_track_button_cooldown_text(language: str | None, seconds: int) -> str:
    normalized_language = normalize_message_language(language)
    template = TRACK_BUTTON_COOLDOWN_MESSAGES.get(
        normalized_language,
        TRACK_BUTTON_COOLDOWN_MESSAGES["uz_latn"],
    )
    return template.format(seconds=max(1, int(seconds)))


def _send_with_retry(send_func, *args, retries=3, delay=1.5, **kwargs):
    last_error = None
    for attempt in range(max(1, int(retries))):
        try:
            return send_func(*args, **kwargs)
        except Exception as exc:
            last_error = exc
            if attempt < retries - 1:
                time.sleep(delay)
    if last_error:
        raise last_error
    return None


def _extract_telegram_file_id(payload: dict | None, media_key: str) -> str:
    if not isinstance(payload, dict):
        return ""
    result = payload.get("result") or {}
    media = result.get(media_key) or {}
    if isinstance(media, dict):
        return str(media.get("file_id") or "").strip()
    return ""


def _get_cached_welcome_media_file_id(kind: str) -> str:
    with WELCOME_MEDIA_FILE_IDS_LOCK:
        return WELCOME_MEDIA_FILE_IDS.get(kind, "")


def _set_cached_welcome_media_file_id(kind: str, file_id: str) -> None:
    normalized = str(file_id or "").strip()
    if not normalized:
        return
    with WELCOME_MEDIA_FILE_IDS_LOCK:
        WELCOME_MEDIA_FILE_IDS[kind] = normalized


def telegram_send_video_by_file_id(chat_id, file_id: str):
    payload = {
        "chat_id": chat_id,
        "video": file_id,
        "supports_streaming": True,
    }
    return telegram_api("sendVideo", json=payload, timeout=60)


def telegram_send_voice_by_file_id(chat_id, file_id: str):
    payload = {
        "chat_id": chat_id,
        "voice": file_id,
    }
    return telegram_api("sendVoice", json=payload, timeout=60)


def _send_welcome_video(chat_id):
    cached_file_id = _get_cached_welcome_media_file_id("video")
    if cached_file_id:
        try:
            return _send_with_retry(
                telegram_send_video_by_file_id,
                chat_id,
                cached_file_id,
                retries=2,
                delay=1,
            )
        except Exception:
            with WELCOME_MEDIA_FILE_IDS_LOCK:
                if WELCOME_MEDIA_FILE_IDS.get("video") == cached_file_id:
                    WELCOME_MEDIA_FILE_IDS["video"] = ""

    payload = _send_with_retry(
        telegram_send_video,
        chat_id,
        WELCOME_VIDEO_PATH,
        "Buraq Logistics guide.mp4",
        retries=3,
        delay=2,
    )
    _set_cached_welcome_media_file_id("video", _extract_telegram_file_id(payload, "video"))
    return payload


def _send_welcome_voice(chat_id):
    cached_file_id = _get_cached_welcome_media_file_id("voice")
    if cached_file_id:
        try:
            return _send_with_retry(
                telegram_send_voice_by_file_id,
                chat_id,
                cached_file_id,
                retries=2,
                delay=1,
            )
        except Exception:
            with WELCOME_MEDIA_FILE_IDS_LOCK:
                if WELCOME_MEDIA_FILE_IDS.get("voice") == cached_file_id:
                    WELCOME_MEDIA_FILE_IDS["voice"] = ""

    payload = _send_with_retry(
        telegram_send_voice,
        chat_id,
        WELCOME_VOICE_PATH,
        "Buraq Logistics instruktsiya.ogg",
        retries=2,
        delay=1,
    )
    _set_cached_welcome_media_file_id("voice", _extract_telegram_file_id(payload, "voice"))
    return payload


def _send_welcome_media(chat_id):
    try:
        with WELCOME_MEDIA_SEMAPHORE:
            if os.path.exists(WELCOME_VIDEO_PATH):
                try:
                    _send_welcome_video(chat_id)
                except Exception:
                    pass
            if os.path.exists(WELCOME_VOICE_PATH):
                try:
                    _send_welcome_voice(chat_id)
                except Exception:
                    pass
    finally:
        with WELCOME_MEDIA_LOCK:
            WELCOME_MEDIA_PENDING.discard(str(chat_id))


def enqueue_welcome_media(chat_id):
    chat_key = str(chat_id)
    with WELCOME_MEDIA_LOCK:
        if chat_key in WELCOME_MEDIA_PENDING:
            return
        WELCOME_MEDIA_PENDING.add(chat_key)
    threading.Thread(
        target=_send_welcome_media,
        args=(chat_id,),
        name=f"welcome-media-{chat_key}",
        daemon=True,
    ).start()


def is_group_chat_id(chat_id) -> bool:
    try:
        return int(chat_id) < 0
    except Exception:
        return False


# The "Yuk holati" reply keyboard is RETIRED (owner request, 19.08.2026):
# the button is removed from every group and private chat. Both builders
# now return REMOVE_REPLY_MARKUP so every outgoing message actively strips
# any keyboard still cached on the client side. Typing the button text
# manually still works (the TRACK_BUTTON_TEXTS handler is kept).
# ИСКЛЮЧЕНИЕ: в группах, где включён Mini App трекинга (/formon), у поля
# ввода держится постоянная кнопка «📝 Treking forma» — тап по ней даёт
# кнопку открытия формы.
TRACKING_FORM_BUTTON_TEXT = "📝 Treking forma"


def _tgform_reply_keyboard() -> dict:
    return {
        "keyboard": [[{"text": TRACKING_FORM_BUTTON_TEXT}]],
        "resize_keyboard": True,
        "one_time_keyboard": False,
        "is_persistent": True,
    }


def build_main_reply_markup(*, chat_id=None, language: str | None = None) -> dict:
    return REMOVE_REPLY_MARKUP


def build_group_track_reply_markup(*, chat_id=None, language: str | None = None) -> dict:
    try:
        if chat_id is not None and str(chat_id) in _tgform_enabled_groups():
            return _tgform_reply_keyboard()
    except Exception:
        pass
    return REMOVE_REPLY_MARKUP


def get_group_welcome_text(button_text: str | None = None) -> str:
    # button_text kept for signature compatibility — the reply button is
    # retired, so the welcome no longer instructs clients to press it.
    return (
        "👋Assalomu alaykum hurmatli mijoz! \n\n"
        "🤖Ushbu bot yuklaringiz bo‘yicha ma’lumotlarni tez va qulay tarzda olish uchun yaratilgan.\n\n"
        "✅Ushbu platformada siz quyidagi imkoniyatlardan foydalanasiz:\n\n"
        "• yuk statusi bo‘yicha yangilanishlarni olasiz\n"
        "• yetkazib berish jarayonini nazorat qilasiz\n"
        "• menejer bilan bog‘lanasiz\n\n"
        "🎥 Botdan foydalanish bo‘yicha qisqa videoqo‘llanma quyida taqdim etilgan.\n\n"
        "Bir marta ko‘rib chiqish tavsiya etiladi 👇\n\n"
        "━━━━━━━━━━━━━━━\n\n"
        "👋Здравствуйте, уважаемый клиент!\n\n"
        "🤖Этот бот создан для того, чтобы вы могли быстро и удобно получать информацию по вашим грузам.\n\n"
        "✅На этой платформе вы сможете:\n\n"
        "• получать обновления по статусу груза\n"
        "• контролировать процесс доставки\n"
        "• связываться с менеджером\n\n"
        "🎥 Ниже представлена короткая видеоинструкция по использованию бота.\n\n"
        "Рекомендуем посмотреть её один раз 👇"
    )


def get_no_active_cargo_text(language: str | None = None) -> str:
    normalized_language = normalize_message_language(language)
    return NO_ACTIVE_CARGO_MESSAGES.get(normalized_language, NO_ACTIVE_CARGO_MESSAGES["uz_latn"])


def get_menu_restore_text(language: str | None = None) -> str:
    # The reply button is retired — /menu now just reports that updates
    # arrive automatically instead of pretending to restore a keyboard.
    normalized_language = normalize_message_language(language)
    if normalized_language == "uz_cyrl":
        return "ℹ️ Юк ҳолати бўйича янгиланишлар автоматик юборилади."
    if normalized_language == "ru":
        return "ℹ️ Обновления по статусу груза приходят автоматически."
    if normalized_language == "en":
        return "ℹ️ Cargo status updates are delivered automatically."
    return "ℹ️ Yuk holati bo'yicha yangilanishlar avtomatik yuboriladi."


def get_chat_admin_ids(chat_id):
    cache_key = str(chat_id)
    cached = CHAT_ADMIN_CACHE.get(cache_key)
    now = time.time()
    if cached and now - cached.get("loaded_at", 0) < CHAT_ADMIN_CACHE_TTL:
        return cached.get("admin_ids", set())

    admin_ids = set()
    try:
        payload = telegram_api("getChatAdministrators", json={"chat_id": chat_id})
        for member in (payload or {}).get("result", []):
            user = member.get("user") or {}
            user_id = user.get("id")
            if user_id is not None:
                admin_id = str(user_id)
                admin_ids.add(admin_id)
                db.remember_chat_member(
                    chat_id,
                    admin_id,
                    telegram_user_name(user),
                    user.get("username") or "",
                    is_admin=True,
                )
    except Exception:
        if cached:
            return cached.get("admin_ids", set())
    CHAT_ADMIN_CACHE[cache_key] = {"loaded_at": now, "admin_ids": admin_ids}
    return admin_ids


def extract_telegram_message_text(message: dict) -> str:
    text = (message.get("text") or message.get("caption") or "").strip()
    if text:
        return text
    if message.get("photo"):
        return "[photo]"
    if message.get("video"):
        return "[video]"
    if message.get("voice"):
        return "[voice]"
    if message.get("audio"):
        return "[audio]"
    if message.get("sticker"):
        return "[sticker]"
    document = message.get("document") or {}
    if document:
        filename = (document.get("file_name") or "").strip()
        return f"[document] {filename}".strip()
    return ""


def telegram_user_name(user: dict) -> str:
    if not user:
        return ""
    full_name = " ".join(
        part for part in [user.get("first_name") or "", user.get("last_name") or ""] if part
    ).strip()
    return full_name or ""


def telegram_unix_to_local(value) -> str:
    try:
        return datetime.fromtimestamp(int(value), db.TASHKENT_TZ).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return db.current_ts()


def get_responsible_response_role(assignments: dict | None, sender_id: str, admin_ids: set[str] | None = None) -> str:
    sender_value = str(sender_id or "").strip()
    if not sender_value:
        return ""
    moderator_id = str((assignments or {}).get("moderator_tg_id") or "").strip()
    sales_manager_id = str((assignments or {}).get("sales_manager_tg_id") or "").strip()
    if moderator_id and sender_value == moderator_id:
        return "moderator"
    if sales_manager_id and sender_value == sales_manager_id:
        return "sales_manager"
    if not moderator_id and not sales_manager_id and admin_ids and sender_value in admin_ids:
        return "moderator"
    return ""


def track_moderator_response_metrics(message: dict):
    chat = message.get("chat") or {}
    chat_id = chat.get("id")
    chat_type = chat.get("type")
    if chat_type not in {"group", "supergroup"} or not chat_id:
        return

    sender = message.get("from") or {}
    sender_id = sender.get("id")
    if sender.get("is_bot") or sender_id is None:
        return

    text = extract_telegram_message_text(message)
    if not text:
        return

    admin_ids = get_chat_admin_ids(chat_id)
    sender_id_str = str(sender_id)
    is_admin_sender = sender_id_str in admin_ids if admin_ids else False
    db.remember_chat_member(
        chat_id,
        sender_id_str,
        telegram_user_name(sender),
        sender.get("username") or "",
        is_admin=is_admin_sender,
    )
    reply_to = message.get("reply_to_message") or {}
    reply_sender = reply_to.get("from") or {}
    reply_sender_id = reply_sender.get("id")
    reply_sender_id_str = str(reply_sender_id) if reply_sender_id is not None else ""
    if reply_sender_id_str and not reply_sender.get("is_bot"):
        db.remember_chat_member(
            chat_id,
            reply_sender_id_str,
            telegram_user_name(reply_sender),
            reply_sender.get("username") or "",
            is_admin=reply_sender_id_str in admin_ids if admin_ids else False,
        )

    linked_bl = db.find_latest_active_bl_by_chat(chat_id) or db.find_latest_bl_by_chat(chat_id)
    chat_assignments = db.get_chat_response_assignments(chat_id) or {}
    linked_context = {
        "bl_id": linked_bl.get("id") if linked_bl else None,
        "batch_id": linked_bl.get("batch_id") if linked_bl else None,
        "batch_name": linked_bl.get("batch_name") if linked_bl else "",
        "assigned_moderator_id": (chat_assignments.get("moderator_tg_id") or (linked_bl.get("moderator_tg_id") if linked_bl else "") or ""),
        "assigned_sales_manager_id": (chat_assignments.get("sales_manager_tg_id") or (linked_bl.get("sales_manager_tg_id") if linked_bl else "") or ""),
    }

    normalized_text = text.strip()
    if normalized_text in TRACK_BUTTON_TEXTS or normalized_text == CANCEL_BUTTON:
        return
    if normalized_text.startswith("/start") or normalized_text.startswith("/chatid"):
        return
    if re.match(r"^/([A-Za-z0-9_]+)(?:@\w+)?$", normalized_text):
        return

    if reply_to and reply_sender_id_str and reply_sender_id_str != sender_id_str and not reply_sender.get("is_bot"):
        db.record_moderator_request(
            chat_id=chat_id,
            chat_title=chat.get("title") or "",
            request_message_id=reply_to.get("message_id"),
            request_user_id=reply_sender_id_str,
            request_user_name=telegram_user_name(reply_sender),
            request_username=reply_sender.get("username") or "",
            request_text=extract_telegram_message_text(reply_to),
            **linked_context,
            requested_at=telegram_unix_to_local(reply_to.get("date")),
        )

        response_role = get_responsible_response_role(
            {
                "moderator_tg_id": linked_context["assigned_moderator_id"],
                "sales_manager_tg_id": linked_context["assigned_sales_manager_id"],
            },
            sender_id_str,
            admin_ids,
        )
        if response_role:
            db.mark_moderator_response(
                chat_id=chat_id,
                request_message_id=reply_to.get("message_id"),
                responder_user_id=sender_id_str,
                responder_name=telegram_user_name(sender),
                responder_username=sender.get("username") or "",
                response_text=normalized_text,
                responded_at=telegram_unix_to_local(message.get("date")),
                response_role=response_role,
            )
        return

    if is_admin_sender or get_responsible_response_role(
        {
            "moderator_tg_id": linked_context["assigned_moderator_id"],
            "sales_manager_tg_id": linked_context["assigned_sales_manager_id"],
        },
        sender_id_str,
        admin_ids,
    ):
        return

    db.record_moderator_request(
        chat_id=chat_id,
        chat_title=chat.get("title") or "",
        request_message_id=message.get("message_id"),
        request_user_id=sender_id_str,
        request_user_name=telegram_user_name(sender),
        request_username=sender.get("username") or "",
        request_text=normalized_text,
        **linked_context,
        requested_at=telegram_unix_to_local(message.get("date")),
    )


def login_required(func):
    @wraps(func)
    def decorated(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login"))
        return func(*args, **kwargs)

    return decorated


def editor_required(func):
    @wraps(func)
    def decorated(*args, **kwargs):
        if not session.get("logged_in"):
            if request.path.startswith("/api/"):
                return jsonify({"error": "Login required"}), 401
            return redirect(url_for("login"))
        if session.get("role") != ROLE_EDITOR:
            if request.path.startswith("/api/"):
                return jsonify({"error": "View-only access"}), 403
            return redirect(url_for("index"))
        return func(*args, **kwargs)

    return decorated


def get_auth_users():
    guest_logins = [
        os.getenv("GUEST1_LOGIN", "Guest1"),
        os.getenv("GUEST2_LOGIN", "Guest2"),
        os.getenv("GUEST3_LOGIN", "Guest3"),
    ]
    users = {}

    def add_user(username: str, password: str, role: str):
        username = (username or "").strip()
        password = (password or "").strip()
        if username and password:
            users[username] = {"password": password, "role": role}

    add_user(ADMIN_LOGIN, ADMIN_PASSWORD, ROLE_EDITOR)
    add_user(ADMIN1_LOGIN, ADMIN1_PASSWORD, ROLE_EDITOR)
    for guest_login in guest_logins:
        add_user(guest_login, GUEST_PASSWORD, ROLE_VIEWER)
    # TV kiosk user — sees only the Sales Monitor full-screen view
    add_user(KIOSK_LOGIN, KIOSK_PASSWORD, ROLE_KIOSK)

    return users


def get_role_label(role: str) -> str:
    return "Editor" if role == ROLE_EDITOR else "View only"


def get_request_ip() -> str:
    forwarded = (request.headers.get("X-Forwarded-For") or "").split(",")[0].strip()
    return (
        forwarded
        or (request.headers.get("X-Real-IP") or "").strip()
        or (request.remote_addr or "").strip()
    )


@app.context_processor
def inject_auth_context():
    role = session.get("role", ROLE_VIEWER)
    return {
        "is_editor": role == ROLE_EDITOR,
        "current_role": role,
        "current_role_label": get_role_label(role) if session.get("logged_in") else "",
    }


# СТРОГО КОНФИДЕНЦИАЛЬНЫЕ чаты (Savdo bo'limi / Buraq): бот НИКОГДА не
# отправляет туда сообщения — гард стоит на самом нижнем уровне API,
# поэтому покрывает трекинг, объявления, AI, sweep и всё остальное.
CONFIDENTIAL_CHAT_IDS = {
    part.strip()
    for part in os.getenv("CONFIDENTIAL_CHAT_IDS", "-1002687342009").split(",")
    if part.strip()
}


def telegram_api(method: str, *, timeout: int = 15, **kwargs):
    if not BOT_TOKEN:
        raise RuntimeError("BOT_TOKEN is not configured")
    target_chat = None
    for container in (kwargs.get("json"), kwargs.get("data")):
        if isinstance(container, dict) and container.get("chat_id") is not None:
            target_chat = str(container.get("chat_id"))
            break
    if target_chat is not None and target_chat in CONFIDENTIAL_CHAT_IDS and method.lower().startswith("send"):
        app.logger.warning("BLOCKED %s to confidential chat %s", method, target_chat)
        return {"ok": False, "blocked_confidential": True}
    response = req.post(
        f"https://api.telegram.org/bot{BOT_TOKEN}/{method}",
        timeout=timeout,
        **kwargs,
    )
    try:
        payload = response.json()
    except ValueError:
        payload = {}
    if not response.ok or payload.get("ok") is False:
        description = payload.get("description") or response.text or f"HTTP {response.status_code}"
        raise RuntimeError(f"Telegram {method}: {description}")
    return payload


def telegram_send_message(
    chat_id,
    text: str,
    reply_markup: dict | None = None,
    parse_mode: str | None = "HTML",
    disable_notification: bool = False,
    link_preview_options: dict | None = None,
):
    payload = {
        "chat_id": chat_id,
        "text": text,
    }
    if disable_notification:
        payload["disable_notification"] = True
    if parse_mode:
        payload["parse_mode"] = parse_mode
    if reply_markup:
        payload["reply_markup"] = reply_markup
    if link_preview_options:
        payload["link_preview_options"] = link_preview_options
    return telegram_api("sendMessage", json=payload)


def telegram_send_chat_action(chat_id, action: str = "typing"):
    """Show "typing…" in the chat header. Telegram clears it after ~5s
    or as soon as a message arrives, so long operations use
    TypingIndicator instead of a single call."""
    try:
        telegram_api("sendChatAction", json={"chat_id": chat_id, "action": action})
    except Exception:
        pass


class TypingIndicator:
    """Context manager: keeps the "typing…" status visible while a slow
    operation (DeepSeek call, file delivery) is producing the reply."""

    def __init__(self, chat_id, action: str = "typing"):
        self.chat_id = chat_id
        self.action = action
        self._stop = threading.Event()

    def __enter__(self):
        def _loop():
            while not self._stop.is_set():
                telegram_send_chat_action(self.chat_id, self.action)
                self._stop.wait(4.0)

        threading.Thread(target=_loop, daemon=True).start()
        return self

    def __exit__(self, *exc):
        self._stop.set()
        return False


# Network errors we treat as "retry-worthy" — usually a transient
# Railway↔Telegram connectivity blip. requests raises slightly different
# exception classes depending on what stage failed, so we catch them all.
_RETRIABLE_NETWORK_ERRORS = (
    req.exceptions.Timeout,
    req.exceptions.ConnectionError,
    req.exceptions.ChunkedEncodingError,
)


def _parse_tg_retry_after(response) -> float:
    """Extract Telegram's retry_after (seconds) from a 429 response.

    Telegram returns:
      {"ok":false,"error_code":429,"description":"Too Many Requests: retry after N",
       "parameters":{"retry_after":N}}
    """
    try:
        payload = response.json()
    except Exception:
        return 0.0
    params = payload.get("parameters") or {}
    retry_after = params.get("retry_after")
    if isinstance(retry_after, (int, float)) and retry_after > 0:
        return float(retry_after)
    # Fallback — parse "retry after N" from description text
    desc = str(payload.get("description") or "")
    match = re.search(r"retry after\s+(\d+)", desc, flags=re.IGNORECASE)
    if match:
        try:
            return float(match.group(1))
        except ValueError:
            return 0.0
    return 0.0


def telegram_send_document(
    chat_id,
    file_path: str,
    filename: str,
    *,
    caption: str | None = None,
    parse_mode: str | None = None,
):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Файл не найден: {file_path}")

    safe_filename = filename or os.path.basename(file_path)
    mime_type = mimetypes.guess_type(safe_filename)[0] or "application/octet-stream"
    data = {
        "chat_id": chat_id,
        "disable_content_type_detection": "true",
    }
    if caption:
        data["caption"] = caption
    if parse_mode:
        data["parse_mode"] = parse_mode

    last_exc: Exception | None = None
    # Up to 3 attempts. Catches:
    #   - Transient TimeoutError / ConnectionError from Railway → Telegram
    #     (growing backoff: 1.5s, 3s).
    #   - HTTP 429 Too Many Requests — read retry_after from the Telegram
    #     response and sleep exactly that long before retrying.
    # The file handle is reopened on each attempt so the upload restarts
    # cleanly from byte 0.
    for attempt in range(3):
        try:
            with open(file_path, "rb") as file_handle:
                response = req.post(
                    f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument",
                    data=data,
                    files={"document": (safe_filename, file_handle, mime_type)},
                    timeout=120,
                )
            if response.status_code == 429:
                retry_after = _parse_tg_retry_after(response) or 5.0
                app.logger.warning(
                    "telegram_send_document hit 429 — sleeping %.1fs before retry (attempt %s/3)",
                    retry_after, attempt + 1,
                )
                if attempt < 2:
                    time.sleep(min(retry_after + 0.5, 60.0))
                    continue
                # Final attempt — surface the description
                try:
                    payload = response.json()
                    description = payload.get("description") or response.text
                except ValueError:
                    description = response.text
                raise RuntimeError(description)
            if not response.ok:
                try:
                    payload = response.json()
                    description = payload.get("description") or response.text
                except ValueError:
                    description = response.text
                raise RuntimeError(description)
            return response.json()
        except _RETRIABLE_NETWORK_ERRORS as exc:
            last_exc = exc
            app.logger.warning(
                "telegram_send_document network error (attempt %s/3): %s",
                attempt + 1, exc,
            )
            if attempt < 2:
                time.sleep(1.5 * (attempt + 1))
                continue
            raise
    if last_exc:
        raise last_exc


def telegram_send_document_by_file_id(
    chat_id,
    tg_file_id: str,
    *,
    caption: str | None = None,
    parse_mode: str | None = None,
):
    """Resend a previously-uploaded document by Telegram's cached file_id.

    No multipart upload — Telegram serves the bytes from its own CDN, so
    the round-trip is typically <500ms regardless of file size.
    """
    data = {
        "chat_id": chat_id,
        "document": tg_file_id,
        "disable_content_type_detection": "true",
    }
    if caption:
        data["caption"] = caption
    if parse_mode:
        data["parse_mode"] = parse_mode

    last_exc: Exception | None = None
    # CDN resend is fast (~500ms). Retry transient network errors and
    # respect Telegram's 429 retry_after.
    for attempt in range(3):
        try:
            response = req.post(
                f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument",
                data=data,
                timeout=15,
            )
            if response.status_code == 429:
                retry_after = _parse_tg_retry_after(response) or 3.0
                app.logger.warning(
                    "telegram_send_document_by_file_id hit 429 — sleeping %.1fs (attempt %s/3)",
                    retry_after, attempt + 1,
                )
                if attempt < 2:
                    time.sleep(min(retry_after + 0.5, 60.0))
                    continue
                try:
                    payload = response.json()
                    description = payload.get("description") or response.text
                except ValueError:
                    description = response.text
                raise RuntimeError(description)
            if not response.ok:
                try:
                    payload = response.json()
                    description = payload.get("description") or response.text
                except ValueError:
                    description = response.text
                raise RuntimeError(description)
            return response.json()
        except _RETRIABLE_NETWORK_ERRORS as exc:
            last_exc = exc
            app.logger.warning(
                "telegram_send_document_by_file_id network error (attempt %s/3): %s",
                attempt + 1, exc,
            )
            if attempt < 2:
                time.sleep(1.0 * (attempt + 1))
                continue
            raise
    if last_exc:
        raise last_exc


def _extract_tg_file_id_from_send_response(payload) -> str:
    """Pull the document file_id out of a sendDocument response."""
    if not isinstance(payload, dict):
        return ""
    result = payload.get("result") or {}
    document = result.get("document") or {}
    fid = document.get("file_id") or ""
    return str(fid).strip()


def _extract_message_id(payload) -> int:
    """Pull the message_id out of any Telegram send* response payload."""
    if not isinstance(payload, dict):
        return 0
    result = payload.get("result") or {}
    try:
        return int(result.get("message_id") or 0)
    except (TypeError, ValueError):
        return 0


def telegram_send_photo(
    chat_id,
    file_path: str,
    filename: str | None = None,
    *,
    caption: str | None = None,
    parse_mode: str | None = None,
):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Изображение не найдено: {file_path}")

    safe_filename = filename or os.path.basename(file_path)
    mime_type = mimetypes.guess_type(safe_filename)[0] or "image/jpeg"
    data = {
        "chat_id": chat_id,
    }
    if caption:
        data["caption"] = caption
    if parse_mode:
        data["parse_mode"] = parse_mode
    with open(file_path, "rb") as file_handle:
        response = req.post(
            f"https://api.telegram.org/bot{BOT_TOKEN}/sendPhoto",
            data=data,
            files={"photo": (safe_filename, file_handle, mime_type)},
            timeout=30,
        )
    if not response.ok:
        try:
            payload = response.json()
            description = payload.get("description") or response.text
        except ValueError:
            description = response.text
        raise RuntimeError(description)
    return response.json()


def telegram_send_video(
    chat_id,
    file_path: str,
    filename: str | None = None,
    *,
    caption: str | None = None,
    parse_mode: str | None = None,
):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Видео не найдено: {file_path}")

    safe_filename = filename or os.path.basename(file_path)
    data = {
        "chat_id": chat_id,
        "supports_streaming": "true",
    }
    if caption:
        data["caption"] = caption
    if parse_mode:
        data["parse_mode"] = parse_mode
    with open(file_path, "rb") as file_handle:
        response = req.post(
            f"https://api.telegram.org/bot{BOT_TOKEN}/sendVideo",
            data=data,
            files={"video": (safe_filename, file_handle, "video/mp4")},
            timeout=60,
        )
    if not response.ok:
        try:
            payload = response.json()
            description = payload.get("description") or response.text
        except ValueError:
            description = response.text
        raise RuntimeError(description)
    return response.json()


def telegram_send_animation(
    chat_id,
    file_path: str,
    filename: str | None = None,
    *,
    caption: str | None = None,
    parse_mode: str | None = None,
):
    """GIF (or soundless MP4) via sendAnimation — shows as a looping GIF."""
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"GIF не найден: {file_path}")

    safe_filename = filename or os.path.basename(file_path)
    mime_type = mimetypes.guess_type(safe_filename)[0] or "image/gif"
    data = {"chat_id": chat_id}
    if caption:
        data["caption"] = caption
    if parse_mode:
        data["parse_mode"] = parse_mode
    with open(file_path, "rb") as file_handle:
        response = req.post(
            f"https://api.telegram.org/bot{BOT_TOKEN}/sendAnimation",
            data=data,
            files={"animation": (safe_filename, file_handle, mime_type)},
            timeout=60,
        )
    if not response.ok:
        try:
            payload = response.json()
            description = payload.get("description") or response.text
        except ValueError:
            description = response.text
        raise RuntimeError(description)
    return response.json()


def telegram_send_voice(chat_id, file_path: str, filename: str | None = None):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Озвучка не найдена: {file_path}")

    safe_filename = filename or os.path.basename(file_path)
    with open(file_path, "rb") as file_handle:
        response = req.post(
            f"https://api.telegram.org/bot{BOT_TOKEN}/sendVoice",
            data={
                "chat_id": chat_id,
            },
            files={"voice": (safe_filename, file_handle, "audio/ogg")},
            timeout=60,
        )
    if not response.ok:
        try:
            payload = response.json()
            description = payload.get("description") or response.text
        except ValueError:
            description = response.text
        raise RuntimeError(description)
    return response.json()


def telegram_answer_callback_query(callback_query_id, text: str):
    return telegram_api(
        "answerCallbackQuery",
        json={
            "callback_query_id": callback_query_id,
            "text": text,
            "show_alert": False,
        },
    )


def telegram_delete_message(chat_id, message_id):
    return telegram_api(
        "deleteMessage",
        json={
            "chat_id": chat_id,
            "message_id": message_id,
        },
    )


def telegram_leave_chat(chat_id):
    return telegram_api(
        "leaveChat",
        json={
            "chat_id": chat_id,
        },
    )


def communication_rating_markup(dispatch_id: int):
    options = list(range(1, 11))
    return {
        "inline_keyboard": [
            [
                {
                    "text": str(score),
                    "callback_data": f"{COMM_RATE_PREFIX}:{dispatch_id}:{score}",
                }
                for score in options[:5]
            ],
            [
                {
                    "text": str(score),
                    "callback_data": f"{COMM_RATE_PREFIX}:{dispatch_id}:{score}",
                }
                for score in options[5:]
            ],
        ]
    }


def bl_file_markup(bl_id: int, language: str | None = None):
    """One button that delivers every attached packing list in a single tap.

    Previously we rendered one button per file, which got noisy for BLs
    with many packing lists and forced the client to tap each one. Now
    we expose a single localized button ("Packing listlarni yuklab
    olish") whose callback triggers bulk delivery of all attached files.
    """
    files = db.get_files(bl_id) or []
    has_any = any((file_info.get("public_token") or "").strip() for file_info in files)
    if not has_any:
        return None
    label = packing_list_download_label(language)
    return {
        "inline_keyboard": [[{
            "text": label,
            "callback_data": f"{FILE_ALL_PREFIX}:{bl_id}",
        }]],
    }


def clear_group_reply_keyboard(chat_id):
    try:
        telegram_send_message(chat_id, "ㅤ", reply_markup=REMOVE_REPLY_MARKUP)
    except Exception:
        pass


def _delete_message_later(chat_id, message_id, delay: float = 1.5):
    try:
        time.sleep(max(0.2, float(delay)))
        telegram_delete_message(chat_id, message_id)
    except Exception:
        pass
    finally:
        chat_key = str(chat_id)
        try:
            with TRACK_KEYBOARD_ANCHORS_LOCK:
                current = TRACK_KEYBOARD_ANCHORS.get(chat_key) or {}
                if current.get("message_id") == message_id:
                    TRACK_KEYBOARD_ANCHORS.pop(chat_key, None)
        except Exception:
            pass


def extract_bot_command(text: str) -> str:
    match = re.match(r"^/([A-Za-z0-9_]+)(?:@\w+)?(?:\s+.*)?$", (text or "").strip(), re.S)
    return (match.group(1) or "").lower() if match else ""


def handle_group_remove_request(message: dict, command: str):
    chat = message.get("chat") or {}
    chat_id = chat.get("id")
    chat_type = chat.get("type")
    sender = message.get("from") or {}
    sender_id = sender.get("id")
    if chat_type not in {"group", "supergroup"} or not chat_id:
        return False
    if command not in GROUP_REMOVE_COMMANDS:
        return False

    admin_ids = get_chat_admin_ids(chat_id)
    if str(sender_id or "") not in admin_ids:
        telegram_send_message(
            chat_id,
            "❌ Bu buyruqdan faqat guruh admini foydalanishi mumkin.",
        )
        return True

    telegram_send_message(
        chat_id,
        "✅ Klaviatura yopilmoqda. Bot guruhni tark etadi...",
        reply_markup=REMOVE_REPLY_MARKUP,
    )
    time.sleep(1)
    try:
        telegram_leave_chat(chat_id)
    except Exception:
        pass
    return True


def refresh_track_reply_keyboard(chat_id, *, language: str | None = None):
    # Retired with the "Yuk holati" button: the keyboard anchor messages
    # ("⬇️ Yuk holati") are no longer sent. Keyboard removal rides along
    # every regular outgoing message via REMOVE_REPLY_MARKUP instead.
    return


def send_group_message_with_keyboard(chat_id, text: str, *, language: str | None = None):
    if is_group_chat_id(chat_id):
        telegram_send_message(
            chat_id,
            text,
            reply_markup=build_group_track_reply_markup(chat_id=chat_id, language=language),
        )
        return
    telegram_send_message(chat_id, text)
    refresh_track_reply_keyboard(chat_id, language=language)


def send_group_welcome_bundle(chat_id, button_text: str | None = None):
    send_group_message_with_keyboard(chat_id, get_group_welcome_text(button_text))
    enqueue_welcome_media(chat_id)


def send_with_track_keyboard(
    chat_id,
    text: str,
    *,
    language: str | None = None,
    reply_markup: dict | None = None,
    parse_mode: str | None = "HTML",
):
    """Returns the raw Telegram API response (or None on failure inside).

    Callers that want to remember message_id for later recall can read
    response["result"]["message_id"].
    """
    if is_group_chat_id(chat_id):
        # In groups, suppress the "Yuk holati" reply keyboard (those only
        # belong to private chats), but keep inline keyboards (e.g. file
        # buttons under the tracking message) — those work in groups and
        # are how clients pull packing-list files into the group chat.
        inline_only = None
        if isinstance(reply_markup, dict) and reply_markup.get("inline_keyboard"):
            inline_only = {"inline_keyboard": reply_markup["inline_keyboard"]}
        return telegram_send_message(
            chat_id,
            text,
            reply_markup=inline_only,
            parse_mode=parse_mode,
        )
    return telegram_send_message(
        chat_id,
        text,
        reply_markup=reply_markup or build_main_reply_markup(chat_id=chat_id, language=language),
        parse_mode=parse_mode,
    )


def _plain_text_message(text: str) -> str:
    raw = str(text or "")
    cleaned = re.sub(r"<br\s*/?>", "\n", raw, flags=re.IGNORECASE)
    cleaned = re.sub(r"</?[^>]+>", "", cleaned)
    return html.unescape(cleaned).strip()


def communication_rating_label(score: int) -> str:
    return f"{int(score)}/10"


def send_communication_survey(recipient: dict, month_key: str):
    dispatch_id = db.record_communication_survey_send(month_key, recipient)
    text = db.render_communication_rate_message(recipient, month_key)
    try:
        response = telegram_send_message(
            recipient["chat_id"],
            text,
            reply_markup=communication_rating_markup(dispatch_id),
        )
        message_id = (((response or {}).get("result") or {}).get("message_id"))
        if message_id is not None:
            db.save_communication_survey_dispatch_message_id(dispatch_id, message_id)
    except Exception:
        db.delete_communication_survey_dispatch(dispatch_id)
        raise


# Telegram counts caption/text limits in UTF-16 code units (emoji = 2).
def _tg_text_len(s: str) -> int:
    return len((s or "").encode("utf-16-le")) // 2


TELEGRAM_MAX_CAPTION_UNITS = 1024
TELEGRAM_MAX_TEXT_UNITS = 4096


def _announcement_media_sig(rel_path: str) -> str:
    return hmac.new(
        app.secret_key.encode() if isinstance(app.secret_key, str) else app.secret_key,
        f"announce-media:{rel_path}".encode(),
        "sha256",
    ).hexdigest()[:20]


def announcement_media_url(file_path: str) -> str:
    """Absolute public URL for an announcement attachment (or '' when the
    public base URL isn't configured / the file is outside UPLOAD_FOLDER).

    Stateless: the URL embeds the uploads-relative path plus an HMAC over
    it, so scheduled-broadcast snapshots keep working with no DB lookups.
    Used to attach the photo as a link preview when the announcement text
    exceeds Telegram's 1024-unit caption limit."""
    base = WEBHOOK_BASE_URL.rstrip("/")
    if not base or not file_path:
        return ""
    try:
        rel = os.path.relpath(os.path.abspath(file_path), os.path.abspath(UPLOAD_FOLDER))
    except ValueError:
        return ""
    rel = rel.replace("\\", "/")
    if rel.startswith(".."):
        return ""
    token = base64.urlsafe_b64encode(rel.encode()).decode().rstrip("=")
    return f"{base}/announcement-media/{_announcement_media_sig(rel)}/{token}"


@app.route("/announcement-media/<sig>/<token>")
def serve_announcement_media(sig: str, token: str):
    """Public (signed) endpoint Telegram fetches the preview photo from."""
    try:
        rel = base64.urlsafe_b64decode(token + "=" * (-len(token) % 4)).decode()
    except Exception:
        abort(404)
    if not hmac.compare_digest(sig, _announcement_media_sig(rel)):
        abort(404)
    uploads_root = os.path.abspath(UPLOAD_FOLDER)
    full = os.path.abspath(os.path.join(uploads_root, rel))
    if not full.startswith(uploads_root + os.sep) or not os.path.exists(full):
        abort(404)
    return send_file(full)


def _send_announcement_media(chat_id, attachment_info: dict, caption: str | None = None):
    """Send the announcement attachment with the right Telegram method.

    photo → sendPhoto (falls back to sendDocument when Telegram refuses
    the image), animation (GIF) → sendAnimation, video → sendVideo —
    both fall back to sendDocument if Telegram rejects the media form.
    Everything else goes as a document.
    """
    kind = (attachment_info.get("kind") or "document").strip().lower()
    file_path = (attachment_info.get("file_path") or "").strip()
    filename = (attachment_info.get("filename") or "").strip()

    if kind == "photo":
        send_as_photo = True
        try:
            if os.path.getsize(file_path) > TELEGRAM_MAX_PHOTO_BYTES:
                send_as_photo = False
        except OSError:
            pass
        if send_as_photo:
            try:
                telegram_send_photo(chat_id, file_path, filename or None, caption=caption)
                return
            except Exception as exc:
                if "too big for a photo" not in str(exc).lower():
                    raise
        telegram_send_document(
            chat_id, file_path, filename or os.path.basename(file_path), caption=caption
        )
        return

    if kind in ("animation", "video"):
        sender = telegram_send_animation if kind == "animation" else telegram_send_video
        try:
            sender(chat_id, file_path, filename or None, caption=caption)
            return
        except RuntimeError:
            # Telegram couldn't take it as GIF/video (codec, size…) —
            # deliver as a plain document so the broadcast still lands.
            telegram_send_document(
                chat_id, file_path, filename or os.path.basename(file_path), caption=caption
            )
            return

    telegram_send_document(
        chat_id, file_path, filename or os.path.basename(file_path), caption=caption
    )


def send_announcement_broadcast(
    chat_id,
    text: str,
    attachment: dict | None = None,
    delivery_mode: str = "together",
    media_order: str = "media_first",
):
    """Deliver one announcement to one chat.

    delivery_mode:
      "together" — media + text as ONE message when Telegram allows it
                   (caption ≤1024 units; for photos a longer text goes out
                   as a single message with the photo as a large preview).
                   Falls back to two messages in media_order otherwise.
      "separate" — always two messages.
    media_order: "media_first" | "text_first" — order of the two-message
      form (and of the fallback when "together" doesn't fit).
    """
    attachment_info = attachment or {}
    file_path = (attachment_info.get("file_path") or "").strip()
    caption_text = (text or "").strip()
    mode = "separate" if str(delivery_mode).strip().lower() == "separate" else "together"
    order = "text_first" if str(media_order).strip().lower() == "text_first" else "media_first"

    if not file_path:
        if caption_text:
            telegram_send_message(chat_id, text, parse_mode=None)
        return
    if not caption_text:
        _send_announcement_media(chat_id, attachment_info)
        return

    can_use_single_caption = _tg_text_len(caption_text) <= TELEGRAM_MAX_CAPTION_UNITS

    if mode == "together":
        if can_use_single_caption:
            _send_announcement_media(chat_id, attachment_info, caption=caption_text)
            return
        # Photo + text LONGER than the caption limit: still deliver as ONE
        # message — the text goes out via sendMessage (4096-unit limit)
        # with the photo attached as a large link preview shown ABOVE the
        # text. Needs a public URL (WEBHOOK_BASE_URL); GIF/video previews
        # don't render reliably, so those fall back to two messages.
        if (
            attachment_info.get("kind") == "photo"
            and _tg_text_len(caption_text) <= TELEGRAM_MAX_TEXT_UNITS
        ):
            media_url = announcement_media_url(file_path)
            if media_url:
                telegram_send_message(
                    chat_id,
                    caption_text,
                    parse_mode=None,
                    link_preview_options={
                        "url": media_url,
                        "prefer_large_media": True,
                        "show_above_text": True,
                    },
                )
                return
        # "Together" impossible for this combination — degrade to the
        # two-message form below, honoring the configured order.

    if order == "text_first":
        telegram_send_message(chat_id, text, parse_mode=None)
        _send_announcement_media(chat_id, attachment_info)
    else:
        _send_announcement_media(chat_id, attachment_info)
        telegram_send_message(chat_id, text, parse_mode=None)


# Telegram Bot API hard cap for sendDocument multipart uploads.
# Anything larger is rejected up-front with an explicit error message.
TELEGRAM_BOT_DOCUMENT_LIMIT_BYTES = 50 * 1024 * 1024

# Cooldown so rapid duplicate taps on the same file button in the same
# chat don't spawn N parallel uploads / N error messages.
_FILE_DELIVERY_COOLDOWN_SECONDS = 4.0
_FILE_DELIVERY_LOCK = threading.Lock()
_FILE_DELIVERY_RECENT: dict[tuple[str, int], float] = {}


def _file_delivery_should_skip(chat_id, file_id) -> bool:
    """True if the same (chat, file) was already dispatched a moment ago."""
    if not file_id:
        return False
    key = (str(chat_id), int(file_id))
    now = time.time()
    with _FILE_DELIVERY_LOCK:
        last = _FILE_DELIVERY_RECENT.get(key, 0.0)
        if now - last < _FILE_DELIVERY_COOLDOWN_SECONDS:
            return True
        _FILE_DELIVERY_RECENT[key] = now
        # Periodically prune stale entries so the dict stays small.
        if len(_FILE_DELIVERY_RECENT) > 256:
            cutoff = now - 60.0
            for k, ts in list(_FILE_DELIVERY_RECENT.items()):
                if ts < cutoff:
                    _FILE_DELIVERY_RECENT.pop(k, None)
    return False


def _send_too_large_message(chat_id, file_info: dict, reason: str) -> None:
    """Notify the chat when Telegram refused to carry the file.

    Per product decision we do NOT send a download link here — we surface
    a clear error so the operator knows the file is too big for Telegram
    and can resolve it manually (compress / split / send another way).
    """
    filename = (file_info.get("filename") or "").strip() or "fayl"
    pretty_reason = html.escape(reason) if reason else "слишком большой для Telegram"
    body = (
        f"📎 <b>{html.escape(filename)}</b>\n"
        f"❌ Не удалось отправить файл: {pretty_reason}.\n"
        "Telegram ограничивает размер документа от бота 50 МБ — "
        "сожмите файл или пришлите частями."
    )
    try:
        telegram_send_message(chat_id, body)
    except Exception:
        app.logger.exception("Failed to send too-large notification")


def _is_too_large_error(exc: Exception) -> bool:
    msg = str(exc).lower()
    return (
        "request entity too large" in msg
        or "413" in msg
        or "file is too big" in msg
        or "too large" in msg
    )


def _deliver_file_async(chat_id, file_info: dict) -> None:
    """Background-thread file delivery for the BL file inline-button.

    Strategy:
      1. If we have a cached Telegram file_id for this file, use it — no
         disk read, no upload, near-instant delivery from Telegram's CDN.
      2. If cached id fails (Telegram dropped it), wipe the cache and fall
         back to a fresh upload from disk, then save the new file_id.
      3. If the file is larger than Telegram's 50 MB bot limit (or upload
         returns 413), skip the upload entirely and send a clickable
         public download link instead.
      4. On any other error, message the chat with a human-readable
         description so the user knows what happened.
    """
    file_id = file_info.get("id")
    file_path = file_info.get("file_path") or ""
    filename = file_info.get("filename") or ""
    cached_tg_id = (file_info.get("tg_file_id") or "").strip()

    # De-dupe rapid taps on the same button so we don't spawn 4 parallel
    # uploads / 4 identical error messages.
    if _file_delivery_should_skip(chat_id, file_id):
        return

    # Resolve BL id for recall-bookkeeping below (file → BL → batch).
    bl_id_for_recall = None
    batch_id_for_recall = None
    try:
        if file_id:
            file_row = db.get_file_by_id(int(file_id)) or {}
            bl_id_for_recall = file_row.get("bl_id")
            if bl_id_for_recall:
                bl_row = db.get_bl_by_id(int(bl_id_for_recall)) or {}
                batch_id_for_recall = bl_row.get("batch_id")
    except Exception:
        pass

    # Fast path: previously uploaded — reuse Telegram's CDN copy.
    if cached_tg_id:
        try:
            cdn_payload = telegram_send_document_by_file_id(chat_id, cached_tg_id)
            try:
                msg_id = _extract_message_id(cdn_payload)
                if msg_id:
                    db.record_sent_telegram_message(
                        bl_id=bl_id_for_recall,
                        batch_id=batch_id_for_recall,
                        chat_id=chat_id,
                        message_id=msg_id,
                        kind="file",
                    )
            except Exception:
                app.logger.exception("Failed to record file message id (CDN path)")
            return
        except Exception as exc:
            app.logger.warning(
                "cached tg_file_id failed (file_id=%s, will re-upload): %s",
                file_id, exc,
            )
            if file_id:
                try:
                    db.clear_file_tg_file_id(file_id)
                except Exception:
                    pass
            # Fall through to fresh upload below.

    if not file_path or not os.path.exists(file_path):
        telegram_send_message(
            chat_id,
            "❌ Файл недоступен на сервере. Свяжитесь с менеджером.",
        )
        return

    # Pre-flight size check: don't even start a doomed multipart upload.
    try:
        file_size = os.path.getsize(file_path)
    except OSError:
        file_size = 0
    if file_size and file_size > TELEGRAM_BOT_DOCUMENT_LIMIT_BYTES:
        size_mb = file_size / (1024 * 1024)
        _send_too_large_message(
            chat_id,
            file_info,
            f"размер {size_mb:.1f} МБ превышает 50 МБ",
        )
        return

    # Slow path: upload from disk and cache the new file_id.
    try:
        payload = telegram_send_document(chat_id, file_path, filename)
        new_id = _extract_tg_file_id_from_send_response(payload)
        if new_id and file_id:
            try:
                db.set_file_tg_file_id(file_id, new_id)
            except Exception as exc:
                app.logger.warning("failed to cache tg_file_id: %s", exc)
        try:
            msg_id = _extract_message_id(payload)
            if msg_id:
                db.record_sent_telegram_message(
                    bl_id=bl_id_for_recall,
                    batch_id=batch_id_for_recall,
                    chat_id=chat_id,
                    message_id=msg_id,
                    kind="file",
                )
        except Exception:
            app.logger.exception("Failed to record file message id (upload path)")
    except Exception as exc:
        app.logger.exception("file delivery failed (file_id=%s)", file_id)
        if _is_too_large_error(exc):
            size_label = f"{file_size / (1024 * 1024):.1f} МБ" if file_size else "размер неизвестен"
            _send_too_large_message(
                chat_id,
                file_info,
                f"Telegram отклонил файл ({size_label})",
            )
            return
        try:
            telegram_send_message(
                chat_id,
                f"❌ Не удалось отправить файл: {html.escape(str(exc)[:200])}",
            )
        except Exception:
            pass


def _send_attached_files_for_bl(chat_id, bl_id) -> None:
    """Synchronously send every packing list attached to a BL, in order.

    Used as the auto-follow-up step right after a tracking message so the
    client sees: info(batch A) → files(batch A) → info(batch B) →
    files(batch B). Reuses _deliver_file_async (which is the per-file
    cached/size-safe sender) — runs it in-thread, not in a new thread,
    so the calling loop stays in order.
    """
    if not bl_id:
        return
    try:
        files = db.get_files(int(bl_id)) or []
    except Exception:
        app.logger.exception("Failed to load files for bl_id=%s", bl_id)
        return

    deliverable = [f for f in files if (f.get("public_token") or "").strip()]
    if not deliverable:
        return

    for index, file_info in enumerate(deliverable):
        try:
            _deliver_file_async(chat_id, dict(file_info))
        except Exception:
            app.logger.exception(
                "Auto file delivery failed (bl_id=%s file_id=%s)",
                bl_id, file_info.get("id"),
            )
        # Per-chat pacing. Telegram throttles bots at ~1 msg/sec into a
        # group chat — sending 5 files at 0.25s gaps reliably trips the
        # 429 rate limit. 1.1s keeps us safely under, and Telegram still
        # rarely flags it; if it does, sendDocument now reads retry_after
        # and waits the prescribed time.
        if index < len(deliverable) - 1:
            try:
                time.sleep(1.1)
            except Exception:
                pass


def _deliver_all_files_async(chat_id, bl_id: int) -> None:
    """Send every file attached to a BL in one go.

    Uses the same cache + size-limit logic as single-file delivery, just
    in a loop. A tiny pause between sends keeps message ordering stable
    in Telegram clients without hitting flood limits.
    """
    try:
        files = db.get_files(bl_id) or []
    except Exception:
        app.logger.exception("Failed to load files for bl_id=%s", bl_id)
        files = []

    deliverable = [f for f in files if (f.get("public_token") or "").strip()]
    if not deliverable:
        try:
            telegram_send_message(chat_id, "❌ К этому BL не прикреплено ни одного файла.")
        except Exception:
            pass
        return

    sent = 0
    failed = 0
    for index, file_info in enumerate(deliverable):
        try:
            _deliver_file_async(chat_id, dict(file_info))
            sent += 1
        except Exception:
            failed += 1
            app.logger.exception("Bulk file delivery failed for file_id=%s", file_info.get("id"))
        # Telegram throttles bots at ~30 msg/sec into a single chat; even
        # for many files this tiny gap keeps us safely under and preserves
        # order on the client side.
        if index < len(deliverable) - 1:
            try:
                time.sleep(0.25)
            except Exception:
                pass

    app.logger.info(
        "Bulk file delivery bl_id=%s chat=%s sent=%s failed=%s total=%s",
        bl_id, chat_id, sent, failed, len(deliverable),
    )


def handle_callback_query(callback_query: dict):
    callback_id = callback_query.get("id")
    data = (callback_query.get("data") or "").strip()
    voter = callback_query.get("from") or {}
    message = callback_query.get("message") or {}
    chat = message.get("chat") or {}
    chat_id = chat.get("id")
    message_id = message.get("message_id")

    if not callback_id or not data or not chat_id:
        return

    if data.startswith("aiact:"):
        handle_ai_action_callback(callback_query, callback_id, data)
        return

    if data.startswith("trkform:"):
        # Старая кнопочная форма заменена мини-приложением (/tgform);
        # старые карточки в истории чата просто гаснут.
        telegram_answer_callback_query(callback_id, "Forma yangilandi — /form buyrug'ini yuboring")
        return

    if data.startswith(f"{FILE_ALL_PREFIX}:"):
        # New bulk button: deliver every packing list attached to the BL.
        try:
            bl_id = int(data.split(":", 1)[1].strip())
        except (TypeError, ValueError):
            telegram_answer_callback_query(callback_id, "Неверный BL")
            return
        # Acknowledge immediately so the spinner disappears.
        try:
            telegram_answer_callback_query(callback_id, "📦 Отправляю packing list...")
        except Exception:
            pass
        threading.Thread(
            target=_deliver_all_files_async,
            args=(chat_id, bl_id),
            daemon=True,
        ).start()
        return

    if data.startswith(f"{FILE_PREFIX}:"):
        # Legacy single-file button (kept for backward-compat with messages
        # sent before the bulk button was introduced).
        token = data.split(":", 1)[1].strip()
        file_info = db.get_file_by_public_token(token)
        if not file_info:
            telegram_answer_callback_query(callback_id, "Файл не найден")
            return

        try:
            telegram_answer_callback_query(callback_id, "📤 Файл отправляется...")
        except Exception:
            pass

        threading.Thread(
            target=_deliver_file_async,
            args=(chat_id, dict(file_info)),
            daemon=True,
        ).start()
        return

    if not data.startswith(f"{COMM_RATE_PREFIX}:"):
        telegram_answer_callback_query(callback_id, "Неизвестное действие")
        return

    parts = data.split(":")
    if len(parts) != 3:
        telegram_answer_callback_query(callback_id, "Неверный формат оценки")
        return

    _, dispatch_or_month, score_raw = parts
    try:
        score = int(score_raw)
    except ValueError:
        telegram_answer_callback_query(callback_id, "Оценка не распознана")
        return

    dispatch_id = None
    month_key = dispatch_or_month
    try:
        dispatch_id = int(dispatch_or_month)
        month_key = ""
    except ValueError:
        pass

    if not db.save_communication_rating(dispatch_id, month_key, chat_id, score, voter=voter):
        telegram_answer_callback_query(callback_id, "Не удалось сохранить оценку")
        return

    telegram_answer_callback_query(
        callback_id,
        f"Спасибо! Оценка {communication_rating_label(score)} сохранена",
    )

    if message_id:
        try:
            telegram_delete_message(chat_id, message_id)
        except Exception:
            pass
def configure_telegram_webhook():
    if not BOT_TOKEN or not WEBHOOK_BASE_URL:
        return False

    payload = {
        "url": f"{WEBHOOK_BASE_URL.rstrip('/')}/telegram/webhook",
        "drop_pending_updates": True,
    }
    if WEBHOOK_SECRET:
        payload["secret_token"] = WEBHOOK_SECRET

    telegram_api("setWebhook", json=payload)
    return True


def remember_group_chat(chat: dict, is_active: bool = True):
    if not chat:
        return

    chat_type = chat.get("type")
    if chat_type not in {"group", "supergroup"}:
        return

    db.upsert_telegram_chat(
        chat_id=chat.get("id"),
        title=chat.get("title") or f"Group {chat.get('id')}",
        chat_type=chat_type,
        username=chat.get("username") or "",
        is_active=is_active,
    )


_BOT_USERNAME_CACHE = {"username": "", "checked_at": 0.0}


def get_bot_username() -> str:
    """Bot's @username via getMe, cached (retry at most every 5 min)."""
    now = time.time()
    if _BOT_USERNAME_CACHE["username"]:
        return _BOT_USERNAME_CACHE["username"]
    if now - _BOT_USERNAME_CACHE["checked_at"] < 300:
        return ""
    _BOT_USERNAME_CACHE["checked_at"] = now
    try:
        payload = telegram_api("getMe") or {}
        username = str(((payload.get("result") or {}).get("username")) or "").strip()
        if username:
            _BOT_USERNAME_CACHE["username"] = username
        return username
    except Exception:
        app.logger.exception("getMe failed while resolving bot username")
        return ""


def extract_bot_mention_question(message: dict) -> str | None:
    """If the bot is @mentioned in the group message, return the text
    with the mention stripped (the actual question). None otherwise."""
    text = message.get("text") or ""
    username = get_bot_username()
    if not username:
        return None
    if ("@" + username.lower()) not in text.lower():
        return None
    stripped = re.sub(re.escape("@" + username), "", text, flags=re.IGNORECASE)
    return " ".join(stripped.split()).strip()


def maybe_handle_group_ai_message(message: dict) -> bool:
    chat = message.get("chat") or {}
    chat_type = chat.get("type")
    if chat_type not in {"group", "supergroup"}:
        return False

    chat_id = chat.get("id")
    text = (message.get("text") or "").strip()
    if not chat_id or not text or text.startswith("/") or text == CANCEL_BUTTON or text in TRACK_BUTTON_TEXTS:
        return False

    from services import ai_assistant

    # ── ACCESS MODEL (19.08.2026) ────────────────────────────────────
    # Exactly ONE staff group (AI_CONTROL_GROUP_ID) has full rights over
    # the bot — its members talk to the FULL assistant (all read tools +
    # approval-gated actions) by @mentioning it. Every other group is a
    # client group where the bot stays COMPLETELY silent for now.
    if not ai_assistant.is_control_chat(chat_id):
        return False

    question = extract_bot_mention_question(message)
    if question is None:
        return False
    if not question:
        # Bare @mention with no actual question — stay silent.
        return True

    chat_title = chat.get("title") or ""
    sender = message.get("from") or {}
    sender_name = telegram_user_name(sender)
    # Speaker attribution: several staff members share this chat, so the
    # assistant sees who is asking. History/pending actions are keyed by
    # the group id.
    assistant_input = f"{sender_name}: {question}" if sender_name else question

    def _worker():
        try:
            with TypingIndicator(chat_id):
                result = ai_assistant.handle_owner_message(f"group:{chat_id}", assistant_input)
        except Exception as exc:
            app.logger.exception("Control-group assistant failure")
            try:
                telegram_send_message(chat_id, f"⚠️ Ошибка ассистента: {exc}", parse_mode=None)
            except Exception:
                pass
            return
        reply = result.get("reply") or "…"
        try:
            telegram_send_message(chat_id, reply)
        except Exception:
            try:
                telegram_send_message(chat_id, reply, parse_mode=None)
            except Exception:
                app.logger.exception("Control-group assistant reply delivery failed")
        for act in result.get("pending") or []:
            try:
                send_ai_action_approval(chat_id, act)
            except Exception:
                app.logger.exception("Control-group approval card delivery failed")
        try:
            db.record_ai_log(
                chat_id=chat_id,
                group_title=chat_title,
                user_id=sender.get("id") or "",
                username=sender.get("username") or sender_name,
                original_text=question,
                detected_intent="control_group_assistant",
                bl_code="",
                ai_response=reply[:1000],
            )
        except Exception:
            app.logger.exception("AI log write failed for chat %s", chat_id)

    threading.Thread(target=_worker, daemon=True).start()
    return True


def send_ai_diagnostic(chat_id, chat: dict):
    chat_title = (chat or {}).get("title") or ""
    chat_type = (chat or {}).get("type") or ""
    chat_id_value = (chat or {}).get("id")
    global_ai_enabled = False
    group_ai_enabled = False
    ai_module_ok = False
    key_present = False
    model_name = ""
    error_text = ""
    try:
        global_ai_enabled = db.get_global_ai_enabled()
        group_ai_enabled = db.get_chat_ai_enabled(chat_id_value)
    except Exception as exc:
        error_text = f"DB settings read failed: {exc}"

    for module_name in ("services.ai_service", "ai_service"):
        try:
            ai_service = __import__(module_name, fromlist=["*"])
            ai_module_ok = True
            runtime_status_getter = getattr(ai_service, "get_runtime_status", None)
            if callable(runtime_status_getter):
                runtime_status = runtime_status_getter() or {}
                key_present = bool(runtime_status.get("openai_api_key_present"))
                model_name = str(runtime_status.get("openai_model") or "").strip()
            break
        except Exception as exc:
            error_text = str(exc)

    lines = [
        "AI diagnostic",
        f"chat_id: {chat_id_value}",
        f"group_title: {chat_title}",
        f"chat_type: {chat_type}",
        f"global_ai_enabled: {global_ai_enabled}",
        f"group_ai_enabled: {group_ai_enabled}",
        f"ai_module_ok: {ai_module_ok}",
        f"openai_api_key_present: {key_present}",
    ]
    if model_name:
        lines.append(f"openai_model: {model_name}")
    if error_text:
        lines.append(f"error: {error_text}")
    telegram_send_message(chat_id, "<pre>" + html.escape("\n".join(lines)) + "</pre>")


def run_ai_test(chat_id, chat: dict, raw_text: str):
    test_text = (raw_text or "").strip()
    if not test_text:
        telegram_send_message(
            chat_id,
            "Usage:\n<code>/aitest Salom</code>\n<code>/aitest BL123 qayerda?</code>",
        )
        return

    ai_service = None
    for module_name in ("services.ai_service", "ai_service"):
        try:
            ai_service = __import__(module_name, fromlist=["*"])
            break
        except Exception:
            continue

    if not ai_service:
        telegram_send_message(chat_id, "AI test error: ai_service module is missing")
        return

    analyzer = getattr(ai_service, "analyze_message", None) or getattr(ai_service, "handle_group_message", None)
    if not callable(analyzer):
        telegram_send_message(chat_id, "AI test error: analyzer is missing")
        return

    try:
        if analyzer.__name__ == "handle_group_message":
            result = analyzer({"text": test_text})
        else:
            result = analyzer(test_text)
        payload = json.dumps(result or {}, ensure_ascii=False, indent=2)
        telegram_send_message(chat_id, "<pre>" + html.escape(payload) + "</pre>")
    except Exception as exc:
        telegram_send_message(chat_id, "<pre>" + html.escape(f"AI test error: {exc}") + "</pre>")


def _send_single_bl_status(chat_id, bl: dict, batch_name: str):
    """Send tracking for one specific BL, then drop its packing-list files.

    Order on the wire: tracking message for this BL → all of this BL's
    attached files → return so the caller can move on to the next BL.
    """
    text = db.render_message(bl, batch_name, include_related_batches=False)
    batch = db.get_batch(bl.get("batch_id")) if bl.get("batch_id") else None
    show_packing_list = not db.is_customer_delivery_eta((batch or {}).get("eta_destination") or "")
    language = normalize_message_language(bl.get("message_language"))
    sent_response = None
    try:
        sent_response = send_with_track_keyboard(chat_id, text, language=language)
    except Exception:
        try:
            sent_response = send_with_track_keyboard(
                chat_id,
                _plain_text_message(text),
                language=language,
                parse_mode=None,
            )
        except Exception:
            app.logger.exception("Yuk holati: both send attempts failed")
            sent_response = None

    # Same recall-bookkeeping as the admin-send path.
    if sent_response is not None:
        try:
            msg_id = _extract_message_id(sent_response)
            if msg_id:
                db.record_sent_telegram_message(
                    bl_id=bl.get("id"),
                    batch_id=bl.get("batch_id"),
                    chat_id=chat_id,
                    message_id=msg_id,
                    kind="tracking",
                )
        except Exception:
            app.logger.exception("Failed to record Yuk-holati message id")

    if show_packing_list:
        try:
            _send_attached_files_for_bl(chat_id, bl.get("id"))
        except Exception:
            app.logger.exception("Auto file delivery failed for bl_id=%s", bl.get("id"))


def send_bl_status(chat_id, bl: dict, in_transit_only: bool = False):
    """Handle the "Yuk holati" button: send each active batch as its own message.

    Mirrors the admin-panel logic in send_bl_package — discover related
    batches for this client and emit one Telegram message per BL so each
    one carries its own file inline-keyboard.
    """
    primary_language = normalize_message_language(bl.get("message_language"))
    bundle = db.get_tracking_bundle_bls(bl, include_related_batches=True, in_transit_only=in_transit_only)
    if not bundle:
        bundle = [bl]

    batch_name_cache: dict[int, str] = {}
    if bl.get("batch_id"):
        batch_name_cache[int(bl["batch_id"])] = bl.get("batch_name") or ""

    for index, item in enumerate(bundle):
        item_batch_id = item.get("batch_id")
        item_batch_name = item.get("batch_name") or ""
        if item_batch_id and not item_batch_name:
            cached = batch_name_cache.get(int(item_batch_id))
            if cached is None:
                related_batch = db.get_batch(int(item_batch_id)) or {}
                cached = related_batch.get("name") or ""
                batch_name_cache[int(item_batch_id)] = cached
            item_batch_name = cached

        # Force the primary language for every message so the client doesn't
        # get a jarring mix when related batches were saved in another locale.
        item_with_lang = dict(item)
        item_with_lang["message_language"] = primary_language

        _send_single_bl_status(chat_id, item_with_lang, item_batch_name)

        # Tiny gap between messages so they appear in order and don't trip
        # Telegram per-chat flood limits.
        if index < len(bundle) - 1:
            try:
                time.sleep(0.4)
            except Exception:
                pass


def send_requested_file(chat_id, file_info: dict | None):
    if not file_info:
        telegram_send_message(chat_id, "❌ Fayl topilmadi.")
        return
    # Same delivery path as the inline-button callback: prefers cached
    # tg_file_id, falls back to upload. Runs in a background thread so the
    # webhook returns immediately and the chat feels responsive.
    threading.Thread(
        target=_deliver_file_async,
        args=(chat_id, dict(file_info)),
        daemon=True,
    ).start()


def handle_bl_lookup(chat_id, raw_code: str):
    code = raw_code.strip().upper()
    if not code:
        telegram_send_message(
            chat_id,
            "Введи <b>BL-код</b> текстом.\nНапример: <code>BL171</code>",
            reply_markup=CANCEL_REPLY_MARKUP,
        )
        return

    bl = db.find_bl_by_code(code)
    if not bl:
        telegram_send_message(
            chat_id,
            f"❌ BL-код <b>{code}</b> не найден.\n\nПроверь код и отправь его ещё раз.",
            reply_markup=CANCEL_REPLY_MARKUP,
        )
        return

    db.clear_chat_state(chat_id)
    # Threaded so the webhook returns immediately while files are sent.
    threading.Thread(
        target=send_bl_status,
        args=(chat_id, bl),
        daemon=True,
    ).start()


def handle_telegram_message(message: dict):
    chat = message.get("chat") or {}
    chat_id = chat.get("id")
    chat_type = chat.get("type")
    sender = message.get("from") or {}
    sender_id = sender.get("id") or chat_id
    text = (message.get("text") or "").strip()

    remember_group_chat(chat, is_active=True)

    # ZIP with packing lists in the control group — handled before the
    # text check (documents have no text).
    if maybe_handle_control_group_document(message):
        return

    if not chat_id or not text:
        return


    if chat_type in {"group", "supergroup"} and not text.startswith("/"):
        app.logger.info(
            "TG_GROUP_TEXT chat_id=%s group_title=%r sender_id=%s text=%r",
            chat_id,
            chat.get("title") or "",
            sender_id,
            text,
        )

    bot_command = extract_bot_command(text)
    if handle_group_remove_request(message, bot_command):
        return

    if bot_command in {"formon", "form_on", "formoff", "form_off"} and chat_type in {"group", "supergroup"}:
        handle_tgform_toggle_command(message, bot_command)
        return

    # /form в группе (нативное меню команд «/» видно у поля ввода) —
    # мгновенно даёт кнопку открытия Mini App
    if bot_command == "form" and chat_type in {"group", "supergroup"}:
        if str(chat_id) in _tgform_enabled_groups():
            telegram_send_message(
                chat_id,
                "📝 Formani ochish:",
                reply_markup=_tgform_group_keyboard(chat_id),
                disable_notification=True,
            )
        return

    if bot_command in AI_STATUS_COMMANDS:
        send_ai_diagnostic(chat_id, chat)
        return

    if bot_command in AI_TEST_COMMANDS:
        raw_test_text = re.sub(r"^/[A-Za-z0-9_]+(?:@\w+)?\s*", "", text, count=1).strip()
        run_ai_test(chat_id, chat, raw_test_text)
        return

    if bot_command in MENU_RESTORE_COMMANDS:
        db.clear_chat_state(chat_id)
        language = get_chat_message_language(chat_id)
        if chat_type in {"group", "supergroup"}:
            send_group_message_with_keyboard(chat_id, get_menu_restore_text(language), language=language)
        else:
            send_with_track_keyboard(chat_id, get_menu_restore_text(language), language=language)
        return

    if text == "/start" or text.startswith("/start "):
        # Bot relaunch (19.08.2026): no welcome bundle anymore. The only
        # /start payload is "form" — the group button deep-links here so
        # the private chat can get the Web App button (в группах Telegram
        # web_app-кнопки не разрешает).
        db.clear_chat_state(chat_id)
        payload = text[7:].strip().lower() if text.startswith("/start ") else ""
        if payload == "form" and chat_type == "private":
            send_tracking_form_button(chat_id)
        return

    if text == "/chatid":
        title = chat.get("title") or "Личный чат"
        telegram_send_message(
            chat_id,
            f"📍 Чат: <b>{title}</b>\n🆔 ID: <code>{chat_id}</code>",
        )
        return

    # ── AI assistant (DeepSeek): private chat, admin + read-only ids ──
    # Admin ids get the full assistant; read-only ids (например,
    # ответственный за трекинг) — те же READ-инструменты без права
    # что-либо менять. Group messages are NOT routed here.
    if chat_type == "private":
        from services import ai_assistant

        if ai_assistant.is_admin(sender_id) or ai_assistant.is_readonly_user(sender_id):
            handle_ai_assistant_private_message(chat_id, sender_id, text)
            return

    file_match = re.match(r"^/([A-Za-z0-9_]+)(?:@\w+)?$", text)
    if file_match:
        file_info = db.get_file_by_command_alias(file_match.group(1))
        if file_info:
            send_requested_file(chat_id, file_info)
            return

    if text in TRACK_BUTTON_TEXTS:
        # The "Yuk holati" flow is FULLY retired (owner request,
        # 19.08.2026): the button is gone and the bot no longer answers
        # this text at all — consumed silently so it can't fall through
        # to the group-AI handler either.
        return

    # Постоянная кнопка «📝 Treking forma» у поля ввода (включённые группы)
    if text == TRACKING_FORM_BUTTON_TEXT and chat_type in {"group", "supergroup"}:
        if str(chat_id) in _tgform_enabled_groups():
            telegram_send_message(
                chat_id,
                "📝 Formani ochish uchun tugmani bosing:",
                reply_markup=_tgform_group_keyboard(chat_id),
                disable_notification=True,
            )
        return

    # ZIP-ссылка (Google Drive / прямой URL) в управляющей группе —
    # большие архивы packing list приходят так (Telegram >20MB не отдаёт).
    if maybe_handle_control_group_zip_link(message):
        return

    if maybe_handle_group_ai_message(message):
        return

    if text == CANCEL_BUTTON or db.get_chat_state(chat_id) == STATE_WAITING_BL:
        # Legacy ask-for-BL flow retired alongside the button: clear any
        # lingering state silently, no reply.
        db.clear_chat_state(chat_id)
        return


# ═══════════════════════════════════════════════════════════════
# AI ASSISTANT (DeepSeek) — owner private chat + approval flow
# ═══════════════════════════════════════════════════════════════

def handle_ai_assistant_private_message(chat_id, sender_id, text: str):
    """Route one owner message to the assistant on a worker thread."""
    from services import ai_assistant

    lowered = (text or "").strip().lower()
    if lowered in {"/reset", "/aireset"}:
        ai_assistant.reset_history(sender_id)
        telegram_send_message(chat_id, "🧹 История диалога с ассистентом очищена.")
        return
    if lowered in {"/form", "/trekform", "/tracking"}:
        send_tracking_form_button(chat_id)
        return

    readonly = ai_assistant.is_readonly_user(sender_id)
    # Владельческий режим (прямые инструменты) — только личка админа.
    owner_direct = ai_assistant.is_admin(sender_id)

    def _worker():
        try:
            # Keep "typing…" visible while the assistant thinks — the
            # agent loop with tools can take a while and the chat looked
            # dead in the meantime.
            with TypingIndicator(chat_id):
                result = ai_assistant.handle_owner_message(sender_id, text, readonly=readonly, owner_direct=owner_direct)
        except Exception as exc:
            app.logger.exception("AI assistant failure")
            try:
                telegram_send_message(chat_id, f"⚠️ Ошибка ассистента: {exc}", parse_mode=None)
            except Exception:
                pass
            return
        reply = result.get("reply") or "…"
        try:
            telegram_send_message(chat_id, reply)
        except Exception:
            # Model text can contain broken HTML — retry as plain text.
            try:
                telegram_send_message(chat_id, reply, parse_mode=None)
            except Exception:
                app.logger.exception("AI assistant reply delivery failed")
        for act in result.get("pending") or []:
            try:
                send_ai_action_approval(chat_id, act)
            except Exception:
                app.logger.exception("AI approval card delivery failed")

    threading.Thread(target=_worker, daemon=True).start()


def send_ai_action_approval(chat_id, act: dict):
    from html import escape as html_escape

    action_id = act.get("id")
    summary = act.get("summary") or ""
    keyboard = {
        "inline_keyboard": [[
            {"text": "✅ Подтвердить", "callback_data": f"aiact:ok:{action_id}"},
            {"text": "❌ Отклонить", "callback_data": f"aiact:no:{action_id}"},
        ]]
    }
    telegram_send_message(
        chat_id,
        (
            "🤖 <b>Требуется подтверждение</b>\n\n"
            f"{html_escape(summary)}\n\n"
            f"<i>Заявка #{action_id}. Без подтверждения ничего не изменится.</i>"
        ),
        reply_markup=keyboard,
    )


def execute_ai_action(action: dict):
    """Run one APPROVED assistant action. Returns (ok, human_result)."""
    kind = action.get("kind")
    try:
        params = json.loads(action.get("params_json") or "{}")
    except Exception:
        params = {}

    if kind == "set_batch_status":
        batch_id = int(params.get("batch_id") or 0)
        status = str(params.get("status") or "").strip()
        batch = db.get_batch(batch_id)
        if not batch:
            return False, "Партия не найдена"
        valid = set(db.STATUSES) | {db.DELIVERED_STATUS}
        if status not in valid:
            return False, f"Недопустимый статус «{status}»"
        db.update_batch(
            batch_id,
            batch["name"],
            status,
            batch.get("eta_to_toshkent") or "",
            batch.get("eta_destination") or "Toshkent",
            batch.get("client_delivery_date") or "",
        )
        return True, f"Статус партии «{batch['name']}»: {batch['status']} → {status}"

    if kind == "send_group_message":
        target_chat = str(params.get("chat_id") or "").strip()
        text = str(params.get("text") or "").strip()
        if not target_chat or not text:
            return False, "Пустой chat_id или текст"
        telegram_send_message(target_chat, text)
        return True, f"Сообщение отправлено в чат {target_chat}"

    if kind == "sync_batch_from_plan":
        batch_id = int(params.get("batch_id") or 0)
        batch = db.get_batch(batch_id)
        if not batch:
            return False, "Партия не найдена"
        marks = params.get("marks") or []
        if not marks:
            return False, "В заявке нет снапшота BL из плана"
        plan_keys = {_normalize_bl_code(str(m.get("mark") or "")) for m in marks if str(m.get("mark") or "").strip()}
        existing = {
            _normalize_bl_code(str(bl.get("code") or ""))
            for bl in db.get_bl_by_batch(batch_id)
        }
        chat_lookup = _build_chat_code_lookup()
        known_chat_ids = _known_chat_id_set()
        added = []
        linked = 0
        already = 0
        for m in marks:
            code = str(m.get("mark") or "").strip()
            if not code:
                continue
            norm = _normalize_bl_code(code)
            if norm in existing:
                already += 1
                continue
            # Same auto-attach chain as the manual sheet import: match a
            # group by title first, then the permanent link history.
            target_chat = _find_chat_for_bl_code(code, chat_lookup)
            if not target_chat:
                hist = db.lookup_bl_link_history(code)
                if hist and (not known_chat_ids or hist["chat_id"] in known_chat_ids):
                    target_chat = hist["chat_id"]
            created = db.add_bl(
                batch_id=batch_id,
                code=code,
                client_name="",
                chat_id=target_chat,
                cargo_type="",
                weight_kg=m.get("kg", 0),
                volume_cbm=m.get("cbm", 0),
                quantity_places=m.get("ctn", 0),
                quantity_places_breakdown="",
                cargo_description="",
                message_language=getattr(db, "DEFAULT_MESSAGE_LANGUAGE", "uz_latn"),
            )
            if created:
                added.append(code)
                existing.add(norm)
                if target_chat:
                    linked += 1
        # BLs that are NOT in the plan get REMOVED from the batch: the
        # cargo isn't on this truck, so leaving them would send wrong
        # tracking to their groups on the next broadcast. The permanent
        # bl_link_history keeps the group link for when the BL reappears
        # in another plan/batch.
        removed = []
        remove_failed = []
        for bl in db.get_bl_by_batch(batch_id):
            code = str(bl.get("code") or "")
            if _normalize_bl_code(code) in plan_keys:
                continue
            try:
                db.delete_bl(bl["id"])
                removed.append(code)
            except Exception as exc:
                app.logger.exception("Plan sync: failed to delete bl_id=%s", bl.get("id"))
                remove_failed.append(f"{code} ({exc})")
        msg = (
            f"Синхронизация «{batch['name']}» с планом «{params.get('plan_title', '')}» "
            f"({params.get('plan_date', '')}): добавлено {len(added)} BL"
            f" (групп привязано автоматически: {linked}), уже были: {already}, удалено вне плана: {len(removed)}."
        )
        if added:
            msg += f" Новые: {', '.join(added[:15])}."
        if removed:
            msg += f" Удалены: {', '.join(removed[:15])}."
        if remove_failed:
            msg += f" ⚠️ Не удалось удалить: {', '.join(remove_failed[:5])}."
        return not remove_failed, msg

    if kind == "send_tracking_batch":
        batch_id = int(params.get("batch_id") or 0)
        batch = db.get_batch(batch_id)
        if not batch:
            return False, "Партия не найдена"
        bl_rows = [
            bl for bl in db.get_bl_by_batch(batch_id)
            if bl.get("chat_id") and not bl.get("send_excluded") and not bl.get("tracking_sent_current")
        ]
        sent_chats: set = set()
        dispatch = []
        for bl in bl_rows:
            cid = str(bl.get("chat_id") or "").strip()
            if not cid or cid in sent_chats:
                continue
            sent_chats.add(cid)
            dispatch.append(bl)
        if not dispatch:
            return False, "Нет BL для отправки: все уже отправлены, исключены или без групп"
        ok_count = 0
        fail_count = 0
        for bl in dispatch:
            with _BULK_SEND_SLOTS:
                try:
                    success, error_msg = send_bl_package(bl, batch["name"], include_related_batches=False)
                except Exception as exc:
                    success, error_msg = False, str(exc)
                    app.logger.exception("AI tracking send failed for bl_id=%s", bl.get("id"))
            try:
                db.add_log(bl["id"], bl["code"], batch["name"], bl["chat_id"], bl["status"], success, error_msg)
            except Exception:
                app.logger.exception("AI tracking send: log write failed")
            if success:
                ok_count += 1
            else:
                fail_count += 1
        return fail_count == 0, f"Трекинг партии «{batch['name']}»: ✅ {ok_count} · ❌ {fail_count}"

    return False, f"Неизвестное действие: {kind}"


def handle_ai_action_callback(callback_query: dict, callback_id, data: str):
    from services import ai_assistant

    voter = callback_query.get("from") or {}
    origin_chat_id = ((callback_query.get("message") or {}).get("chat") or {}).get("id")
    # Approvals: the owner (admin ids) anywhere, or ANY member of the
    # control group when the buttons live in that group.
    if not (ai_assistant.is_admin(voter.get("id")) or ai_assistant.is_control_chat(origin_chat_id)):
        telegram_answer_callback_query(callback_id, "Недостаточно прав")
        return

    parts = data.split(":")
    if len(parts) != 3:
        telegram_answer_callback_query(callback_id, "Неверный формат")
        return
    verdict = parts[1]
    try:
        action_id = int(parts[2])
    except (TypeError, ValueError):
        telegram_answer_callback_query(callback_id, "Неверный id заявки")
        return

    action = db.ai_get_pending_action(action_id)
    origin_chat = ((callback_query.get("message") or {}).get("chat") or {}).get("id")
    if not action:
        telegram_answer_callback_query(callback_id, "Заявка не найдена")
        return
    if action.get("status") != "pending":
        telegram_answer_callback_query(callback_id, f"Уже обработано: {action.get('status')}")
        return

    if verdict == "no":
        db.ai_finish_pending_action(action_id, "rejected")
        telegram_answer_callback_query(callback_id, "Отклонено")
        if origin_chat:
            telegram_send_message(origin_chat, f"❌ Заявка #{action_id} отклонена. Ничего не изменено.")
        return

    if verdict != "ok":
        telegram_answer_callback_query(callback_id, "Неизвестное действие")
        return

    # Mark as approved synchronously so a double-click can't run it twice,
    # then execute on a worker thread (tracking sends can take minutes).
    db.ai_finish_pending_action(action_id, "approved")
    telegram_answer_callback_query(callback_id, "✅ Выполняю...")

    def _worker():
        try:
            ok, result_text = execute_ai_action(action)
        except Exception as exc:
            ok, result_text = False, f"Ошибка выполнения: {exc}"
            app.logger.exception("AI action execution failed")
        db.ai_finish_pending_action(action_id, "executed" if ok else "failed", result_text)
        if origin_chat:
            icon = "✅" if ok else "⚠️"
            try:
                telegram_send_message(origin_chat, f"{icon} Заявка #{action_id}: {result_text}")
            except Exception:
                app.logger.exception("AI action result delivery failed")

    threading.Thread(target=_worker, daemon=True).start()


# ═══════════════════════════════════════════════════════════════
# MORNING PACKING-LIST FLOW (Tracking gruppa)
# ═══════════════════════════════════════════════════════════════
# Every morning the bot @mentions the responsible person (Jigar) in the
# control group with the list of BLs that still have no packing list.
# He replies with a ZIP; the bot downloads it, unpacks, matches every
# file to a BL of an ACTIVE batch (same resolver as the panel's bulk
# upload) and attaches them. Summary is reported back in Uzbek.

PACKING_RESPONSIBLE_TG_ID = os.getenv("PACKING_RESPONSIBLE_TG_ID", "8526226966").strip()
PACKING_RESPONSIBLE_NAME = os.getenv("PACKING_RESPONSIBLE_NAME", "Jigar").strip() or "Jigar"
PACKING_REMINDER_HOUR = int(os.getenv("PACKING_REMINDER_HOUR", "9") or 9)
_PACKING_REMINDER_SETTING = "packing_reminder_last_date"
_PACKING_ZIP_MAX_FILES = 200
_PACKING_FILE_MAX_BYTES = 30 * 1024 * 1024


def get_missing_packing_bls() -> list[dict]:
    """BLs of ACTIVE batches that have no packing list attached."""
    conn = db.get_conn()
    try:
        rows = conn.execute(
            """
            SELECT bl.id, bl.code, bl.client_name, b.name AS batch_name
            FROM bl_codes bl JOIN batches b ON b.id = bl.batch_id
            WHERE COALESCE(b.client_delivery_date, '') = ''
              AND NOT EXISTS (SELECT 1 FROM files f WHERE f.bl_id = bl.id)
            ORDER BY b.id DESC, bl.code
            """
        ).fetchall()
    finally:
        conn.close()
    return [dict(r) for r in rows]


def send_packing_list_reminder(force: bool = False):
    """Morning ask in the control group. Returns (sent, info)."""
    from html import escape as html_escape
    from services import ai_assistant

    chat_id = ai_assistant.control_group_id()
    if not chat_id or not BOT_TOKEN:
        return False, "control group or BOT_TOKEN not configured"
    today = datetime.now(db.TASHKENT_TZ).strftime("%Y-%m-%d")
    if not force and db.get_setting(_PACKING_REMINDER_SETTING) == today:
        return False, "already sent today"
    missing = get_missing_packing_bls()
    if not missing:
        db.set_setting(_PACKING_REMINDER_SETTING, today)
        return False, "nothing missing"

    mention = f'<a href="tg://user?id={PACKING_RESPONSIBLE_TG_ID}">{html_escape(PACKING_RESPONSIBLE_NAME)}</a>'
    by_batch: dict[str, list] = {}
    for r in missing:
        by_batch.setdefault(r["batch_name"], []).append(r["code"])
    lines = [
        f"🌅 Assalomu alaykum, {mention}!",
        "",
        "Quyidagi BL larga packing list hali biriktirilmagan:",
    ]
    shown = 0
    for batch_name, codes in by_batch.items():
        lines.append(f"\n📦 <b>{html_escape(batch_name)}</b>:")
        for code in codes:
            if shown >= 60:
                break
            lines.append(f"  • <code>{html_escape(code)}</code>")
            shown += 1
        if shown >= 60:
            break
    rest = len(missing) - shown
    if rest > 0:
        lines.append(f"\n…va yana {rest} ta BL.")
    lines.append(
        f"\n📎 Iltimos, packing listlarni bitta ZIP arxiv qilib shu guruhga tashlang — "
        f"o'zim ochib, tahlil qilib, tegishli BL larga biriktiraman. "
        f"Katta arxiv bo'lsa (20 MB+), Google Drive papkaga yuklab, havolasini tashlang "
        f"(ZIP formatida — RAR ochilmaydi). Rahmat, {html_escape(PACKING_RESPONSIBLE_NAME)}! 🙏"
    )
    telegram_send_message(chat_id, "\n".join(lines))
    db.set_setting(_PACKING_REMINDER_SETTING, today)
    return True, f"{len(missing)} BL without packing list"


def _run_due_ai_tasks():
    """Исполнение запланированных задач ассистента (владельческий режим):
    отправка сообщений/упоминаний и опросов по расписанию."""
    from html import escape as html_escape

    now_str = datetime.now(db.TASHKENT_TZ).strftime("%Y-%m-%d %H:%M")
    for task in db.ai_due_scheduled_tasks(now_str):
        try:
            params = json.loads(task.get("params_json") or "{}")
        except Exception:
            params = {}
        ok = True
        try:
            chat_id = str(params.get("chat_id") or "").strip()
            if task["kind"] == "send_message":
                text = str(params.get("text") or "")
                mention_id = str(params.get("mention_user_id") or "").strip()
                if mention_id:
                    label = html_escape(str(params.get("mention_label") or "👤"))
                    text = f'<a href="tg://user?id={mention_id}">{label}</a>, {text}'
                telegram_send_message(chat_id, text)
            elif task["kind"] == "send_poll":
                telegram_api("sendPoll", json={
                    "chat_id": chat_id,
                    "question": str(params.get("question") or "")[:300],
                    "options": json.dumps([str(o)[:100] for o in (params.get("options") or [])][:10], ensure_ascii=False),
                    "is_anonymous": bool(params.get("is_anonymous", False)),
                })
            else:
                ok = False
        except Exception:
            ok = False
            app.logger.exception("Scheduled AI task %s failed", task["id"])
        if ok and task.get("recurrence") == "daily":
            try:
                from datetime import timedelta

                next_dt = datetime.strptime(task["run_at"], "%Y-%m-%d %H:%M") + timedelta(days=1)
                # если бот долго не работал — не «догоняем» по дню за тик,
                # а сразу перематываем на ближайшее будущее время
                now_naive = datetime.now(db.TASHKENT_TZ).replace(tzinfo=None)
                while next_dt <= now_naive:
                    next_dt += timedelta(days=1)
                db.ai_mark_scheduled_task(task["id"], "pending", next_dt.strftime("%Y-%m-%d %H:%M"))
            except Exception:
                db.ai_mark_scheduled_task(task["id"], "failed")
        else:
            db.ai_mark_scheduled_task(task["id"], "done" if ok else "failed")
        if not ok:
            try:
                creator = str(task.get("created_by") or "").strip()
                if creator.isdigit():
                    telegram_send_message(creator, f"⚠️ Запланированная задача #{task['id']} не выполнилась.", parse_mode=None)
            except Exception:
                pass


def _packing_reminder_scheduler():
    """Morning tick (every 5 min): packing-list ask + tracking digest.
    Both are self-guarded by per-day settings keys."""
    while True:
        try:
            now = datetime.now(db.TASHKENT_TZ)
            if now.hour == PACKING_REMINDER_HOUR:
                sent, info = send_packing_list_reminder()
                if sent:
                    app.logger.info("Packing reminder sent: %s", info)
            if now.hour == TRACKING_ASK_HOUR:
                sent, info = send_tracking_update_request()
                if sent:
                    app.logger.info("Tracking update request sent: %s", info)
            if now.hour >= TRACKING_DIGEST_HOUR:
                sent, info = send_tracking_digest()
                if sent:
                    app.logger.info("Tracking digest sent: %s", info)
            # авто-жизненный цикл партий из «Fura statuslari»
            synced = sync_batches_from_fura_statuses()
            if synced:
                app.logger.info("Fura-status sync: %s batch updates", synced)
            # запланированные задачи ассистента (сообщения/опросы)
            _run_due_ai_tasks()
        except Exception:
            app.logger.exception("Morning scheduler tick failed")
        time.sleep(60)


def _build_active_bl_index():
    """(code_index, rows) across ALL active batches — for zip matching."""
    conn = db.get_conn()
    try:
        rows = conn.execute(
            """
            SELECT bl.*, b.name AS batch_name
            FROM bl_codes bl JOIN batches b ON b.id = bl.batch_id
            WHERE COALESCE(b.client_delivery_date, '') = ''
            ORDER BY b.id DESC
            """
        ).fetchall()
    finally:
        conn.close()
    rows = [dict(r) for r in rows]
    index: dict[str, dict] = {}
    for row in rows:
        primary = _normalize_bl_code(row.get("code") or "")
        if primary:
            index.setdefault(primary, row)
        for part in re.split(r"[,;\s]+", str(row.get("merged_codes") or "")):
            piece = _normalize_bl_code(part)
            if piece:
                index.setdefault(piece, row)
    return index, rows


# Packing-list filename convention: "<БРЕНД/BL> <N> MESTA <товар>.xlsx",
# e.g. "LUCEAT 10 MESTA LYUSTRA.xlsx", "BL-65 10 MESTA ELEKTROMOBIL….xlsx".
# The brand (before the number) says WHOSE packing list it is; N = carton
# count (CTN/件数 from the Sklad sheet → bl.quantity_places) and is used
# as the secondary check when the brand matches several BLs.
_MESTA_RE = re.compile(r"^(?P<brand>.+?)\s+(?P<mesta>\d{1,5})\s+MESTA\b", re.IGNORECASE)


def _parse_packing_filename(base_name: str):
    """(brand, mesta_count|None) from a packing-list filename."""
    stem = base_name.rsplit(".", 1)[0].strip()
    m = _MESTA_RE.match(stem)
    if not m:
        return stem, None
    return m.group("brand").strip(), int(m.group("mesta"))


def _mesta_matches(bl: dict, mesta: int | None) -> bool:
    if mesta is None:
        return True
    try:
        places = int(round(float(bl.get("quantity_places") or 0)))
    except (TypeError, ValueError):
        places = 0
    if places == mesta:
        return True
    # разбивка мест ("4+11+14") — файл может закрывать одну её часть
    breakdown = str(bl.get("quantity_places_breakdown") or "")
    return mesta in [int(x) for x in re.findall(r"\d+", breakdown)]


def _find_bl_candidates_by_brand(brand: str, rows: list, chat_titles: dict) -> list:
    """Active-batch BLs whose code / client / group title match the brand."""
    brand_norm = _normalize_bl_code(brand)
    brand_up = brand.upper().strip()
    scored = []
    for row in rows:
        code_norm = _normalize_bl_code(str(row.get("code") or ""))
        client = str(row.get("client_name") or "").upper().strip()
        title = str(chat_titles.get(str(row.get("chat_id") or "").strip(), "")).upper()
        score = 0
        if brand_norm and code_norm and brand_norm == code_norm:
            score = 3
        else:
            for part in re.split(r"[,;\s]+", str(row.get("merged_codes") or "")):
                if brand_norm and _normalize_bl_code(part) == brand_norm:
                    score = 3
                    break
        if not score and len(brand_up) >= 3:
            if client and (brand_up == client or brand_up in client or client in brand_up):
                score = 2
            elif title and brand_up in title:
                score = 2
        if score:
            scored.append((score, row))
    if not scored:
        return []
    best = max(s for s, _ in scored)
    return [row for s, row in scored if s == best]


def maybe_handle_control_group_document(message: dict) -> bool:
    """ZIP dropped in the control group → unpack & attach packing lists."""
    from services import ai_assistant

    chat = message.get("chat") or {}
    if chat.get("type") not in {"group", "supergroup"}:
        return False
    chat_id = chat.get("id")
    if not chat_id or not ai_assistant.is_control_chat(chat_id):
        return False
    doc = message.get("document") or {}
    filename = str(doc.get("file_name") or "").lower()
    if not filename.endswith(".zip"):
        return False
    threading.Thread(target=process_packing_zip, args=(chat_id, doc), daemon=True).start()
    return True


def process_packing_zip(chat_id, doc: dict):
    """ZIP sent as a Telegram document (≤20MB — Bot API download limit).
    Larger archives come in via a link (see process_packing_zip_url)."""
    try:
        with TypingIndicator(chat_id):
            info = telegram_api("getFile", json={"file_id": doc.get("file_id")}) or {}
            tg_path = ((info.get("result") or {}).get("file_path")) or ""
            if not tg_path:
                telegram_send_message(
                    chat_id,
                    "⚠️ ZIP faylni yuklab bo'lmadi — fayl 20 MB dan katta bo'lishi mumkin.\n"
                    "Katta arxivlarni Google Drive'ga yuklab, havolasini shu guruhga tashlang — o'zim yuklab olaman.",
                )
                return
            url = f"https://api.telegram.org/file/bot{BOT_TOKEN}/{tg_path}"
            content = req.get(url, timeout=180).content
            _ingest_packing_zip(chat_id, io.BytesIO(content))
    except Exception as exc:
        app.logger.exception("Packing zip (telegram) failed")
        try:
            telegram_send_message(chat_id, f"⚠️ ZIP tahlilida xato: {exc}", parse_mode=None)
        except Exception:
            pass


def _new_packing_results() -> dict:
    return {
        "attached": [], "skipped_dup": [], "unmatched": [],
        "rejected": [], "ambiguous": [], "mesta_warns": [], "unsupported": [],
    }


def _attach_packing_bytes(base: str, data: bytes, index, rows, chat_titles, results: dict):
    """Match ONE packing-list file (brand → MESTA) and attach it."""
    ext = base.rsplit(".", 1)[-1].lower() if "." in base else ""
    if ext not in ALLOWED_EXT:
        results["rejected"].append(base)
        return
    if len(data) > _PACKING_FILE_MAX_BYTES:
        results["rejected"].append(base)
        return
    # 1) главный ключ — бренд/название из имени файла;
    # 2) при нескольких кандидатах решает MESTA (CTN).
    brand, mesta = _parse_packing_filename(base)
    candidates = _find_bl_candidates_by_brand(brand, rows, chat_titles)
    bl = None
    if len(candidates) == 1:
        bl = candidates[0]
        if mesta is not None and not _mesta_matches(bl, mesta):
            results["mesta_warns"].append(
                f"{base}: faylda {mesta} mesta, tizimda {bl.get('quantity_places') or 0}"
            )
    elif len(candidates) > 1:
        filtered = [c for c in candidates if _mesta_matches(c, mesta)] if mesta is not None else []
        if len(filtered) == 1:
            bl = filtered[0]
        else:
            results["ambiguous"].append(
                base + " → " + ", ".join(
                    f"{c.get('code')}({c.get('batch_name')})" for c in candidates[:4]
                )
            )
            return
    else:
        # запасной путь — общий резолвер (нестандартные имена)
        hit = _resolve_filename_to_bl(base, index, rows, chat_titles)
        if hit:
            bl = hit["row"]
    if not bl:
        results["unmatched"].append(base)
        return
    existing = {
        str(f.get("filename") or "").strip().lower()
        for f in (db.get_files(bl["id"]) or [])
    }
    if base.lower() in existing:
        results["skipped_dup"].append(base)
        return
    ext_storage = secure_filename(base) or f"file_{secrets.token_hex(4)}.{ext}"
    unique = f"bl{bl['id']}_{secrets.token_hex(4)}_{ext_storage}"
    path = os.path.join(UPLOAD_FOLDER, unique)
    with open(path, "wb") as fh:
        fh.write(data)
    db.add_file(bl["id"], base, path)
    results["attached"].append((base, bl.get("code") or "", bl.get("batch_name") or ""))


def _ingest_zip_source(zip_source, index, rows, chat_titles, results: dict):
    """Iterate one ZIP (path or file-like) into shared results."""
    import zipfile

    with zipfile.ZipFile(zip_source) as zf:
        names = [n for n in zf.namelist() if not n.endswith("/")]
        for name in names[:_PACKING_ZIP_MAX_FILES]:
            base = os.path.basename(name.replace("\\", "/")).strip()
            if not base or base.startswith("._") or "__MACOSX" in name:
                continue
            _attach_packing_bytes(base, zf.read(name), index, rows, chat_titles, results)


def _send_packing_report(chat_id, results: dict, empty_note: str = "ℹ️ Biriktiradigan fayl topilmadi."):
    from html import escape as html_escape

    lines = []
    attached = results["attached"]
    if attached:
        lines.append(f"✅ Rahmat, {html_escape(PACKING_RESPONSIBLE_NAME)}! {len(attached)} ta packing list biriktirildi:")
        for base, code, batch_name in attached[:30]:
            lines.append(f"  • <code>{html_escape(code)}</code> ← {html_escape(base)} ({html_escape(batch_name)})")
        if len(attached) > 30:
            lines.append(f"  …va yana {len(attached) - 30} ta")
    if results["mesta_warns"]:
        lines.append("⚠️ Mesta (karobka soni) mos kelmadi, lekin nomi aniq bo'lgani uchun biriktirdim:")
        for w in results["mesta_warns"][:10]:
            lines.append(f"  • {html_escape(w)}")
    if results["ambiguous"]:
        lines.append(f"🤔 Bir nechta BL mos keldi, mesta ham aniqlik kiritmadi ({len(results['ambiguous'])} ta):")
        for a in results["ambiguous"][:10]:
            lines.append(f"  • {html_escape(a)}")
        lines.append("Bularni aniqlashtirib qayta yuboring (fayl nomiga BL kodini yozing).")
    if results["skipped_dup"]:
        lines.append(f"↩️ Allaqachon biriktirilgan edi ({len(results['skipped_dup'])} ta) — o'tkazib yubordim.")
    if results["unmatched"]:
        lines.append(
            f"❓ Mos BL topilmadi ({len(results['unmatched'])} ta): "
            + ", ".join(html_escape(u) for u in results["unmatched"][:15])
        )
        lines.append("Bu fayllar nomiga BL kodini yozib qayta yuboring yoki paneldan qo'lda biriktiring.")
    if results["rejected"]:
        lines.append(f"🚫 Format qabul qilinmadi ({len(results['rejected'])} ta): " + ", ".join(html_escape(x) for x in results["rejected"][:10]))
    if results["unsupported"]:
        lines.append(
            f"📦 RAR/7z arxivlarni ocholmayman ({len(results['unsupported'])} ta): "
            + ", ".join(html_escape(x) for x in results["unsupported"][:5])
        )
        lines.append("Iltimos, ZIP formatida yuboring yoki fayllarni papkaga alohida-alohida yuklang.")
    if not lines:
        lines = [empty_note]
    telegram_send_message(chat_id, "\n".join(lines))


def _ingest_packing_zip(chat_id, zip_source):
    """Shared pipeline: open the archive (path or file-like), match every
    file (brand → MESTA), attach and report in Uzbek."""
    import zipfile

    results = _new_packing_results()
    try:
        index, rows = _build_active_bl_index()
        chat_titles = _build_chat_title_lookup(rows)
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        _ingest_zip_source(zip_source, index, rows, chat_titles, results)
    except zipfile.BadZipFile:
        telegram_send_message(chat_id, "⚠️ Arxiv ochilmadi — ZIP fayl buzilgan ko'rinadi. Qayta yuborib ko'ring.")
        return
    except Exception as exc:
        app.logger.exception("Packing zip processing failed")
        try:
            telegram_send_message(chat_id, f"⚠️ ZIP tahlilida xato: {exc}", parse_mode=None)
        except Exception:
            pass
        return
    _send_packing_report(chat_id, results, empty_note="ℹ️ ZIP ichida biriktiradigan fayl topilmadi.")


# ── Большие ZIP по ссылке (Google Drive / прямой URL) ────────────────
# Telegram Bot API не отдаёт файлы >20MB, поэтому большие архивы
# приходят ссылкой: бот скачивает по HTTP (стрим на диск, до 512MB)
# и прогоняет через тот же конвейер.
_PACKING_URL_MAX_BYTES = 512 * 1024 * 1024
_DRIVE_ID_RE = re.compile(
    r"drive\.google\.com/(?:file/d/([\w-]{20,})|open\?id=([\w-]{20,})|uc\?[^\s]*?id=([\w-]{20,}))"
)
_DIRECT_ZIP_RE = re.compile(r"https?://\S+?\.zip(?:\?\S*)?(?=\s|$)", re.IGNORECASE)


_DRIVE_FOLDER_RE = re.compile(r"drive\.google\.com/drive/(?:u/\d+/)?folders/([\w-]{15,})")


def maybe_handle_control_group_zip_link(message: dict) -> bool:
    from services import ai_assistant

    chat = message.get("chat") or {}
    if chat.get("type") not in {"group", "supergroup"}:
        return False
    chat_id = chat.get("id")
    if not chat_id or not ai_assistant.is_control_chat(chat_id):
        return False
    text = (message.get("text") or "").strip()
    if not text:
        return False
    mf = _DRIVE_FOLDER_RE.search(text)
    if mf:
        threading.Thread(target=process_drive_folder, args=(chat_id, mf.group(1)), daemon=True).start()
        return True
    m = _DRIVE_ID_RE.search(text)
    if m:
        file_id = next(g for g in m.groups() if g)
        threading.Thread(target=process_packing_zip_url, args=(chat_id, "drive", file_id), daemon=True).start()
        return True
    m2 = _DIRECT_ZIP_RE.search(text)
    if m2:
        threading.Thread(target=process_packing_zip_url, args=(chat_id, "direct", m2.group(0)), daemon=True).start()
        return True
    return False


def _list_drive_folder(folder_id: str) -> list:
    """[(file_id, name)] of a PUBLIC Drive folder via the legacy
    embeddedfolderview endpoint (the modern folder page is JS-only)."""
    url = f"https://drive.google.com/embeddedfolderview?id={folder_id}#list"
    html_text = req.get(url, timeout=60).text
    return re.findall(
        r'id="entry-(1[A-Za-z0-9_-]{15,})".*?flip-entry-title">([^<]+)<',
        html_text,
        re.S,
    )


def _download_drive_bytes(file_id: str, max_bytes: int) -> bytes | None:
    """Download ONE public Drive file, None on failure/HTML stub."""
    for url in (
        f"https://drive.usercontent.google.com/download?id={file_id}&export=download&confirm=t",
        f"https://drive.google.com/uc?export=download&id={file_id}&confirm=t",
    ):
        try:
            with req.get(url, stream=True, timeout=300) as resp:
                resp.raise_for_status()
                chunks = resp.iter_content(chunk_size=512 * 1024)
                first = next(chunks, b"")
                if first.lstrip()[:1] in (b"<", b""):
                    continue  # HTML-заглушка (нет доступа)
                buf = bytearray(first)
                for chunk in chunks:
                    buf.extend(chunk)
                    if len(buf) > max_bytes:
                        raise ValueError("file too big")
                return bytes(buf)
        except Exception:
            continue
    return None


def process_drive_folder(chat_id, folder_id: str):
    """Drive FOLDER link in the control group: enumerate the folder and
    run every file (and every ZIP inside it) through the packing
    pipeline. RAR/7z are reported as unsupported."""
    import zipfile

    try:
        with TypingIndicator(chat_id):
            entries = _list_drive_folder(folder_id)
            if not entries:
                telegram_send_message(
                    chat_id,
                    "⚠️ Papkada fayl topilmadi yoki ruxsat yo'q. "
                    "Papkaga \"Anyone with the link\" ruxsati berilganini tekshiring.",
                )
                return
            index, rows = _build_active_bl_index()
            chat_titles = _build_chat_title_lookup(rows)
            os.makedirs(UPLOAD_FOLDER, exist_ok=True)
            results = _new_packing_results()
            for file_id, name in entries[:_PACKING_ZIP_MAX_FILES]:
                base = str(name or "").strip()
                low = base.lower()
                if low.endswith((".rar", ".7z")):
                    results["unsupported"].append(base)
                    continue
                if low.endswith(".zip"):
                    data = _download_drive_bytes(file_id, _PACKING_URL_MAX_BYTES)
                    if data is None:
                        results["rejected"].append(base + " (yuklab bo'lmadi)")
                        continue
                    try:
                        _ingest_zip_source(io.BytesIO(data), index, rows, chat_titles, results)
                    except zipfile.BadZipFile:
                        results["rejected"].append(base + " (buzilgan arxiv)")
                    continue
                data = _download_drive_bytes(file_id, _PACKING_FILE_MAX_BYTES)
                if data is None:
                    results["rejected"].append(base + " (yuklab bo'lmadi)")
                    continue
                _attach_packing_bytes(base, data, index, rows, chat_titles, results)
            _send_packing_report(chat_id, results, empty_note="ℹ️ Papkada biriktiradigan fayl topilmadi.")
    except Exception as exc:
        app.logger.exception("Drive folder processing failed")
        try:
            telegram_send_message(chat_id, f"⚠️ Papka tahlilida xato: {exc}", parse_mode=None)
        except Exception:
            pass


def process_packing_zip_url(chat_id, kind: str, ref: str):
    import tempfile

    try:
        with TypingIndicator(chat_id):
            if kind == "drive":
                candidates = [
                    f"https://drive.usercontent.google.com/download?id={ref}&export=download&confirm=t",
                    f"https://drive.google.com/uc?export=download&id={ref}&confirm=t",
                ]
            else:
                candidates = [ref]
            tmp_path = None
            last_err = ""
            for url in candidates:
                try:
                    with req.get(url, stream=True, timeout=600) as resp:
                        resp.raise_for_status()
                        chunks = resp.iter_content(chunk_size=512 * 1024)
                        first = next(chunks, b"")
                        if not first.startswith(b"PK"):
                            # HTML-заглушка Google (нет доступа) или не ZIP
                            last_err = "havola ZIP faylga olib bormaydi (ruxsat yo'q bo'lishi mumkin)"
                            continue
                        fd, tmp_path = tempfile.mkstemp(suffix=".zip")
                        total = 0
                        with os.fdopen(fd, "wb") as out:
                            out.write(first)
                            total += len(first)
                            for chunk in chunks:
                                total += len(chunk)
                                if total > _PACKING_URL_MAX_BYTES:
                                    raise ValueError("arxiv 512 MB dan katta")
                                out.write(chunk)
                        break
                    # unreachable
                except Exception as exc:
                    last_err = str(exc)
                    if tmp_path and os.path.exists(tmp_path):
                        try:
                            os.remove(tmp_path)
                        except OSError:
                            pass
                    tmp_path = None
                    continue
            if not tmp_path:
                telegram_send_message(
                    chat_id,
                    f"⚠️ Havoladan ZIP yuklab bo'lmadi: {last_err}\n"
                    "Google Drive'da faylga \"Anyone with the link\" ruxsati berilganini tekshiring.",
                    parse_mode=None,
                )
                return
            try:
                _ingest_packing_zip(chat_id, tmp_path)
            finally:
                try:
                    os.remove(tmp_path)
                except OSError:
                    pass
    except Exception as exc:
        app.logger.exception("Packing zip via URL failed")
        try:
            telegram_send_message(chat_id, f"⚠️ Havola ZIP tahlilida xato: {exc}", parse_mode=None)
        except Exception:
            pass


# ═══════════════════════════════════════════════════════════════
# MORNING TRACKING DIGEST (личка ответственного за трекинг)
# ═══════════════════════════════════════════════════════════════
# Каждое утро после того, как логисты обновят статусы в системе, бот
# шлёт пользователю TRACKING_DIGEST_TG_ID сводку по всем активным
# партиям. С 9:00 бот ждёт первого обновления статусов за сегодня; если
# к 13:00 обновлений так и нет — шлёт сводку с пометкой об этом.

TRACKING_DIGEST_TG_ID = os.getenv("TRACKING_DIGEST_TG_ID", "7713376668").strip()
TRACKING_DIGEST_HOUR = int(os.getenv("TRACKING_DIGEST_HOUR", "9") or 9)
_TRACKING_DIGEST_FALLBACK_HOUR = 13
_TRACKING_DIGEST_SETTING = "tracking_digest_last_date"


# ═══════════════════════════════════════════════════════════════
# TRACKING MINI APP (Telegram Web App «Параметры партии»)
# ═══════════════════════════════════════════════════════════════
# Утром бот присылает в управляющую группу список активных партий и
# кнопку, открывающую МИНИ-ПРИЛОЖЕНИЕ — страницу в стиле окна
# «Параметры партии» на сайте: сначала список активных партий, тап по
# партии открывает её форму (статус, ETA-текст, точка, дата выдачи).
# Авторизация: подпись Telegram initData + (админ/просмотровый id или
# членство в управляющей группе через getChatMember).

TRACKING_ASK_HOUR = int(os.getenv("TRACKING_ASK_HOUR", "7") or 7)
_TRACKING_ASK_SETTING = "tracking_ask_last_date"
_TGFORM_GROUPS_SETTING = "tgform_enabled_groups"
_FORM_MEMBER_CACHE: dict = {}


def _tgform_enabled_groups() -> set:
    """Группы, где Mini App активирован (/formon / /formoff).

    Пока настройка ни разу не менялась — по умолчанию управляющая
    группа. "__none__" = явно выключено везде."""
    from services import ai_assistant

    raw = (db.get_setting(_TGFORM_GROUPS_SETTING) or "").strip()
    if not raw:
        return {ai_assistant.control_group_id()}
    if raw == "__none__":
        return set()
    return {p.strip() for p in raw.split(",") if p.strip()}


def _save_tgform_groups(groups: set) -> None:
    db.set_setting(_TGFORM_GROUPS_SETTING, ",".join(sorted(groups)) if groups else "__none__")


def handle_tgform_toggle_command(message: dict, command: str) -> None:
    """/formon | /formoff в группе — активация/деактивация Mini App.
    Разрешено владельцу (admin ids) и администраторам самой группы."""
    from services import ai_assistant

    chat = message.get("chat") or {}
    chat_id = str(chat.get("id") or "")
    sender_id = str(((message.get("from") or {}).get("id")) or "")
    if sender_id not in ai_assistant.admin_ids() and sender_id not in get_chat_admin_ids(chat_id):
        telegram_send_message(chat_id, "❌ Bu buyruqni faqat guruh admini ishlatishi mumkin.")
        return
    if chat_id in CONFIDENTIAL_CHAT_IDS:
        telegram_send_message(chat_id, "❌ Bu guruhda forma yoqilmaydi.")
        return

    groups = _tgform_enabled_groups()
    enable = command in {"formon", "form_on"}
    if enable:
        groups.add(chat_id)
        _save_tgform_groups(groups)
        _FORM_MEMBER_CACHE.clear()
        telegram_send_message(
            chat_id,
            (
                "✅ Treking Mini App bu guruhda <b>YOQILDI</b>.\n"
                f"Har kuni ertalab soat {TRACKING_ASK_HOUR}:00 da yangilash so'rovi keladi. "
                "Guruh a'zolari formadan foydalana oladi:"
            ),
            reply_markup=_tgform_group_keyboard(chat_id),
        )
        # если раньше закрепляли сообщение формы — снимаем (закреп отменён)
        old_pin = db.get_setting(f"tgform_pin_{chat_id}")
        if old_pin:
            try:
                telegram_api("unpinChatMessage", json={"chat_id": chat_id, "message_id": int(old_pin)})
            except Exception:
                pass
            db.set_setting(f"tgform_pin_{chat_id}", "")
        # постоянная кнопка у поля ввода (reply-клавиатура)
        telegram_send_message(
            chat_id,
            "⬇️ Pastdagi doimiy «📝 Treking forma» tugmasi qo'shildi — istalgan vaqtda bosing. /form buyrug'i ham ishlaydi.",
            reply_markup=_tgform_reply_keyboard(),
        )
    else:
        groups.discard(chat_id)
        _save_tgform_groups(groups)
        _FORM_MEMBER_CACHE.clear()
        # снимаем закреп и постоянную кнопку
        pinned_id = db.get_setting(f"tgform_pin_{chat_id}")
        if pinned_id:
            try:
                telegram_api("unpinChatMessage", json={"chat_id": chat_id, "message_id": int(pinned_id)})
            except Exception:
                pass
            db.set_setting(f"tgform_pin_{chat_id}", "")
        telegram_send_message(
            chat_id,
            (
                "❌ Treking Mini App bu guruhda <b>O'CHIRILDI</b> — "
                "ertalabki so'rovlar kelmaydi, guruh a'zolarining forma ruxsati bekor qilindi."
            ),
            reply_markup=REMOVE_REPLY_MARKUP,
        )


def _tgform_url() -> str:
    base = (WEBHOOK_BASE_URL or "").rstrip("/")
    return f"{base}/tgform" if base else ""


# Короткое имя Mini App из BotFather (/newapp): ссылка t.me/<bot>/<name>
# открывает приложение ПРЯМО в группе (web_app-кнопки в группах Telegram
# запрещает — BUTTON_TYPE_INVALID).
TGFORM_APP_SHORTNAME = os.getenv("TGFORM_APP_SHORTNAME", "form").strip() or "form"


def _tgform_group_keyboard(chat_id) -> dict | None:
    """Кнопка для группы: direct-link Mini App со startapp=src<chat_id> —
    страница узнаёт, из какой группы её открыли, и шлёт подтверждения
    туда же."""
    username = get_bot_username() or ""
    if not username:
        return None
    url = f"https://t.me/{username}/{TGFORM_APP_SHORTNAME}?startapp=src{chat_id}"
    return {"inline_keyboard": [[{"text": "📝 Formani ochish", "url": url}]]}


def _resolve_src_group(src_chat_id) -> str:
    """Группа-источник формы для уведомлений/подтверждений: должна быть
    из включённых; иначе — управляющая группа."""
    from services import ai_assistant

    src = str(src_chat_id or "").strip()
    if src and src in _tgform_enabled_groups():
        return src
    return ai_assistant.control_group_id()


def send_tracking_form_button(chat_id):
    """Личная кнопка Web App (в группах Telegram их не разрешает)."""
    url = _tgform_url()
    if not url:
        telegram_send_message(chat_id, "⚠️ WEBHOOK_BASE_URL sozlanmagan — forma ochilmaydi.")
        return
    telegram_send_message(
        chat_id,
        "📝 Partiya parametrlarini yangilash formasi:",
        reply_markup={"inline_keyboard": [[
            {"text": "📝 Formani ochish", "web_app": {"url": url}}
        ]]},
    )


def send_tracking_update_request(force: bool = False, target_chat_id=None):
    """Утренний запрос в управляющую группу (или по требованию).

    В группах web_app-кнопки запрещены Telegram'ом, поэтому там кнопка-
    ссылка t.me/<bot>?start=form: тап открывает личку бота, где он сразу
    присылает web_app-кнопку формы."""
    from services import ai_assistant
    from html import escape as html_escape

    if not BOT_TOKEN:
        return False, "not configured"
    today = datetime.now(db.TASHKENT_TZ).strftime("%Y-%m-%d")
    if target_chat_id is None:
        if not force and db.get_setting(_TRACKING_ASK_SETTING) == today:
            return False, "already asked today"
        targets = sorted(_tgform_enabled_groups())
        if not targets:
            db.set_setting(_TRACKING_ASK_SETTING, today)
            return False, "mini app disabled in all groups"
    else:
        targets = [target_chat_id]
    batches = [b for b in db.get_batches() if not (b.get("client_delivery_date") or "")]
    if not batches:
        if target_chat_id is None:
            db.set_setting(_TRACKING_ASK_SETTING, today)
        else:
            telegram_send_message(target_chat_id, "ℹ️ Aktiv partiyalar yo'q.")
        return False, "no active batches"

    lines = [
        "🌅 Assalomu alaykum! Treking ma'lumotlarini yangilash vaqti bo'ldi.",
        "",
    ]
    for b in batches[:30]:
        lines.append(f"📦 <b>{html_escape(b['name'])}</b> — {html_escape(b['status'])}")
    lines.append("")
    lines.append("Quyidagi tugma orqali formani oching:")
    text = "\n".join(lines)

    for chat_id in targets:
        target_is_private = not str(chat_id).startswith("-")
        if target_is_private:
            keyboard = {"inline_keyboard": [[
                {"text": "📝 Formani ochish", "web_app": {"url": _tgform_url()}}
            ]]}
        else:
            # direct-link Mini App открывается прямо в группе
            keyboard = _tgform_group_keyboard(chat_id)
        try:
            telegram_send_message(chat_id, text, reply_markup=keyboard)
        except Exception:
            app.logger.exception("Tracking ask delivery failed for chat %s", chat_id)
    if target_chat_id is None:
        db.set_setting(_TRACKING_ASK_SETTING, today)
    return True, f"{len(batches)} batches → {len(targets)} chats"


def _validate_webapp_init_data(init_data: str):
    """Проверка подписи Telegram WebApp initData → user dict или None."""
    try:
        from urllib.parse import parse_qsl

        pairs = dict(parse_qsl(str(init_data or ""), keep_blank_values=True))
        their_hash = pairs.pop("hash", "")
        if not their_hash:
            return None
        data_check = "\n".join(f"{k}={v}" for k, v in sorted(pairs.items()))
        secret = hmac.new(b"WebAppData", BOT_TOKEN.encode(), hashlib.sha256).digest()
        calc = hmac.new(secret, data_check.encode(), hashlib.sha256).hexdigest()
        if not hmac.compare_digest(calc, their_hash):
            return None
        auth_date = int(pairs.get("auth_date") or 0)
        if auth_date and time.time() - auth_date > 86400:
            return None
        user = json.loads(pairs.get("user") or "{}")
        return user if user.get("id") else None
    except Exception:
        return None


def _webapp_user_allowed(user_id) -> bool:
    """Админы/просмотровые — сразу; остальные — по членству в ЛЮБОЙ
    группе, где Mini App активирован (getChatMember, кэш 10 минут)."""
    from services import ai_assistant

    uid = str(user_id)
    if ai_assistant.is_admin(uid) or ai_assistant.is_readonly_user(uid):
        return True
    now = time.time()
    cached = _FORM_MEMBER_CACHE.get(uid)
    if cached and cached[1] > now:
        return cached[0]
    allowed = False
    for group_id in _tgform_enabled_groups():
        try:
            resp = telegram_api(
                "getChatMember",
                json={"chat_id": group_id, "user_id": int(uid)},
            )
            status = ((resp.get("result") or {}).get("status")) or ""
            if status in {"creator", "administrator", "member"}:
                allowed = True
                break
        except Exception:
            continue
    _FORM_MEMBER_CACHE[uid] = (allowed, now + 600)
    return allowed


def _webapp_auth_or_403():
    data = request.json or {}
    user = _validate_webapp_init_data(data.get("init_data") or "")
    if not user:
        return None, (jsonify({"error": "auth"}), 403)
    if not _webapp_user_allowed(user.get("id")):
        return None, (jsonify({"error": "forbidden"}), 403)
    return user, None


@app.route("/tgform")
def tgform_page():
    return render_template("tgform.html")


@app.route("/tgform/api/list", methods=["POST"])
def tgform_api_list():
    user, err = _webapp_auth_or_403()
    if err:
        return err
    batches = [
        {
            "id": b["id"],
            "name": b["name"],
            "status": b["status"],
            "eta_to_toshkent": b.get("eta_to_toshkent") or "",
            "eta_destination": b.get("eta_destination") or "Toshkent",
            "incident_note": b.get("incident_note") or "",
            "bl_count": b.get("bl_count") or 0,
        }
        for b in db.get_batches()
        if not (b.get("client_delivery_date") or "")
    ]
    return jsonify({
        "ok": True,
        "user_name": (user.get("first_name") or ""),
        "batches": batches,
        "statuses": list(db.STATUSES),
        "destinations": [
            {"value": key, "label": label}
            for key, label in db.ETA_DESTINATION_LABELS.items()
        ],
    })


def _tgform_apply_save(user: dict, data: dict):
    """Сохранение параметров партии из мини-формы. Возвращает
    (batch_dict, None) либо (None, (json_response, http_code))."""
    from services import ai_assistant

    if ai_assistant.is_readonly_user((user or {}).get("id")):
        return None, (jsonify({"error": "Sizda faqat ko'rish huquqi bor"}), 403)
    try:
        batch_id = int(data.get("batch_id") or 0)
    except (TypeError, ValueError):
        return None, (jsonify({"error": "bad batch_id"}), 400)
    batch = db.get_batch(batch_id)
    if not batch:
        return None, (jsonify({"error": "Партия не найдена"}), 404)
    status = str(data.get("status") or batch.get("status") or "Xitoy").strip()
    valid_statuses = set(db.STATUSES) | {db.DELIVERED_STATUS, db.LEGACY_DELIVERED_STATUS}
    if status not in valid_statuses:
        return None, (jsonify({"error": "Недопустимый статус"}), 400)
    destination = str(data.get("eta_destination") or batch.get("eta_destination") or "Toshkent").strip()
    if destination not in db.ETA_DESTINATION_LABELS:
        destination = "Toshkent"
    eta_text = str(data.get("eta_to_toshkent") or "").strip()
    db.update_batch(
        batch_id, batch["name"], status, eta_text, destination,
        batch.get("client_delivery_date") or "",
    )
    if "incident_note" in data:
        db.set_batch_incident(batch_id, str(data.get("incident_note") or ""))
    app.logger.info(
        "TGFORM save by %s (%s): batch %s -> status=%r dest=%r eta=%r incident=%r",
        user.get("id"), user.get("first_name"), batch["name"], status, destination, eta_text,
        str(data.get("incident_note") or ""),
    )
    return db.get_batch(batch_id), None


@app.route("/tgform/api/save", methods=["POST"])
def tgform_api_save():
    from html import escape as html_escape

    user, err = _webapp_auth_or_403()
    if err:
        return err
    data = request.json or {}
    batch, save_err = _tgform_apply_save(user, data)
    if save_err:
        return save_err
    # Учёт: об изменении через форму сообщаем в группу, из которой
    # форма была открыта (или в управляющую).
    notify_chat = _resolve_src_group(data.get("src_chat_id"))
    sender_name = str(user.get("first_name") or user.get("id") or "").strip()
    incident = (batch.get("incident_note") or "").strip()
    try:
        telegram_send_message(
            notify_chat,
            (
                f"📝 <b>{html_escape(sender_name)}</b> partiya <b>{html_escape(batch['name'])}</b> "
                f"ma'lumotini yangiladi:\n"
                f"🚚 {html_escape(batch.get('status') or '')} · "
                f"📍 {html_escape(batch.get('eta_destination') or 'Toshkent')} · "
                f"⏱ {html_escape(batch.get('eta_to_toshkent') or '—')}"
                + (f"\n⚠️ {html_escape(incident)}" if incident else "")
            ),
            disable_notification=True,
        )
    except Exception:
        app.logger.exception("TGFORM save notify failed")
    return jsonify({"ok": True})


@app.route("/tgform/api/confirm", methods=["POST"])
def tgform_api_confirm():
    """Большая кнопка TASDIQLASH: сохраняет форму и запускает ВТОРОЕ
    подтверждение — бот пишет в управляющую группу карточку «Hammasi
    to'g'rimi?»; рассылка трекинга по клиентским группам уходит только
    после ✅ в группе (существующий approval-механизм)."""
    from html import escape as html_escape
    from services import ai_assistant

    user, err = _webapp_auth_or_403()
    if err:
        return err
    data = request.json or {}
    batch, save_err = _tgform_apply_save(user, data)
    if save_err:
        return save_err

    # Карточка «Hammasi to'g'rimi?» уходит в группу, из которой открыта
    # форма (если она из включённых), иначе — в управляющую.
    group_id = _resolve_src_group(data.get("src_chat_id"))
    summary = f"Разослать трекинг партии «{batch['name']}» по группам клиентов"
    action_id = db.ai_create_pending_action(
        f"tgform:{user.get('id')}",
        "send_tracking_batch",
        json.dumps({"batch_id": batch["id"], "batch_name": batch["name"]}, ensure_ascii=False),
        summary,
    )
    sender_name = str(user.get("first_name") or user.get("id") or "").strip()
    incident = (batch.get("incident_note") or "").strip()
    incident_line = f"⚠️ Kutilmagan vaziyat: <b>{html_escape(incident)}</b>\n" if incident else ""
    text = (
        f"📝 <b>{html_escape(sender_name)}</b> treking ma'lumotlarini to'ldirdi:\n\n"
        f"📦 Partiya: <b>{html_escape(batch['name'])}</b>\n"
        f"🚚 Holat: <b>{html_escape(batch.get('status') or '')}</b>\n"
        f"📍 Nuqta: {html_escape(batch.get('eta_destination') or 'Toshkent')}\n"
        f"⏱ ETA: {html_escape(batch.get('eta_to_toshkent') or '—')}\n"
        f"{incident_line}\n"
        "❓ <b>Hammasi to'g'rimi?</b> Tasdiqlansa, treking mijoz guruhlariga yuboriladi."
    )
    keyboard = {
        "inline_keyboard": [[
            {"text": "✅ TASDIQLAYMAN — yuborilsin", "callback_data": f"aiact:ok:{action_id}"},
            {"text": "❌ Yo'q", "callback_data": f"aiact:no:{action_id}"},
        ]]
    }
    telegram_send_message(group_id, text, reply_markup=keyboard)
    return jsonify({"ok": True})



def sync_batches_from_fura_statuses():
    """Auto-lifecycle from the «Fura statuslari» sheet:

    - BOJXONAGA TUSHDI (V) filled → batch arrived in Tashkent: if the
      live status hasn't reached an arrived state yet, set
      «Toshkent(Chuqursoy ULS da)» and write the V date into
      toshkent_arrived_at (same event, sheet vs live).
    - YUKLAR TARQATILDI (W) filled → batch delivered: set the delivery
      date and DELIVERED status — the batch automatically moves to the
      inactive list. The control group gets a notification.
    """
    from services import ai_assistant, fura_status_service

    data = fura_status_service.get_fura_statuses()
    if not data.get("ok"):
        return 0
    records = data["records"]
    changed = 0
    arrived_set = set(db.ARRIVED_STATUSES)
    for batch in db.get_batches():
        digits = re.sub(r"\D", "", str(batch.get("name") or ""))
        if len(digits) != 8:
            continue
        rec = records.get(digits)
        if not rec:
            continue
        was_active = not (batch.get("client_delivery_date") or "").strip()

        # W: дата выдачи клиентам. Шитс — ИСТОЧНИК ИСТИНЫ, в том числе для
        # уже закрытых партий (вручную вбитые даты корректируются).
        if rec["tarqatildi"]:
            delivery = rec["tarqatildi"].strftime("%d.%m.%Y")
            if (batch.get("client_delivery_date") or "").strip() != delivery:
                new_status = db.DELIVERED_STATUS if was_active else (batch.get("status") or db.DELIVERED_STATUS)
                db.update_batch(
                    batch["id"], batch["name"], new_status,
                    batch.get("eta_to_toshkent") or "",
                    batch.get("eta_destination") or "Toshkent",
                    delivery,
                )
                changed += 1
                if was_active:
                    # уведомляем только о ПЕРЕХОДЕ в неактивные, не о правках дат
                    try:
                        telegram_send_message(
                            ai_assistant.control_group_id(),
                            (
                                f"📦 Partiya <b>{batch['name']}</b> yakunlandi: "
                                f"Yuklar tarqatildi — {delivery} (Fura statuslari jadvalidan). "
                                "Partiya avtomatik NOAKTIV bo'ldi."
                            ),
                        )
                    except Exception:
                        app.logger.exception("Fura-status deactivation notify failed")
        elif was_active and rec["bojxona"] and batch.get("status") not in arrived_set:
            # прибыла в Ташкент, но ещё не выдана
            db.update_batch(
                batch["id"], batch["name"], "Toshkent(Chuqursoy ULS da)",
                batch.get("eta_to_toshkent") or "",
                batch.get("eta_destination") or "Toshkent",
                "",
            )
            changed += 1

        # V: дата прибытия в Ташкент. Тоже источник истины — исправляем,
        # если отличается (в т.ч. штамп «вживую», поставленный логистами
        # раньше/позже реальной даты из шитса).
        if rec["bojxona"]:
            iso_arr = rec["bojxona"].strftime("%Y-%m-%d") + " 00:00:00"
            if (batch.get("toshkent_arrived_at") or "").strip()[:10] != iso_arr[:10]:
                conn = db.get_conn()
                try:
                    conn.execute(
                        "UPDATE batches SET toshkent_arrived_at = ? WHERE id = ?",
                        (iso_arr, batch["id"]),
                    )
                    conn.commit()
                finally:
                    conn.close()
                changed += 1
    return changed


def send_tracking_digest(force: bool = False):
    from html import escape as html_escape

    if not TRACKING_DIGEST_TG_ID or not BOT_TOKEN:
        return False, "not configured"
    now = datetime.now(db.TASHKENT_TZ)
    today = now.strftime("%Y-%m-%d")
    if not force and db.get_setting(_TRACKING_DIGEST_SETTING) == today:
        return False, "already sent today"

    batches = [b for b in db.get_batches() if not (b.get("client_delivery_date") or "")]
    if not batches:
        db.set_setting(_TRACKING_DIGEST_SETTING, today)
        return False, "no active batches"

    def _updated_today(b) -> bool:
        return str(b.get("status_updated_at") or "")[:10] == today

    updated = [b for b in batches if _updated_today(b)]
    if not force and not updated and now.hour < _TRACKING_DIGEST_FALLBACK_HOUR:
        # логисты ещё не обновили статусы — ждём следующего тика
        return False, "waiting for status updates"

    lines = [f"🚚 <b>Treking ma'lumoti — {now.strftime('%d.%m.%Y')}</b>", ""]
    for b in batches:
        mark = "✅" if _updated_today(b) else "⏳"
        lines.append(f"📦 <b>{html_escape(b.get('name') or '')}</b> — {html_escape(b.get('status') or '')} {mark}")
        eta = (b.get("eta_to_toshkent") or "").strip()
        dest = (b.get("eta_destination") or "").strip()
        detail = []
        if eta:
            detail.append(f"ETA: {html_escape(eta)}")
        if dest:
            detail.append(html_escape(dest))
        if detail:
            lines.append("   " + " · ".join(detail))
        incident = (b.get("incident_note") or "").strip()
        if incident:
            lines.append(f"   ⚠️ {html_escape(incident)}")
    lines.append("")
    lines.append("✅ — status bugun yangilangan · ⏳ — hali yangilanmagan")
    if not updated:
        lines.append("⚠️ Logistlar bugun statuslarni hali yangilashmadi.")
    telegram_send_message(TRACKING_DIGEST_TG_ID, "\n".join(lines))
    db.set_setting(_TRACKING_DIGEST_SETTING, today)
    return True, f"{len(batches)} batches ({len(updated)} updated today)"


def handle_my_chat_member_update(chat_update: dict):
    chat = chat_update.get("chat") or {}
    chat_id = chat.get("id")
    chat_type = chat.get("type")
    if chat_type not in {"group", "supergroup"}:
        return

    old_status = ((chat_update.get("old_chat_member") or {}).get("status") or "").lower()
    new_status = ((chat_update.get("new_chat_member") or {}).get("status") or "").lower()
    is_active = new_status not in {"left", "kicked"}
    remember_group_chat(chat, is_active=is_active)
    if not chat_id:
        return
    if not is_active:
        clear_group_reply_keyboard(chat_id)
        return
    if new_status in {"member", "administrator"} and old_status in {"", "left", "kicked"}:
        return


def _send_single_bl_message(bl: dict, batch_name: str) -> tuple[bool, str]:
    """Render and send tracking for a single BL (no merging with other batches).

    Right after the message goes out we also deliver every packing list
    attached to this BL so the client receives "info → files → info →
    files" in strict batch order instead of all infos first and then a
    bucket of files at the end.
    """
    chat_id = bl.get("chat_id")
    if not chat_id:
        return False, "Нет chat_id"

    language = normalize_message_language(bl.get("message_language"))
    rendered_message = db.render_message(bl, batch_name, include_related_batches=False)

    delivered = False
    last_error = ""
    sent_response = None
    try:
        sent_response = send_with_track_keyboard(
            chat_id,
            rendered_message,
            language=language,
        )
        db.record_tracking_delivery(bl, include_related_batches=False)
        delivered = True
    except Exception as exc:
        last_error = str(exc)
        try:
            fallback_message = _plain_text_message(rendered_message)
            sent_response = send_with_track_keyboard(
                chat_id,
                fallback_message,
                language=language,
                parse_mode=None,
            )
            db.record_tracking_delivery(bl, include_related_batches=False)
            delivered = True
            last_error = ""
        except Exception as fallback_exc:
            return False, str(fallback_exc or exc)

    # Remember the message_id so the operator can recall it later via
    # /api/batches/<id>/recall-tracking. Best-effort — never blocks send.
    if delivered:
        try:
            msg_id = _extract_message_id(sent_response)
            if msg_id:
                db.record_sent_telegram_message(
                    bl_id=bl.get("id"),
                    batch_id=bl.get("batch_id"),
                    chat_id=chat_id,
                    message_id=msg_id,
                    kind="tracking",
                )
        except Exception:
            app.logger.exception("Failed to record tracking message id")

    # Auto-attach packing lists right after the message, before moving on
    # to the next batch in the bundle.
    if delivered:
        try:
            _send_attached_files_for_bl(chat_id, bl.get("id"))
        except Exception:
            app.logger.exception("Auto file delivery failed for bl_id=%s", bl.get("id"))

    return True, last_error


def send_bl_package(bl: dict, batch_name: str, include_related_batches: bool = True):
    """Send tracking messages.

    Even when `include_related_batches=True`, each BL/batch is sent as its own
    separate Telegram message (the "send together" trigger still discovers all
    related batches for this client, but they are no longer merged into one).
    """
    if not bl["chat_id"]:
        return False, "Нет chat_id"

    # Build the full bundle (primary + related batches if requested)
    bundle = db.get_tracking_bundle_bls(bl, include_related_batches=include_related_batches)
    if not bundle:
        bundle = [bl]

    # Resolve batch name per item (related items may belong to other batches)
    batch_name_cache: dict[int, str] = {}
    if bl.get("batch_id"):
        batch_name_cache[int(bl["batch_id"])] = batch_name or ""

    last_error = ""
    success_any = False
    for index, item in enumerate(bundle):
        item_batch_id = item.get("batch_id")
        item_batch_name = ""
        if item_batch_id:
            cached = batch_name_cache.get(int(item_batch_id))
            if cached is None:
                related_batch = db.get_batch(int(item_batch_id)) or {}
                cached = related_batch.get("name") or ""
                batch_name_cache[int(item_batch_id)] = cached
            item_batch_name = cached
        if not item_batch_name:
            item_batch_name = batch_name or ""

        ok, err = _send_single_bl_message(item, item_batch_name)
        if ok:
            success_any = True
        else:
            last_error = err
        # Small pause between messages helps Telegram avoid flood limits and
        # keeps the per-batch messages visually separated in the chat.
        if index < len(bundle) - 1:
            try:
                time.sleep(0.4)
            except Exception:
                pass

    if not success_any:
        return False, last_error or "Не удалось отправить сообщения"
    return True, last_error


@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        auth_user = get_auth_users().get(username)
        role = auth_user["role"] if auth_user else ""
        if auth_user and password == auth_user["password"]:
            session.clear()
            session["logged_in"] = True
            session["username"] = username
            session["role"] = role
            db.record_login_history(
                username=username,
                role=role,
                success=True,
                ip_address=get_request_ip(),
                user_agent=request.headers.get("User-Agent", ""),
            )
            # Kiosk users go straight to the full-screen Sales Monitor for TV display
            if role == ROLE_KIOSK:
                return redirect(url_for("analytics_monitor_page"))
            return redirect(url_for("index"))
        db.record_login_history(
            username=username,
            role=role,
            success=False,
            ip_address=get_request_ip(),
            user_agent=request.headers.get("User-Agent", ""),
        )
        error = "Invalid login or password"
    return render_template("index.html", login_page=True, error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/")
@login_required
def index():
    # Kiosk profile (TV display) is restricted to the Sales Monitor only
    if session.get("role") == ROLE_KIOSK:
        return redirect(url_for("analytics_monitor_page"))
    return render_template(
        "index.html",
        initial_view=(request.args.get("view") or "dashboard").strip() or "dashboard",
        initial_analytics_tab=(request.args.get("tab") or "monitor").strip() or "monitor",
    )


def _render_analytics_page(tab: str):
    return render_template("index.html", initial_view="analytics", initial_analytics_tab=tab)


@app.route("/analytics")
@login_required
def analytics_index():
    return _render_analytics_page("monitor")


@app.route("/analytics/sales-growth")
@login_required
def analytics_sales_growth_page():
    return _render_analytics_page("sales-growth")


@app.route("/analytics/cashflow")
@login_required
def analytics_cashflow_page():
    return _render_analytics_page("cashflow")


@app.route("/analytics/managers")
@login_required
def analytics_managers_page():
    return _render_analytics_page("managers")


@app.route("/analytics/logists")
@login_required
def analytics_logists_page():
    return _render_analytics_page("logists")


@app.route("/analytics/shipments")
@login_required
def analytics_shipments_page():
    return _render_analytics_page("shipments")


@app.route("/analytics/debts")
@login_required
def analytics_debts_page():
    return _render_analytics_page("debts")


@app.route("/analytics/export")
@login_required
def analytics_export_page():
    return _render_analytics_page("export")


@app.route("/analytics/sync")
@login_required
def analytics_sync_page():
    return _render_analytics_page("sync")


@app.route("/analytics/monitor")
@login_required
def analytics_monitor_page():
    plans = analytics_service.list_sales_plans()
    active_plan = next((plan for plan in plans if int(plan.get("is_active") or 0) == 1), None)
    return render_template(
        "analytics/monitor.html",
        sales_plans=plans,
        active_plan_id=(active_plan or {}).get("id"),
        is_kiosk=(session.get("role") == ROLE_KIOSK),
    )


@app.route("/analytics/monitor/legacy")
@login_required
def analytics_monitor_legacy():
    """Frozen snapshot of the pre-redesign (neon-green) Sales Monitor.
    Uses sales_monitor_legacy.css — a byte-for-byte copy of the stylesheet
    taken before the Glass Orbit theme landed — with the same live JS, so
    it still shows real data. To fully restore the old design: copy
    sales_monitor_legacy.css over sales_monitor.css and revert the
    gradient stops + font link in monitor.html."""
    plans = analytics_service.list_sales_plans()
    active_plan = next((plan for plan in plans if int(plan.get("is_active") or 0) == 1), None)
    return render_template(
        "analytics/monitor_legacy.html",
        sales_plans=plans,
        active_plan_id=(active_plan or {}).get("id"),
        is_kiosk=(session.get("role") == ROLE_KIOSK),
    )


@app.route("/analytics/monitor/preview")
def analytics_monitor_preview():
    """Side-by-side preview of 10 candidate 3D circle styles for the Sales Monitor.
    Public — no login required (pure CSS demo, no real data)."""
    return render_template("analytics/monitor_preview.html")


@app.route("/analytics/monitor/redesign")
def analytics_monitor_redesign():
    """10 full-page redesign variants for the Sales Monitor with a style
    switcher. Static mock data only — safe to share for design review.
    Public — no login required."""
    return render_template("analytics/monitor_redesign.html")


@app.route("/analytics/monitor/redesign2")
def analytics_monitor_redesign2():
    """Round 2: 10 variants built on the user's picks from round 1
    (Glass / Aurora / Split Bold) with 3D depth and smooth motion —
    animated blobs, spinning conic rings, liquid fill, orbiting
    particles, count-up numbers, ticker. Public — no login required."""
    return render_template("analytics/monitor_redesign2.html")


@app.route("/analytics/monitor/redesign3")
def analytics_monitor_redesign3():
    """Round 3: the 4 finalists the user picked from round 2 (Glass Orbit,
    Aurora Flow, Liquid, Pulse Grid — marked with a star) plus 6 new
    variants in the same 3D/motion family (Hologram, Wave Horizon,
    Crystal, Starfield, Infinity, Nebula). Public — no login required."""
    return render_template("analytics/monitor_redesign3.html")


@app.route("/analytics/monitor/redesign4")
def analytics_monitor_redesign4():
    """Round 4: 5 finalists (Glass Orbit, Aurora Flow, Liquid, Pulse Grid,
    Infinity) plus 5 new animation mechanics (Radar sweep, EKG pulse line,
    rising bubbles, highway with driving truck, dancing equalizer bars).
    Public — no login required."""
    return render_template("analytics/monitor_redesign4.html")


@app.route("/analytics/monitor/preview2")
def analytics_monitor_preview2():
    """Round 2 of preview ideas — 8 fresh scenes with enhanced breathing.
    Public — no login required."""
    return render_template("analytics/monitor_preview2.html")


@app.route("/analytics/monitor/preview3")
def analytics_monitor_preview3():
    """Round 3 of preview ideas — 10 brand-new scenes (volcanic, galaxy, pillars,
    waves, ripples, neon tube, prism, matrix rain, turbine, hologram).
    Public — no login required."""
    return render_template("analytics/monitor_preview3.html")


@app.route("/health")
def health():
    return jsonify(
        {
            "ok": True,
            "bot_configured": bool(BOT_TOKEN),
            "webhook_configured": bool(BOT_TOKEN and WEBHOOK_BASE_URL),
        }
    )


@app.route("/telegram/webhook", methods=["POST"])
def telegram_webhook():
    if WEBHOOK_SECRET:
        incoming_secret = request.headers.get("X-Telegram-Bot-Api-Secret-Token", "")
        if not secrets.compare_digest(incoming_secret, WEBHOOK_SECRET):
            return jsonify({"ok": False, "error": "invalid webhook secret"}), 403

    update = request.get_json(silent=True) or {}
    callback_query = update.get("callback_query")
    if callback_query:
        handle_callback_query(callback_query)

    chat_update = update.get("my_chat_member")
    if chat_update:
        handle_my_chat_member_update(chat_update)

    message = update.get("message") or update.get("edited_message")
    if message:
        track_moderator_response_metrics(message)
        handle_telegram_message(message)
    return jsonify({"ok": True})


@app.route("/api/stats")
@login_required
def api_stats():
    return jsonify(db.get_stats())


@app.route("/api/remove-track-keyboards", methods=["POST"])
@editor_required
def api_remove_track_keyboards():
    """One-time sweep: strip the retired "Yuk holati" reply keyboard from
    every chat the bot knows (groups + privates). Per chat: send an
    invisible message with remove_keyboard, then delete it — the cached
    keyboard disappears without leaving a visible message."""
    conn = db.get_conn()
    try:
        rows = conn.execute(
            """
            SELECT DISTINCT chat_id FROM (
                SELECT chat_id FROM telegram_chats
                UNION SELECT chat_id FROM telegram_sessions
                UNION SELECT chat_id FROM bl_codes WHERE chat_id != ''
                UNION SELECT chat_id FROM send_logs WHERE chat_id != ''
            ) WHERE COALESCE(chat_id, '') != ''
            """
        ).fetchall()
    finally:
        conn.close()
    chat_ids = [str(r["chat_id"]).strip() for r in rows if str(r["chat_id"]).strip()]

    def _sweep():
        ok = 0
        fail = 0
        for cid in chat_ids:
            try:
                resp = telegram_send_message(
                    cid,
                    "ㅤ",
                    reply_markup=REMOVE_REPLY_MARKUP,
                    parse_mode=None,
                    disable_notification=True,
                )
                mid = ((resp or {}).get("result") or {}).get("message_id")
                if mid:
                    try:
                        telegram_delete_message(cid, mid)
                    except Exception:
                        pass
                    ok += 1
                else:
                    fail += 1
            except Exception:
                fail += 1
            time.sleep(0.08)
        app.logger.info("Track-keyboard sweep done: ok=%s fail=%s of %s", ok, fail, len(chat_ids))
        from services import ai_assistant

        for admin_id in ai_assistant.admin_ids():
            try:
                telegram_send_message(
                    admin_id,
                    f"🧹 Кнопка Yuk holati убрана: ✅ {ok} чатов, ❌ {fail} недоступно (бот удалён из чата и т.п.).",
                )
            except Exception:
                pass

    threading.Thread(target=_sweep, daemon=True).start()
    return jsonify({"ok": True, "targets": len(chat_ids)})


@app.route("/api/late-cargo")
@login_required
def api_late_cargo():
    from services import late_cargo_service

    force = request.args.get("force") == "1"
    return jsonify(late_cargo_service.get_late_cargo_report(force=force))


@app.route("/api/dashboard/history/clear", methods=["POST"])
@editor_required
def api_clear_dashboard_history():
    deleted_count = db.clear_dashboard_history()
    return jsonify({"ok": True, "deleted_count": deleted_count})


@app.route("/api/attention")
@login_required
def api_attention():
    limit = int(request.args.get("limit", 10))
    return jsonify(db.get_attention_items(limit))


@app.route("/api/batches")
@login_required
def api_batches():
    return jsonify(db.get_batches())


@app.route("/api/batches", methods=["POST"])
@editor_required
def api_create_batch():
    data = request.json or {}
    name = (data.get("name") or "").strip()
    status = (data.get("status") or "Xitoy").strip() or "Xitoy"
    eta_to_toshkent = (data.get("eta_to_toshkent") or "").strip()
    eta_destination = (data.get("eta_destination") or "Toshkent").strip() or "Toshkent"
    client_delivery_date = (data.get("client_delivery_date") or "").strip()
    if not name:
        return jsonify({"error": "Имя партии обязательно"}), 400
    if not db.create_batch(name, status, eta_to_toshkent, eta_destination, client_delivery_date):
        return jsonify({"error": "Не удалось создать партию"}), 400
    return jsonify({"ok": True})


@app.route("/api/batches/<int:batch_id>", methods=["PUT"])
@editor_required
def api_update_batch(batch_id):
    data = request.json or {}
    name = (data.get("name") or "").strip()
    status = (data.get("status") or "Xitoy").strip() or "Xitoy"
    client_delivery_date = (data.get("client_delivery_date") or "").strip()
    if not name:
        return jsonify({"error": "Имя партии обязательно"}), 400
    if not db.update_batch(
        batch_id,
        name,
        status,
        (data.get("eta_to_toshkent") or "").strip(),
        (data.get("eta_destination") or "Toshkent").strip() or "Toshkent",
        client_delivery_date,
    ):
        return jsonify({"error": "Не удалось обновить партию"}), 400
    return jsonify({"ok": True})


@app.route("/api/batches/<int:batch_id>", methods=["DELETE"])
@editor_required
def api_delete_batch(batch_id):
    db.delete_batch(batch_id)
    return jsonify({"ok": True})


@app.route("/api/batches/<int:batch_id>/bl")
@login_required
def api_bl_list(batch_id):
    batch = db.get_batch(batch_id)
    if not batch:
        abort(404)
    bl_codes = db.get_bl_by_batch(batch_id)
    return jsonify({"batch": batch, "bl_codes": bl_codes, "statuses": db.STATUSES})


@app.route("/api/chats")
@login_required
def api_chats():
    include_inactive = request.args.get("all") == "1"
    return jsonify(db.get_telegram_chats(include_inactive=include_inactive))


@app.route("/api/chats/config")
@login_required
def api_chats_config():
    return jsonify({"global_ai_enabled": db.get_global_ai_enabled()})


@app.route("/api/chats/<chat_id>/toggle-ai", methods=["POST"])
@editor_required
def api_toggle_chat_ai(chat_id):
    enabled = db.toggle_chat_ai_enabled(chat_id)
    return jsonify({"ok": True, "chat_id": str(chat_id), "ai_enabled": enabled})


@app.route("/api/settings/toggle-global-ai", methods=["POST"])
@editor_required
def api_toggle_global_ai():
    enabled = db.toggle_global_ai_enabled()
    return jsonify({"ok": True, "global_ai_enabled": enabled})


@app.route("/api/google-sheets/config")
@login_required
def api_google_sheets_config():
    return jsonify({"url": db.get_setting(GOOGLE_SHEETS_URL_SETTING_KEY, "")})


@app.route("/api/google-sheets/config", methods=["POST"])
@editor_required
def api_save_google_sheets_config():
    data = request.json or {}
    url = (data.get("url") or "").strip()
    db.set_setting(GOOGLE_SHEETS_URL_SETTING_KEY, url)
    return jsonify({"ok": True, "url": url})


@app.route("/api/google-sheets/preview", methods=["POST"])
@login_required
def api_google_sheets_preview():
    data = request.json or {}
    url = (data.get("url") or "").strip() or db.get_setting(GOOGLE_SHEETS_URL_SETTING_KEY, "")
    if not url:
        return jsonify({"error": "Сначала укажи ссылку на Google Sheets"}), 400
    try:
        rows = _parse_google_sheet_rows(url)
    except req.RequestException as exc:
        return jsonify({"error": f"Не удалось прочитать Google Sheets: {exc}"}), 400
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    return jsonify(
        {
            "ok": True,
            "url": url,
            "rows": rows,
            "count": len(rows),
            "sheet_dates": sorted({row.get("sheet_date", "") for row in rows if row.get("sheet_date")}),
        }
    )


# Generic stop-words for chat-title token extraction (used by the sheet
# import's auto-chat-match). These appear in nearly every group title and
# would otherwise route every BL to the first chat seen.
_CHAT_TITLE_STOPWORDS = {
    "BURAQ", "LOGISTICS", "LOGISTIC", "BURAQLOGISTICS",
    "COMPANY", "LLC", "LTD", "INC", "GROUP", "TRADE", "TRADING",
}


def _build_chat_code_lookup() -> dict[str, str]:
    """Map normalized identifier tokens → chat_id for every Telegram chat.

    Example, chat title "HI-TECH (BL-146) & BURAQ LOGISTICS" registers:
        {"HITECH": cid, "HI": cid, "TECH": cid, "BL146": cid, "146": cid}
    So an incoming BL row with `code` = "BL-146" / "146" / "HI-TECH" all
    auto-pick the same chat. Stopwords filter out "BURAQ"/"LOGISTICS"
    so we don't accidentally route every BL to the first chat in the list.

    First chat wins on collision (stable insertion order).
    """
    lookup: dict[str, str] = {}
    try:
        chats = db.get_telegram_chats(include_inactive=True) or []
    except Exception:
        return lookup

    for chat in chats:
        chat_id = str(chat.get("chat_id") or "").strip()
        title = str(chat.get("title") or "").strip()
        if not chat_id or not title:
            continue
        upper = title.upper()
        # Individual alphanumeric tokens — only register tokens 3+ chars to
        # avoid noise from "BL", "HI", "DV" etc. fragments. The 2-char
        # exclusion stops "BL" from gobbling every incoming BL code.
        for tok in re.findall(r"[A-Za-z0-9]+", upper):
            if len(tok) < 3 or tok in _CHAT_TITLE_STOPWORDS:
                continue
            lookup.setdefault(tok, chat_id)
        # Concatenated form of the "meaningful" identifier — strips BL-NNN
        # and stopwords, then joins what's left. This catches multi-word
        # client names that appear hyphenated or space-separated in the
        # chat title, e.g.:
        #   "HI-TECH (BL-146) & BURAQ LOGISTICS"   -> "HITECH"
        #   "LUX LIGHTING (BL-234) & BURAQ ..."    -> "LUXLIGHTING"
        # So incoming codes "HI-TECH" or "LUX LIGHTING" auto-pick the chat.
        cleaned = re.sub(r"\bBL[-_ ]?\d+[A-Z0-9]*\b", " ", upper)
        meaningful = [
            t for t in re.findall(r"[A-Za-z0-9]+", cleaned)
            if len(t) >= 2 and t not in _CHAT_TITLE_STOPWORDS
        ]
        if meaningful:
            concat = "".join(meaningful)
            if len(concat) >= 3:
                lookup.setdefault(concat, chat_id)
        # BL-NNN patterns get registered in three forms so a BL code typed
        # any of these ways auto-matches: "BL-146", "BL146", "146".
        for m in re.finditer(r"BL[-_ ]?(\d+)", upper):
            number = m.group(1)
            lookup.setdefault(f"BL{number}", chat_id)
            lookup.setdefault(number, chat_id)
    return lookup


def _known_chat_id_set() -> set[str]:
    """Chat ids the bot has ever seen — history matches are only applied
    when the remembered chat still exists in telegram_chats."""
    try:
        chats = db.get_telegram_chats(include_inactive=True) or []
    except Exception:
        return set()
    return {
        str(chat.get("chat_id") or "").strip()
        for chat in chats
        if str(chat.get("chat_id") or "").strip()
    }


def _find_chat_for_bl_code(code: str, lookup: dict[str, str]) -> str:
    """Return best-match chat_id for a BL code, or '' if nothing fits."""
    if not code or not lookup:
        return ""
    norm = _normalize_bl_code(code)  # "BL-146" → "BL146", "ARTE" → "ARTE"
    if not norm:
        return ""
    # Exact normalized match (covers BL-146, BL146, ARTE, 5077, ...).
    if norm in lookup:
        return lookup[norm]
    # "BL146" → try "146" too.
    if norm.startswith("BL") and norm[2:].isdigit():
        stripped = norm[2:]
        if stripped in lookup:
            return lookup[stripped]
    # Pure digits "146" → try "BL146" too.
    if norm.isdigit():
        if f"BL{norm}" in lookup:
            return lookup[f"BL{norm}"]
    return ""


@app.route("/api/batches/<int:batch_id>/import-sheet", methods=["POST"])
@editor_required
def api_import_google_sheet_rows(batch_id):
    batch = db.get_batch(batch_id)
    if not batch:
        abort(404)
    data = request.json or {}
    rows = data.get("rows") or []
    if not isinstance(rows, list) or not rows:
        return jsonify({"error": "Не выбраны строки для импорта"}), 400

    # Auto-chat-match: build lookup once per request so we don't query the
    # chats table per BL.
    chat_lookup = _build_chat_code_lookup()
    known_chat_ids = _known_chat_id_set()

    imported = []
    skipped = []
    auto_matched = 0
    history_matched = 0
    for row in rows:
        code = str((row or {}).get("code") or "").strip()
        if not code:
            continue
        # Auto-find the Telegram chat whose title contains this BL code.
        # Falls back to the permanent link memory (this BL was linked to
        # a group in some earlier batch). Empty string if neither knows
        # the code — operator fills it manually later.
        chat_id = _find_chat_for_bl_code(code, chat_lookup)
        matched_via = "title" if chat_id else ""
        if not chat_id:
            hist = db.lookup_bl_link_history(code)
            if hist and (not known_chat_ids or hist["chat_id"] in known_chat_ids):
                chat_id = hist["chat_id"]
                matched_via = "history"
        success = db.add_bl(
            batch_id=batch_id,
            code=code,
            client_name="",
            chat_id=chat_id,
            cargo_type="",
            weight_kg=(row or {}).get("weight_kg", 0),
            volume_cbm=(row or {}).get("volume_cbm", 0),
            quantity_places=(row or {}).get("quantity_places", 0),
            quantity_places_breakdown=(row or {}).get("quantity_places_display", ""),
            cargo_description="",
            message_language=getattr(db, "DEFAULT_MESSAGE_LANGUAGE", "uz_latn"),
        )
        if success:
            imported.append(code)
            if matched_via == "title":
                auto_matched += 1
            elif matched_via == "history":
                history_matched += 1
        else:
            skipped.append({"code": code, "reason": "duplicate"})

    return jsonify(
        {
            "ok": True,
            "imported_count": len(imported),
            "skipped_count": len(skipped),
            "auto_matched_count": auto_matched,
            "history_matched_count": history_matched,
            "imported": imported,
            "skipped": skipped,
        }
    )


@app.route("/api/batches/<int:batch_id>/relink-from-history", methods=["POST"])
@editor_required
def api_relink_bls_from_history(batch_id):
    """Attach groups to the batch's unlinked BLs: first by chat-title
    match, then by the permanent BL↔group link memory built from all
    past batches."""
    batch = db.get_batch(batch_id)
    if not batch:
        abort(404)
    chat_lookup = _build_chat_code_lookup()
    known_chat_ids = _known_chat_id_set()
    try:
        chats_by_id = {
            str(c.get("chat_id") or "").strip(): str(c.get("title") or "")
            for c in (db.get_telegram_chats(include_inactive=True) or [])
        }
    except Exception:
        chats_by_id = {}

    linked = []
    unlinked_left = 0
    title_count = 0
    history_count = 0
    for bl in db.get_bl_by_batch(batch_id):
        if str(bl.get("chat_id") or "").strip():
            continue
        code = str(bl.get("code") or "").strip()
        chat_id = _find_chat_for_bl_code(code, chat_lookup)
        via = "title" if chat_id else ""
        if not chat_id:
            hist = db.lookup_bl_link_history(code)
            if hist and (not known_chat_ids or hist["chat_id"] in known_chat_ids):
                chat_id = hist["chat_id"]
                via = "history"
        if chat_id and db.set_bl_chat_id(bl["id"], chat_id):
            if via == "title":
                title_count += 1
            else:
                history_count += 1
            linked.append(
                {
                    "code": code,
                    "chat_id": chat_id,
                    "chat_title": chats_by_id.get(chat_id, ""),
                    "via": via,
                }
            )
        else:
            unlinked_left += 1

    return jsonify(
        {
            "ok": True,
            "linked_count": len(linked),
            "title_count": title_count,
            "history_count": history_count,
            "unlinked_left": unlinked_left,
            "linked": linked,
        }
    )


@app.route("/api/bl", methods=["POST"])
@editor_required
def api_add_bl():
    data = request.json or {}
    batch_id = data.get("batch_id")
    code = (data.get("code") or "").strip()
    client_name = (data.get("client_name") or "").strip()
    chat_id = (data.get("chat_id") or "").strip()
    message_language = (data.get("message_language") or "").strip()
    moderator_tg_id = (data.get("moderator_tg_id") or "").strip()
    sales_manager_tg_id = (data.get("sales_manager_tg_id") or "").strip()
    cargo_type = (data.get("cargo_type") or "").strip()
    weight_kg = data.get("weight_kg", 0)
    volume_cbm = data.get("volume_cbm", 0)
    quantity_places = data.get("quantity_places", 0)
    quantity_places_breakdown = (data.get("quantity_places_breakdown") or data.get("quantity_places") or "").strip()
    cargo_description = (data.get("cargo_description") or "").strip()

    if not batch_id or not code:
        return jsonify({"error": "batch_id и code обязательны"}), 400

    if not db.add_bl(
        batch_id,
        code,
        client_name,
        chat_id,
        moderator_tg_id,
        sales_manager_tg_id,
        cargo_type,
        weight_kg,
        volume_cbm,
        quantity_places,
        quantity_places_breakdown,
        cargo_description,
        message_language,
    ):
        return jsonify({"error": "BL-код уже существует в этой партии"}), 400

    return jsonify({"ok": True})


@app.route("/api/bl/<int:bl_id>", methods=["PUT"])
@editor_required
def api_update_bl(bl_id):
    data = request.json or {}
    try:
        db.update_bl(
            bl_id,
            data.get("code", ""),
            data.get("client_name", ""),
            data.get("chat_id", ""),
            data.get("status"),
            data.get("moderator_tg_id", ""),
            data.get("sales_manager_tg_id", ""),
            data.get("cargo_type", ""),
            data.get("weight_kg", 0),
            data.get("volume_cbm", 0),
            data.get("quantity_places", 0),
            data.get("quantity_places_breakdown", data.get("quantity_places", "")),
            data.get("cargo_description", ""),
            data.get("message_language", ""),
        )
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    return jsonify({"ok": True})


@app.route("/api/bl/<int:bl_id>", methods=["DELETE"])
@editor_required
def api_delete_bl(bl_id):
    try:
        db.delete_bl(bl_id)
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500
    return jsonify({"ok": True})


@app.route("/api/bl/bulk-delete", methods=["POST"])
@editor_required
def api_bulk_delete_bl():
    """Delete many BL rows in one call.

    Body: {"ids": [int, int, ...]}
    Returns: {ok: True, deleted: N, failed: [{id, error}, ...]}
    Each row uses the same robust delete_bl path (handles attached files,
    problems, send logs, etc). Failures don't stop the batch — we
    accumulate per-id errors so the operator can see which ones survived.
    """
    data = request.json or {}
    raw_ids = data.get("ids") or []
    if not isinstance(raw_ids, list) or not raw_ids:
        return jsonify({"error": "Не выбраны BL для удаления"}), 400

    ids: list[int] = []
    for item in raw_ids:
        try:
            ids.append(int(item))
        except (TypeError, ValueError):
            continue
    if not ids:
        return jsonify({"error": "Список ID пуст"}), 400

    deleted = 0
    failed: list[dict] = []
    for bl_id in ids:
        try:
            db.delete_bl(bl_id)
            deleted += 1
        except Exception as exc:
            app.logger.exception("Bulk delete failed for bl_id=%s", bl_id)
            failed.append({"id": bl_id, "error": str(exc)})
    return jsonify({"ok": True, "deleted": deleted, "failed": failed})


@app.route("/api/batches/<int:batch_id>/recall-tracking", methods=["POST"])
@editor_required
def api_recall_tracking(batch_id):
    """Delete previously-sent Telegram messages for this batch.

    Body (optional):
      {
        "bl_ids":     [int, ...],   # restrict to specific BLs
        "scope":      "last_dispatch" | "all"  (default "last_dispatch")
      }

    "last_dispatch" — only messages from the most recent broadcast window
    (everything within 60 seconds of the newest sent_at timestamp). This
    is the typical use case: "I just sent wrong tracking — undo it."

    "all" — every still-recallable message in this batch (or for the
    selected BLs if bl_ids was provided).

    Per-message: Telegram's deleteMessage works for ~48 h on messages the
    bot sent itself. Failures are tagged on the row so the operator can
    see WHY (e.g. "message to delete not found", "not enough rights").
    """
    if not BOT_TOKEN:
        return jsonify({"error": "BOT_TOKEN не настроен"}), 500
    batch = db.get_batch(batch_id)
    if not batch:
        abort(404)

    data = request.json or {}
    raw_bl_ids = data.get("bl_ids") or []
    if not isinstance(raw_bl_ids, list):
        raw_bl_ids = []
    bl_ids: list[int] = []
    for item in raw_bl_ids:
        try:
            bl_ids.append(int(item))
        except (TypeError, ValueError):
            continue
    scope = str(data.get("scope") or "last_dispatch").strip().lower()
    if scope not in {"last_dispatch", "all"}:
        scope = "last_dispatch"

    rows = db.get_recallable_messages(
        batch_id=batch_id,
        bl_ids=bl_ids if bl_ids else None,
        only_unrecalled=True,
    )
    if not rows:
        return jsonify({"ok": True, "deleted": 0, "failed": 0, "skipped": 0,
                        "message": "Нечего отзывать — нет сообщений в журнале"})

    # last_dispatch: keep only rows within 60 s of the freshest sent_at.
    if scope == "last_dispatch":
        newest = rows[0].get("sent_at") or ""
        try:
            from datetime import datetime as _dt
            newest_dt = _dt.strptime(newest, "%Y-%m-%d %H:%M:%S")
            cutoff = newest_dt.timestamp() - 60.0
            filtered = []
            for r in rows:
                try:
                    sent_dt = _dt.strptime(r.get("sent_at") or "", "%Y-%m-%d %H:%M:%S")
                    if sent_dt.timestamp() >= cutoff:
                        filtered.append(r)
                except ValueError:
                    continue
            rows = filtered
        except ValueError:
            pass  # malformed timestamp; just recall everything

    deleted = 0
    failed = 0
    deleted_bl_ids: set[int] = set()
    for row in rows:
        try:
            telegram_delete_message(row["chat_id"], row["message_id"])
            db.mark_message_recalled(row["id"])
            deleted += 1
            if row.get("bl_id"):
                try:
                    deleted_bl_ids.add(int(row["bl_id"]))
                except (TypeError, ValueError):
                    pass
        except Exception as exc:
            # Telegram returned an error (most often: message too old, or
            # bot lost permission). Tag the row so the operator can see why
            # and we won't keep retrying it on every subsequent recall.
            err_str = str(exc)[:300]
            # "message to delete not found" is a normal outcome if the user
            # already deleted it manually — treat as a soft success so the
            # row doesn't keep showing up in the "still recallable" list.
            if "to delete not found" in err_str.lower():
                db.mark_message_recalled(row["id"])
                deleted += 1
                if row.get("bl_id"):
                    try:
                        deleted_bl_ids.add(int(row["bl_id"]))
                    except (TypeError, ValueError):
                        pass
            else:
                db.mark_message_recalled(row["id"], error=err_str)
                failed += 1

    # If we recalled the tracking message for a BL, clear its
    # tracking_delivery_coverage so the system shows it as "needs to send"
    # again — otherwise the green ✅ "отправлено" badge stays misleading.
    cleared_coverage = 0
    if deleted_bl_ids:
        try:
            cleared_coverage = db.clear_tracking_delivery_coverage(list(deleted_bl_ids))
        except Exception:
            app.logger.exception("Failed to clear tracking_delivery_coverage after recall")

    return jsonify({
        "ok": True,
        "deleted": deleted,
        "failed": failed,
        "scope": scope,
        "cleared_coverage": cleared_coverage,
    })


@app.route("/api/batches/<int:batch_id>/recallable", methods=["GET"])
@login_required
def api_recallable_summary(batch_id):
    """Cheap summary for the UI: how many messages can be recalled, what's
    the newest sent_at? Used to enable/disable the recall button and show
    "Last dispatch: 16.05.2026 17:31, 23 messages".
    """
    rows = db.get_recallable_messages(batch_id=batch_id, only_unrecalled=True)
    total = len(rows)
    newest = rows[0].get("sent_at") if rows else ""
    # Count messages within the last-dispatch window (60s)
    last_dispatch_count = 0
    if rows and newest:
        try:
            from datetime import datetime as _dt
            newest_dt = _dt.strptime(newest, "%Y-%m-%d %H:%M:%S")
            cutoff = newest_dt.timestamp() - 60.0
            for r in rows:
                try:
                    sent_dt = _dt.strptime(r.get("sent_at") or "", "%Y-%m-%d %H:%M:%S")
                    if sent_dt.timestamp() >= cutoff:
                        last_dispatch_count += 1
                except ValueError:
                    continue
        except ValueError:
            last_dispatch_count = total
    return jsonify({
        "ok": True,
        "total_recallable": total,
        "last_dispatch_count": last_dispatch_count,
        "newest_sent_at": newest or "",
    })


@app.route("/api/bl/<int:bl_id>/move", methods=["POST"])
@editor_required
def api_move_bl(bl_id):
    data = request.json or {}
    target_batch_id = data.get("target_batch_id")
    if not target_batch_id:
        return jsonify({"error": "target_batch_id обязателен"}), 400
    try:
        result = db.move_bl_to_batch(bl_id, target_batch_id)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    return jsonify({"ok": True, "result": result})


@app.route("/api/bl/<int:bl_id>/merge", methods=["POST"])
@editor_required
def api_merge_bl(bl_id):
    data = request.json or {}
    target_bl_id = data.get("target_bl_id")
    if not target_bl_id:
        return jsonify({"error": "target_bl_id обязателен"}), 400
    try:
        result = db.merge_bl_into_target(bl_id, target_bl_id)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    return jsonify({"ok": True, "result": result})


@app.route("/api/bl/<int:bl_id>/files")
@login_required
def api_files(bl_id):
    return jsonify(db.get_files(bl_id))


@app.route("/api/bl/<int:bl_id>/send-exclusion", methods=["POST"])
@editor_required
def api_set_bl_send_exclusion(bl_id):
    data = request.json or {}
    excluded = bool(data.get("excluded"))
    try:
        result = db.set_batch_send_exclusion(bl_id, excluded)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 404
    return jsonify({"ok": True, "result": result})


@app.route("/api/bl/<int:bl_id>/files", methods=["POST"])
@editor_required
def api_upload(bl_id):
    if "file" not in request.files:
        return jsonify({"error": "Файл не выбран"}), 400

    uploaded_file = request.files["file"]
    if not uploaded_file.filename:
        return jsonify({"error": "Выбери файл для загрузки"}), 400
    ext = uploaded_file.filename.rsplit(".", 1)[-1].lower() if "." in uploaded_file.filename else ""
    if ext not in ALLOWED_EXT:
        return jsonify({"error": f"Тип файла .{ext} не разрешён"}), 400

    original_filename = (uploaded_file.filename or "").strip()
    storage_name = secure_filename(original_filename) or f"file_{secrets.token_hex(4)}.{ext}"
    filename = original_filename or storage_name
    unique = f"bl{bl_id}_{secrets.token_hex(4)}_{storage_name}"
    file_path = os.path.join(UPLOAD_FOLDER, unique)
    uploaded_file.save(file_path)
    db.add_file(bl_id, filename, file_path)
    return jsonify({"ok": True, "filename": filename})


@app.route("/api/files/<int:file_id>", methods=["DELETE"])
@editor_required
def api_delete_file(file_id):
    db.delete_file(file_id)
    return jsonify({"ok": True})


# ─────────────────────────────────────────────────────────────────────────
# Bulk Packing Lists upload: extract BL code from filename and auto-attach
# ─────────────────────────────────────────────────────────────────────────
_BL_CODE_RE = re.compile(r"[A-Z]{2,4}[-_ ]?\d+[A-Z0-9]*")
# Strip everything that is not a letter/digit (any script) — used for fuzzy
# client-name matching: drops spaces, punctuation, parentheses, dashes, etc.
_NON_ALNUM_RE = re.compile(r"[^\w]+", re.UNICODE)


def _normalize_bl_code(value: str) -> str:
    """Canonicalize a BL code for index lookups.

    Strips dashes, underscores, spaces, dots and uppercases — so all of
    "BL-190", "bl_190", "BL 190", "BL.190", "bl190" collapse to "BL190".
    """
    return (
        str(value or "")
        .strip()
        .upper()
        .replace("-", "")
        .replace("_", "")
        .replace(" ", "")
        .replace(".", "")
    )


def _extract_bl_code_from_filename(name: str) -> str:
    base = os.path.basename(str(name or "").replace("\\", "/"))
    no_ext = base.rsplit(".", 1)[0] if "." in base else base
    m = _BL_CODE_RE.search(no_ext.upper())
    if not m:
        return ""
    return _normalize_bl_code(m.group(0))


def _normalize_for_fuzzy(text: str) -> str:
    """Lowercase + drop all non-alphanumeric chars (handles spaces, dashes, parens, etc.)."""
    return _NON_ALNUM_RE.sub("", str(text or "").casefold())


# Words that occur in nearly every chat title / filename and would otherwise
# match everything ("BURAQ LOGISTICS", "PACKING LIST", "MESTA" etc.).
_TOKEN_STOPWORDS = {
    "buraq", "logistic", "logistics", "buraqlogistics",
    "буракс", "логистик", "логистикс",
    "packing", "list", "lists",
    "mesta", "places", "joy",
    "tovar", "yuk", "fayl", "file", "files",
    "товар", "груз", "файл",
    "company", "llc", "ltd", "inc", "company", "group", "trade", "trading",
    "ооо", "оао", "ао", "ип",
    "xlsx", "xls", "pdf", "doc", "docx", "zip", "png", "jpg", "jpeg",
    # 2-char noise (prepositions/conjunctions) — 2-char tokens are allowed
    # as client identifiers ("AB", "HI"), so the junk ones must be listed.
    "va", "na", "po", "iz", "uz", "da", "of", "in", "on", "at", "to",
    "by", "or", "an", "is", "no", "ва", "на", "по", "из", "до", "от",
}

# Token regex: pulls Latin and Cyrillic word-runs separately so numbers
# and punctuation never bleed into tokens.
_TOKEN_RE = re.compile(r"[A-Za-zА-Яа-яЁё]+", re.UNICODE)


def _tokenize_name(text: str) -> set[str]:
    """Extract distinctive lowercase tokens for matching.

    Drops BL-NNN codes, stopwords, and single-char tokens. 2-char tokens
    survive (many clients ARE a 2-char identifier: "AB", "HI") — the
    scoring layer only lets them match exactly, never by prefix.
    """
    if not text:
        return set()
    # Strip only literal BL-prefixed codes ("BL-241"). The old wider
    # [A-Z]{2,4}+digits pattern also ate real client identifiers followed
    # by a count — "AB 12 MESTA…" lost its "AB" and matched the wrong
    # client through the generic "light" token.
    cleaned = re.sub(r"\bBL[-_ ]?\d+[A-Z0-9]*\b", " ", str(text), flags=re.IGNORECASE)
    tokens: set[str] = set()
    for word in _TOKEN_RE.findall(cleaned):
        norm = word.lower()
        if len(norm) < 2:
            continue
        if norm in _TOKEN_STOPWORDS:
            continue
        tokens.add(norm)
    return tokens


def _bl_salient_tokens(row: dict, chat_titles: dict[str, str]) -> set[str]:
    """Every word we'd consider 'a name of this BL' for matching purposes."""
    tokens: set[str] = set()
    tokens.update(_tokenize_name(row.get("code") or ""))
    tokens.update(_tokenize_name(row.get("merged_codes") or ""))
    tokens.update(_tokenize_name(row.get("client_name") or ""))
    chat_id = str(row.get("chat_id") or "").strip()
    if chat_id:
        tokens.update(_tokenize_name(chat_titles.get(chat_id, "")))
    return tokens


def _token_hits(filename_tokens: set[str], bl_tokens: set[str]) -> dict[str, tuple[int, str]]:
    """Per-filename-token best hit against one BL's tokens.

    Scoring per token:
    - Exact token match: len(token) * 2
    - Prefix-match (≥4 chars on both sides, e.g. "light" / "lighting"):
      min(len(a), len(b))
    Returns {filename_token: (score, matched_bl_token)}.
    """
    hits: dict[str, tuple[int, str]] = {}
    if not filename_tokens or not bl_tokens:
        return hits
    for ft in filename_tokens:
        best_bt = ""
        best_local = 0
        for bt in bl_tokens:
            if ft == bt:
                local = len(ft) * 2
            elif len(ft) >= 4 and len(bt) >= 4 and (ft.startswith(bt) or bt.startswith(ft)):
                local = min(len(ft), len(bt))
            else:
                continue
            if local > best_local:
                best_local = local
                best_bt = bt
        if best_local:
            hits[ft] = (best_local, best_bt)
    return hits


# Minimum TOTAL score to accept a fuzzy name match (roughly "one 3+ char
# exact match" or "one 4-char prefix overlap"), and the minimum score that
# must come from UNIQUE tokens — tokens pointing at exactly one client.
# Words like "light" appear in most client names here, so a match driven
# only by shared tokens is noise, not evidence.
_TOKEN_MATCH_MIN_SCORE = 6
_TOKEN_MATCH_MIN_UNIQUE = 4


def _build_batch_bl_code_index(batch_id: int) -> dict:
    """Map every BL code (primary and merged) within a batch to its row."""
    index: dict[str, dict] = {}
    for row in db.get_bl_by_batch(batch_id) or []:
        primary = _normalize_bl_code(row.get("code") or "")
        if primary:
            index.setdefault(primary, row)
        merged_raw = str(row.get("merged_codes") or "")
        for part in re.split(r"[,;\s]+", merged_raw):
            piece = _normalize_bl_code(part)
            if piece:
                index.setdefault(piece, row)
    return index


def _build_chat_title_lookup(batch_rows: list[dict]) -> dict[str, str]:
    """chat_id → chat title, for every chat referenced by the batch.

    Used to pull the "client" name out of the Telegram chat title when
    bl_codes.client_name is empty (which is the common case here — most
    sheets only store BL code + chat id, and the actual client name only
    lives in the Telegram group title).
    """
    chat_ids = {
        str(row.get("chat_id") or "").strip()
        for row in batch_rows
        if str(row.get("chat_id") or "").strip()
    }
    if not chat_ids:
        return {}
    try:
        chats = db.get_telegram_chats(include_inactive=True) or []
    except Exception:
        return {}
    return {
        str(chat.get("chat_id") or "").strip(): str(chat.get("title") or "").strip()
        for chat in chats
        if str(chat.get("chat_id") or "").strip() in chat_ids
    }


def _resolve_filename_to_bl(
    filename: str,
    code_index: dict,
    batch_rows: list[dict],
    chat_titles: dict[str, str] | None = None,
) -> dict | None:
    """Try multiple matching strategies, return the BL row that wins or None."""
    if not filename:
        return None
    chat_titles = chat_titles or {}
    base = os.path.basename(str(filename).replace("\\", "/"))
    no_ext = base.rsplit(".", 1)[0] if "." in base else base

    # 1. BL code match (BL171, BL-171, bl_171a) — normalized on both sides.
    code_match = _BL_CODE_RE.search(no_ext.upper())
    if code_match:
        normalized_code = _normalize_bl_code(code_match.group(0))
        row = code_index.get(normalized_code)
        if row:
            return {"row": row, "method": "code", "matched_on": code_match.group(0)}

    # 1b. Whole BL code at the START of the filename — catches phrase-like
    #     codes the regex above can't ("N LIGHT" → "N LIGHT 45 MESTA…").
    #     ≥4 normalized chars so 2-letter initials don't false-hit; the
    #     longest matching code wins ("GIGA LIGHT" beats "GIGA").
    flat_name = _NON_ALNUM_RE.sub("", no_ext.casefold())
    if flat_name:
        best_code_row = None
        best_code_len = 0
        best_code_display = ""
        for norm_code, row in code_index.items():
            flat_code = str(norm_code or "").casefold()
            if len(flat_code) >= 4 and flat_name.startswith(flat_code) and len(flat_code) > best_code_len:
                best_code_row = row
                best_code_len = len(flat_code)
                best_code_display = row.get("code") or norm_code
        if best_code_row is not None:
            return {"row": best_code_row, "method": "code", "matched_on": best_code_display}

    # 2. Token-based name match, discriminativeness-aware. A filename
    #    token that hits SEVERAL different clients ("light" appears in
    #    ~80% of our client names) identifies nobody — it only adds
    #    confidence to a row that some UNIQUE token already picked.
    #    Catches:
    #      - "EURO LIGHT 6 MESTA.xlsx" → "EURO LIGHTING" ("euro" is unique)
    #      - "WIZERA 240 MESTA ALYUMIN.xlsx" → "WIZERA" (unique exact)
    #      - "AB 12 STREET LIGHT.xls" → "AB LIGHTING" ("ab" is unique;
    #        "light" alone would have pulled it to any *LIGHT* client)
    #    A file whose only hits are shared tokens returns None — better an
    #    honest "Не найдено" than a confident wrong client.
    filename_tokens = _tokenize_name(no_ext)
    if filename_tokens:
        scored_rows: list[tuple[dict, dict[str, tuple[int, str]]]] = []
        rows_per_token: dict[str, int] = {}
        for row in batch_rows:
            hits = _token_hits(filename_tokens, _bl_salient_tokens(row, chat_titles))
            if not hits:
                continue
            for ft in hits:
                rows_per_token[ft] = rows_per_token.get(ft, 0) + 1
            scored_rows.append((row, hits))

        best_row: dict | None = None
        best_key = (0, 0)
        best_matched: list[str] = []
        tie = False
        for row, hits in scored_rows:
            unique = sum(sc for ft, (sc, _bt) in hits.items() if rows_per_token.get(ft) == 1)
            total = sum(sc for _ft, (sc, _bt) in hits.items())
            key = (unique, total)
            if key > best_key:
                best_key = key
                best_row = row
                best_matched = [
                    (bt or ft) for ft, (_sc, bt) in hits.items() if rows_per_token.get(ft) == 1
                ] or [(bt or ft) for ft, (_sc, bt) in hits.items()]
                tie = False
            elif key == best_key and best_row is not None:
                tie = True  # two clients equally likely — let the operator decide
        unique_score, total_score = best_key
        confident = (
            unique_score >= _TOKEN_MATCH_MIN_SCORE
            or (unique_score >= _TOKEN_MATCH_MIN_UNIQUE and total_score >= _TOKEN_MATCH_MIN_SCORE)
        )
        if best_row is not None and confident and not tie:
            return {
                "row": best_row,
                "method": "name",
                "matched_on": ", ".join(sorted(set(best_matched))) or "имя",
            }

    return None


@app.route("/api/batches/<int:batch_id>/packing-lists/resolve", methods=["POST"])
@editor_required
def api_resolve_packing_list_codes(batch_id):
    batch = db.get_batch(batch_id)
    if not batch:
        abort(404)
    data = request.json or {}
    filenames_raw = data.get("filenames") or []
    codes_raw = data.get("codes") or []  # legacy fallback
    if not isinstance(filenames_raw, list):
        filenames_raw = []
    if not isinstance(codes_raw, list):
        codes_raw = []

    # STRICT batch scoping: a packing list uploaded from batch X may only
    # ever attach to BLs of batch X. The old cross-batch fallback misfired
    # whenever the same BL code lived in two batches — the file silently
    # landed on the other batch's BL and "left" the batch the operator was
    # working in.
    current_batch_rows = db.get_bl_by_batch(batch_id) or []
    chat_titles = _build_chat_title_lookup(current_batch_rows)

    current_code_index = _build_batch_bl_code_index(batch_id)

    # Manual-pick dropdown options — current batch only.
    bl_options: list[dict] = []
    seen_ids: set[int] = set()

    def _push_option(row: dict, is_current: bool, batch_name: str = ""):
        rid = row.get("id")
        if not rid or int(rid) in seen_ids:
            return
        seen_ids.add(int(rid))
        chat_id = str(row.get("chat_id") or "").strip()
        client_name = (row.get("client_name") or "").strip()
        chat_title = chat_titles.get(chat_id, "") if chat_id else ""
        bl_options.append({
            "id": rid,
            "code": row.get("code") or "",
            "display_code": row.get("display_code") or row.get("code") or "",
            "client_name": client_name or chat_title,
            "batch_id": row.get("batch_id"),
            "batch_name": batch_name or "",
            "is_current": bool(is_current),
        })

    for row in current_batch_rows:
        _push_option(row, is_current=True, batch_name=batch.get("name", ""))

    def _enrich(row: dict, hit_method: str, matched_on: str, is_current: bool, batch_name: str) -> dict:
        row_chat_id = str(row.get("chat_id") or "").strip()
        return {
            "bl_id": row.get("id"),
            "code": row.get("code"),
            "client_name": (row.get("client_name") or "").strip()
                            or chat_titles.get(row_chat_id, ""),
            "method": hit_method,
            "matched_on": matched_on,
            "batch_id": row.get("batch_id"),
            "batch_name": batch_name or "",
            "is_current": bool(is_current),
        }

    # New API: resolve by filename (preferred). Current batch ONLY — a
    # file that doesn't fit any BL here stays unmatched so the operator
    # picks by hand instead of the file silently leaving the batch.
    resolved_by_filename: dict[str, dict] = {}
    for raw_name in filenames_raw:
        name = str(raw_name or "").strip()
        if not name:
            continue

        hit = _resolve_filename_to_bl(name, current_code_index, current_batch_rows, chat_titles)
        if hit:
            resolved_by_filename[raw_name] = _enrich(
                hit["row"], hit["method"], hit["matched_on"],
                is_current=True, batch_name=batch.get("name", ""),
            )

    # Legacy API: resolve by extracted code (current batch only — back-compat).
    resolved_by_code: dict[str, dict] = {}
    for raw_code in codes_raw:
        code = _normalize_bl_code(raw_code)
        if not code:
            continue
        row = current_code_index.get(code)
        if not row:
            continue
        resolved_by_code[raw_code] = {
            "bl_id": row.get("id"),
            "code": row.get("code"),
            "client_name": row.get("client_name") or "",
        }

    return jsonify({
        "ok": True,
        "resolved": resolved_by_code,             # legacy
        "resolved_by_filename": resolved_by_filename,
        "bl_options": bl_options,
    })


@app.route("/api/batches/<int:batch_id>/packing-lists/bulk", methods=["POST"])
@editor_required
def api_bulk_packing_lists(batch_id):
    batch = db.get_batch(batch_id)
    if not batch:
        abort(404)

    files = request.files.getlist("files")
    bl_ids = request.form.getlist("bl_ids")
    if not files or not bl_ids or len(files) != len(bl_ids):
        return jsonify({"error": "Не переданы файлы или BL не сопоставлены"}), 400

    # Defensive: make sure the upload folder still exists. On ephemeral
    # filesystems (e.g. fresh container) the directory can be missing.
    try:
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    except Exception as exc:
        return jsonify({"error": f"Не удалось создать папку загрузки: {exc}"}), 500

    # Whitelist = BLs of THIS batch only. Cross-batch attach used to be
    # allowed and misfired when the same BL code existed in two batches:
    # the file silently landed on the other batch's BL. Strict scoping —
    # a file uploaded from batch X can only attach inside batch X.
    allowed_ids: set[int] = set()
    for row in db.get_bl_by_batch(batch_id) or []:
        rid = row.get("id")
        if rid is not None:
            try:
                allowed_ids.add(int(rid))
            except (TypeError, ValueError):
                pass

    # Pre-load existing filenames per BL so we can suppress duplicates
    # server-side. Two scenarios this kills:
    #   1) The user uploaded the same packing list a few minutes ago and
    #      doesn't realize it (page refresh, fresh selection).
    #   2) Two concurrent requests in the same multipart batch both try to
    #      attach the same filename to the same BL (rare race, but cheap
    #      to defend against).
    existing_per_bl: dict[int, set[str]] = {}

    def _existing_names(bl_id_int: int) -> set[str]:
        if bl_id_int not in existing_per_bl:
            try:
                existing_files = db.get_files(bl_id_int) or []
            except Exception:
                existing_files = []
            existing_per_bl[bl_id_int] = {
                str(f.get("filename") or "").strip().lower()
                for f in existing_files
                if (f.get("filename") or "").strip()
            }
        return existing_per_bl[bl_id_int]

    uploaded = 0
    failed = 0
    skipped_duplicates = 0
    results = []
    for uploaded_file, bl_id_raw in zip(files, bl_ids):
        if not uploaded_file or not uploaded_file.filename:
            failed += 1
            results.append({"filename": "", "ok": False, "error": "Пустой файл"})
            continue
        try:
            bl_id = int(bl_id_raw)
        except (TypeError, ValueError):
            failed += 1
            results.append({"filename": uploaded_file.filename, "ok": False, "error": "Неверный BL"})
            continue
        if bl_id not in allowed_ids:
            failed += 1
            results.append({"filename": uploaded_file.filename, "ok": False, "error": "BL не из этой партии"})
            continue

        original_filename = (uploaded_file.filename or "").strip()
        ext = original_filename.rsplit(".", 1)[-1].lower() if "." in original_filename else ""
        if ext not in ALLOWED_EXT:
            failed += 1
            results.append({"filename": original_filename, "ok": False, "error": f"Тип .{ext} не разрешён"})
            continue

        # Strip subdirectory components from the relative path (folder upload)
        clean_base = os.path.basename(original_filename.replace("\\", "/"))
        storage_name = secure_filename(clean_base) or f"file_{secrets.token_hex(4)}.{ext}"
        filename = clean_base or storage_name

        # Dedupe: this BL already has a file with the same basename.
        # Skip silently and report as success so the UI tags the row as ✅
        # rather than ❌ — from the user's perspective the file "is attached".
        name_key = filename.strip().lower()
        already = _existing_names(bl_id)
        if name_key in already:
            skipped_duplicates += 1
            uploaded += 1   # count as success for UI purposes
            results.append({
                "filename": filename,
                "ok": True,
                "bl_id": bl_id,
                "skipped_duplicate": True,
            })
            continue

        try:
            unique = f"bl{bl_id}_{secrets.token_hex(4)}_{storage_name}"
            file_path = os.path.join(UPLOAD_FOLDER, unique)
            uploaded_file.save(file_path)
            db.add_file(bl_id, filename, file_path)
            # Remember it so a same-batch second copy of this file also gets
            # caught by the dedupe above.
            already.add(name_key)
            uploaded += 1
            results.append({"filename": filename, "ok": True, "bl_id": bl_id})
        except Exception as exc:
            failed += 1
            results.append({"filename": original_filename, "ok": False, "error": str(exc)})

    return jsonify({
        "ok": True,
        "uploaded": uploaded,
        "failed": failed,
        "skipped_duplicates": skipped_duplicates,
        "results": results,
    })


@app.route("/public/file/<public_token>")
def public_file(public_token):
    file_info = db.get_file_by_public_token(public_token)
    if not file_info:
        abort(404)
    file_path = file_info.get("file_path") or ""
    if not file_path or not os.path.exists(file_path):
        abort(404)
    return send_file(
        file_path,
        as_attachment=False,
        download_name=file_info.get("filename") or os.path.basename(file_path),
        conditional=True,
    )


# Bulk-send job registry: batch broadcasts run in background threads and
# the frontend polls /api/send-jobs/<id> for progress. The semaphore caps
# concurrent Telegram sends ACROSS jobs so sending two batches at once
# doesn't double the request rate and trip 429 throttling.
_send_jobs: dict[str, dict] = {}
_send_jobs_lock = threading.Lock()
_BULK_SEND_SLOTS = threading.BoundedSemaphore(8)


@app.route("/api/batches/<int:batch_id>/send", methods=["POST"])
@editor_required
def api_send_batch(batch_id):
    if not BOT_TOKEN:
        return jsonify({"error": "BOT_TOKEN не настроен в .env"}), 500

    batch = db.get_batch(batch_id)
    if not batch:
        abort(404)

    data = request.json or {}
    selected_raw = data.get("selected_bl_ids") or []
    selected_ids = set()
    if isinstance(selected_raw, list):
        for item in selected_raw:
            try:
                selected_ids.add(int(item))
            except (TypeError, ValueError):
                continue
    include_related_batches = False

    bl_rows = db.get_bl_by_batch(batch_id)
    if selected_ids:
        bl_rows = [bl for bl in bl_rows if int(bl.get("id") or 0) in selected_ids]
    else:
        bl_rows = [
            bl
            for bl in bl_rows
            if bl.get("chat_id") and not bl.get("send_excluded") and not bl.get("tracking_sent_current")
        ]

    if not bl_rows:
        return jsonify({"error": "Нет выбранных BL для отправки"}), 400

    # Split BLs into:
    #   - skipped_rows: no chat_id or already covered by another BL of the
    #     same client in this same broadcast (the dedupe rule we had)
    #   - dispatch_rows: one BL per chat, to be sent in parallel
    results: list[dict] = []
    sent_chats: set[str] = set()
    dispatch_rows: list[dict] = []
    for bl in bl_rows:
        chat_id = str(bl.get("chat_id") or "").strip()
        if not chat_id:
            results.append({
                "code": bl["code"],
                "client": bl["client_name"],
                "success": False,
                "skipped": True,
                "error": "Нет chat_id",
            })
            continue
        if chat_id in sent_chats:
            results.append({
                "code": bl["code"],
                "client": bl["client_name"],
                "success": False,
                "skipped": True,
                "error": "Уже покрыто сообщением этого клиента в текущей отправке",
            })
            continue
        sent_chats.add(chat_id)
        dispatch_rows.append(bl)

    if not dispatch_rows:
        # Nothing to actually send — reply synchronously in the old shape.
        skipped = sum(1 for item in results if item.get("skipped"))
        return jsonify({"ok": True, "sent": 0, "skipped": skipped, "total": len(results), "results": results})

    def _dispatch_one(bl):
        try:
            # The global semaphore caps Telegram traffic across ALL
            # simultaneously running send jobs (e.g. two batches sent at
            # once) — without it 2×8 parallel senders trip Telegram's
            # rate limit and 429 backoffs balloon the send time.
            with _BULK_SEND_SLOTS:
                success, error_msg = send_bl_package(
                    bl,
                    batch["name"],
                    include_related_batches=include_related_batches,
                )
        except Exception as exc:
            success, error_msg = False, str(exc)
            app.logger.exception(
                "Bulk send unexpected failure for bl_id=%s chat_id=%s",
                bl.get("id"), bl.get("chat_id"),
            )
        try:
            db.add_log(
                bl["id"], bl["code"], batch["name"],
                bl["chat_id"], bl["status"], success, error_msg,
            )
        except Exception:
            app.logger.exception("Failed to record send log for bl_id=%s", bl.get("id"))
        return {
            "code": bl["code"],
            "client": bl["client_name"],
            "success": success,
            "skipped": False,
            "error": error_msg,
        }

    # The broadcast runs as a BACKGROUND JOB and the request returns
    # immediately with a job id the frontend polls. A synchronous reply
    # used to keep the HTTP request open for the whole broadcast — with
    # Telegram 429 backoffs that easily exceeds the proxy/browser timeout,
    # so the UI showed an error while the send actually kept running.
    with _send_jobs_lock:
        for existing in _send_jobs.values():
            if existing["batch_id"] == batch_id and not existing["finished"]:
                return jsonify({
                    "ok": True,
                    "job_id": existing["job_id"],
                    "total": existing["total"],
                    "already_running": True,
                })
        job_id = secrets.token_hex(6)
        job = {
            "job_id": job_id,
            "batch_id": batch_id,
            "batch_name": batch["name"],
            "total": len(results) + len(dispatch_rows),
            "done": len(results),
            "sent": 0,
            "skipped": sum(1 for item in results if item.get("skipped")),
            "results": list(results),
            "finished": False,
            "started_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
        _send_jobs[job_id] = job
        # Prune old finished jobs so the registry doesn't grow forever.
        finished_ids = [jid for jid, j in _send_jobs.items() if j["finished"]]
        for jid in finished_ids[:-20]:
            _send_jobs.pop(jid, None)

    def _run_send_job():
        # Parallel fan-out across distinct chats. Within one chat the
        # order is still strict (send_bl_package iterates the bundle
        # serially), so tracking message → files cadence is preserved.
        MAX_PARALLEL = 8
        try:
            with ThreadPoolExecutor(
                max_workers=min(MAX_PARALLEL, len(dispatch_rows)),
                thread_name_prefix="bulk-send",
            ) as pool:
                futures = [pool.submit(_dispatch_one, bl) for bl in dispatch_rows]
                for future in as_completed(futures):
                    try:
                        item = future.result()
                    except Exception as exc:
                        app.logger.exception("Bulk send future error: %s", exc)
                        item = {
                            "code": "",
                            "client": "",
                            "success": False,
                            "skipped": False,
                            "error": str(exc),
                        }
                    with _send_jobs_lock:
                        job["results"].append(item)
                        job["done"] += 1
                        if item["success"]:
                            job["sent"] += 1
        finally:
            with _send_jobs_lock:
                job["finished"] = True

    threading.Thread(
        target=_run_send_job, name=f"send-job-{job_id}", daemon=True
    ).start()
    return jsonify({"ok": True, "job_id": job_id, "total": job["total"]})


@app.route("/api/send-jobs/<job_id>")
@login_required
def api_send_job_status(job_id):
    with _send_jobs_lock:
        job = _send_jobs.get(str(job_id))
        if not job:
            return jsonify({"error": "Задача отправки не найдена"}), 404
        return jsonify({**job, "results": list(job["results"])})


@app.route("/api/bl/<int:bl_id>/send", methods=["POST"])
@editor_required
def api_send_one(bl_id):
    if not BOT_TOKEN:
        return jsonify({"error": "BOT_TOKEN не настроен в .env"}), 500

    bl = db.get_bl_by_id(bl_id)
    if not bl:
        abort(404)

    if not bl["chat_id"]:
        return jsonify({"error": "Не указан chat_id"}), 400

    batch = db.get_batch(bl["batch_id"])
    batch_name = batch["name"] if batch else "—"
    success, error_msg = send_bl_package(
        bl,
        batch_name,
        include_related_batches=False,
    )
    db.add_log(bl["id"], bl["code"], batch_name, bl["chat_id"], bl["status"], success, error_msg)

    if not success:
        return jsonify({"error": error_msg}), 500

    return jsonify({"ok": True})


@app.route("/api/logs")
@login_required
def api_logs():
    limit = int(request.args.get("limit", 100))
    return jsonify(db.get_logs(limit))


@app.route("/api/login-history")
@editor_required
def api_login_history():
    limit = int(request.args.get("limit", 200))
    return jsonify(db.get_login_history(limit))


@app.route("/api/problems")
@login_required
def api_problems():
    return jsonify(
        db.get_problems(
            problem_type=(request.args.get("type") or "").strip(),
            date_from=(request.args.get("date_from") or "").strip(),
            date_to=(request.args.get("date_to") or "").strip(),
            batch_id=(request.args.get("batch_id") or "").strip(),
        )
    )


@app.route("/api/problems", methods=["POST"])
@editor_required
def api_create_problem():
    data = request.json or {}
    bl_id = data.get("bl_id")
    problem_type = (data.get("problem_type") or "").strip()
    description = (data.get("description") or "").strip()

    if not bl_id:
        return jsonify({"error": "Не указан BL"}), 400
    if problem_type not in db.PROBLEM_TYPES:
        return jsonify({"error": "Неверный тип проблемы"}), 400
    if not db.create_problem(bl_id, problem_type, description):
        return jsonify({"error": "BL не найден"}), 404

    return jsonify({"ok": True})


@app.route("/api/problems/<int:problem_id>", methods=["DELETE"])
@editor_required
def api_delete_problem(problem_id):
    if not db.delete_problem(problem_id):
        return jsonify({"error": "Проблема не найдена"}), 404
    return jsonify({"ok": True})


@app.route("/api/problems/export")
@login_required
def api_export_problems():
    problem_type = (request.args.get("type") or "").strip()
    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()
    batch_id = (request.args.get("batch_id") or "").strip()
    rows = db.get_problems(
        problem_type=problem_type,
        date_from=date_from,
        date_to=date_to,
        batch_id=batch_id,
    )

    filter_items = []
    if problem_type:
        filter_items.append(("Тип", db.PROBLEM_TYPES.get(problem_type, problem_type)))
    if date_from:
        filter_items.append(("С", date_from))
    if date_to:
        filter_items.append(("По", date_to))
    if batch_id:
        filter_items.append(("Партия", batch_id))

    filters_html = "".join(
        f'<span class="chip"><span class="chip-label">{html.escape(label)}</span>{html.escape(value)}</span>'
        for label, value in filter_items
    ) or '<span class="chip chip-muted">Все инциденты</span>'

    body_rows_list = []
    for row in rows:
        problem_type_key = row.get("problem_type", "") or ""
        problem_label = db.PROBLEM_TYPES.get(problem_type_key, problem_type_key or "—")
        problem_class = {
            "damage": "badge-red",
            "delay": "badge-amber",
            "shortage": "badge-blue",
        }.get(problem_type_key, "badge-muted")
        status_text = row.get("bl_status", "") or "—"
        status_class = {
            "Xitoy": "badge-muted",
            "Yiwu": "badge-muted",
            "Zhongshan": "badge-muted",
            "Horgos (Qozoq)": "badge-blue",
            "Kashgar (Qirg'iz)": "badge-blue",
            "Altynko'l": "badge-amber",
            "Jarkent": "badge-amber",
            "Almata": "badge-blue",
            "Taraz": "badge-amber",
            "Shimkent": "badge-amber",
            "Qonusbay": "badge-amber",
            "Saryagash": "badge-amber",
            "Yallama": "badge-blue",
            "Irkeshtam": "badge-blue",
            "Osh": "badge-amber",
            "Chuqur": "badge-amber",
            "Dostlik": "badge-blue",
            "Andijon": "badge-amber",
            "Toshkent": "badge-amber",
            "Доставлен": "badge-green",
        }.get(status_text, "badge-muted")
        body_rows_list.append(
            f"""
            <tr>
              <td class="mono">{html.escape(row.get('incident_detected_at', '') or row.get('created_at', '') or '-')}</td>
              <td>{html.escape(row.get('batch_name', '') or '-')}</td>
              <td class="mono strong">{html.escape(row.get('bl_code', '') or '-')}</td>
              <td>{html.escape(row.get('client_name', '') or '-')}</td>
              <td><span class="badge {problem_class}">{html.escape(problem_label)}</span></td>
              <td class="desc">{html.escape(row.get('description', '') or '-')}</td>
              <td><span class="badge {status_class}">{html.escape(status_text)}</span></td>
              <td>{html.escape(row.get('expected_date', '') or '-')}</td>
              <td>{html.escape(row.get('actual_date', '') or '-')}</td>
            </tr>
            """
        )
    body_rows = "".join(body_rows_list)
    if not body_rows:
        body_rows = """
        <tr>
          <td colspan="9" class="empty">Инцидентов по выбранным фильтрам не найдено</td>
        </tr>
        """

    report_html = """
    <!DOCTYPE html>
    <html lang="en">
    <head>
      <meta charset="UTF-8">
      <meta name="viewport" content="width=device-width, initial-scale=1.0">
      <title>BURAQ logistics - Проблемы</title>
      <style>
        :root {
          --bg: #eef1f6;
          --card: #ffffff;
          --text: #17191f;
          --muted: #667085;
          --line: #d8dde6;
          --accent: #111827;
          --accent-soft: #f5a623;
          --accent-soft-2: #fff3dd;
          --danger: #c7344f;
          --warning: #d97706;
          --info: #2563eb;
          --ok: #0f9f6e;
        }
        * { box-sizing: border-box; }
        body { margin: 0; font-family: Arial, Helvetica, sans-serif; background: linear-gradient(180deg, #f7f8fb 0%, var(--bg) 100%); color: var(--text); }
        .page { max-width: 1400px; margin: 0 auto; padding: 28px; }
        .hero {
          display: flex;
          align-items: flex-start;
          justify-content: space-between;
          gap: 20px;
          margin-bottom: 20px;
          padding: 24px 26px;
          border-radius: 24px;
          background: linear-gradient(135deg, #121722 0%, #1c2433 48%, #141922 100%);
          color: #fff;
          box-shadow: 0 24px 60px rgba(17, 24, 39, .18);
        }
        .brand { font-size: 12px; font-weight: 800; letter-spacing: .28em; text-transform: uppercase; color: #f6c467; margin-bottom: 10px; }
        .title { font-size: 32px; font-weight: 800; letter-spacing: -0.03em; }
        .meta { color: rgba(255,255,255,.72); font-size: 13px; margin-top: 8px; }
        .hero-stats { display: flex; gap: 10px; flex-wrap: wrap; justify-content: flex-end; }
        .stat {
          min-width: 138px;
          padding: 14px 16px;
          border-radius: 16px;
          background: rgba(255,255,255,.06);
          border: 1px solid rgba(255,255,255,.09);
        }
        .stat-v { font-size: 24px; font-weight: 800; line-height: 1; color: #fff; }
        .stat-l { font-size: 11px; margin-top: 8px; color: rgba(255,255,255,.64); text-transform: uppercase; letter-spacing: .12em; }
        .actions { display: flex; gap: 10px; }
        .btn { border: none; border-radius: 12px; padding: 11px 16px; cursor: pointer; font-size: 14px; font-weight: 700; }
        .btn-dark { background: var(--accent-soft); color: #17191f; }
        .btn-light { background: rgba(255,255,255,.1); color: #fff; border: 1px solid rgba(255,255,255,.12); }
        .card { background: var(--card); border: 1px solid var(--line); border-radius: 22px; overflow: hidden; box-shadow: 0 16px 40px rgba(15, 23, 42, .06); }
        .card-head { padding: 18px 22px; border-bottom: 1px solid var(--line); display: flex; justify-content: space-between; gap: 16px; align-items: center; }
        .card-title { font-size: 18px; font-weight: 700; }
        .filters { display: flex; gap: 8px; flex-wrap: wrap; justify-content: flex-end; }
        .chip {
          display: inline-flex;
          align-items: center;
          gap: 8px;
          padding: 7px 10px;
          border-radius: 999px;
          background: var(--accent-soft-2);
          color: #4b5563;
          font-size: 12px;
          font-weight: 600;
        }
        .chip-label { color: #111827; font-weight: 800; text-transform: uppercase; font-size: 10px; letter-spacing: .08em; }
        .chip-muted { background: #eef2f7; }
        table { width: 100%; border-collapse: collapse; }
        th { text-align: left; padding: 14px 16px; background: #f8fafc; color: var(--muted); font-size: 11px; text-transform: uppercase; letter-spacing: .08em; border-bottom: 1px solid var(--line); }
        td { padding: 14px 16px; border-bottom: 1px solid var(--line); vertical-align: top; font-size: 13px; line-height: 1.45; }
        tbody tr:nth-child(even) td { background: #fbfcfe; }
        tr:last-child td { border-bottom: none; }
        .mono { font-family: "Courier New", monospace; white-space: nowrap; }
        .strong { font-weight: 700; color: #111827; }
        .desc { max-width: 280px; }
        .badge {
          display: inline-flex;
          align-items: center;
          padding: 5px 10px;
          border-radius: 999px;
          font-size: 11px;
          font-weight: 700;
          letter-spacing: .03em;
          white-space: nowrap;
        }
        .badge-red { background: rgba(199, 52, 79, .12); color: var(--danger); }
        .badge-amber { background: rgba(217, 119, 6, .12); color: var(--warning); }
        .badge-blue { background: rgba(37, 99, 235, .12); color: var(--info); }
        .badge-green { background: rgba(15, 159, 110, .12); color: var(--ok); }
        .badge-muted { background: #eef2f7; color: #667085; }
        .empty { text-align: center; color: var(--muted); padding: 28px; }
        .footer { margin-top: 14px; color: var(--muted); font-size: 12px; text-align: right; }
        @media print {
          body { background: #fff; }
          .page { max-width: none; padding: 0; }
          .actions { display: none; }
          .hero { box-shadow: none; margin-bottom: 12px; }
          .card { border: none; box-shadow: none; }
          .card-head { padding-left: 0; padding-right: 0; }
          .hero-stats { gap: 6px; }
          .stat { background: rgba(255,255,255,.08); }
          th, td { font-size: 11px; padding: 10px 8px; }
        }
      </style>
    </head>
    <body>
      <div class="page">
        <div class="hero">
          <div>
            <div class="brand">BURAQ logistics</div>
            <div class="title">Отчёт по проблемам</div>
            <div class="meta">Сформирован: {{ exported_at }}</div>
          </div>
          <div>
            <div class="hero-stats">
              <div class="stat">
                <div class="stat-v">{{ rows_count }}</div>
                <div class="stat-l">Инцидентов</div>
              </div>
              <div class="stat">
                <div class="stat-v">{{ exported_at[:10] }}</div>
                <div class="stat-l">Дата отчёта</div>
              </div>
            </div>
            <div class="actions" style="margin-top:14px;justify-content:flex-end">
              <button class="btn btn-light" onclick="window.close()">Закрыть</button>
              <button class="btn btn-dark" onclick="window.print()">Сохранить в PDF</button>
            </div>
          </div>
        </div>
        <div class="card">
          <div class="card-head">
            <div class="card-title">Список инцидентов</div>
            <div class="filters">{{ filters_html|safe }}</div>
          </div>
          <table>
            <thead>
              <tr>
                <th>Выявлено</th>
                <th>Партия</th>
                <th>BL код</th>
                <th>Клиент</th>
                <th>Тип</th>
                <th>Описание</th>
                <th>Статус груза</th>
                <th>Ожидаемая</th>
                <th>Факт дата</th>
              </tr>
            </thead>
            <tbody>{{ body_rows|safe }}</tbody>
          </table>
        </div>
        <div class="footer">BURAQ logistics · Problems export</div>
      </div>
    </body>
    </html>
    """

    return Response(
        render_template_string(
            report_html,
            exported_at=db.current_ts(),
            rows_count=len(rows),
            filters_html=filters_html,
            body_rows=body_rows,
        ),
        mimetype="text/html; charset=utf-8",
    )


@app.route("/api/clients")
@login_required
def api_clients():
    return jsonify(db.get_clients())


@app.route("/api/clients/<path:client_name>")
@login_required
def api_client_detail(client_name):
    detail = db.get_client_detail(client_name)
    if not detail:
        abort(404)
    return jsonify(detail)


@app.route("/api/notifications")
@login_required
def api_notifications():
    limit = int(request.args.get("limit", 30))
    return jsonify(db.get_notifications(limit))


@app.route("/api/moderator-response")
@login_required
def api_moderator_response():
    return jsonify(
        db.get_moderator_response_stats(
            status=(request.args.get("status") or "").strip(),
            date_from=(request.args.get("date_from") or "").strip(),
            date_to=(request.args.get("date_to") or "").strip(),
            role=(request.args.get("role") or "").strip(),
            limit=int(request.args.get("limit", 300)),
        )
    )


@app.route("/api/moderator-response/assignments")
@login_required
def api_moderator_response_assignments():
    return jsonify(
        {
            "groups": db.get_moderator_response_assignment_groups(),
        }
    )


@app.route("/api/moderator-response/assignments", methods=["POST"])
@editor_required
def api_save_moderator_response_assignments():
    data = request.json or {}
    chat_id = str(data.get("chat_id") or "").strip()
    if not chat_id:
        return jsonify({"error": "chat_id обязателен"}), 400
    db.set_chat_response_assignments(
        chat_id,
        moderator_tg_id=(data.get("moderator_tg_id") or "").strip(),
        sales_manager_tg_id=(data.get("sales_manager_tg_id") or "").strip(),
    )
    return jsonify({"ok": True})


@app.route("/api/moderator-response/clear", methods=["POST"])
@editor_required
def api_clear_moderator_response():
    deleted_count = db.clear_moderator_response_requests()
    return jsonify({"ok": True, "deleted_count": deleted_count})


@app.route("/api/communication-rate")
@login_required
def api_communication_rate():
    month_key = (request.args.get("month") or db.current_month_key()).strip()
    return jsonify(
        {
            "summary": db.get_communication_rate_summary(month_key),
            "rows": db.get_communication_rate(month_key),
            "recipients": db.get_communication_recipients(),
            "sent_chat_ids": list(db.get_communication_sent_chat_ids(month_key)),
            "month_key": month_key,
        }
    )


@app.route("/api/communication-rate/template")
@login_required
def api_communication_rate_template():
    return jsonify({"content": db.get_communication_rate_template()})


@app.route("/api/communication-rate/template", methods=["POST"])
@editor_required
def api_save_communication_rate_template():
    data = request.json or {}
    content = (data.get("content") or "").strip()
    if not content:
        return jsonify({"error": "Шаблон опроса не может быть пустым"}), 400
    db.save_communication_rate_template(content)
    return jsonify({"ok": True})


@app.route("/api/communication-rate/send", methods=["POST"])
@login_required
def api_send_communication_rate():
    if not BOT_TOKEN:
        return jsonify({"error": "BOT_TOKEN не настроен в .env"}), 500

    data = request.json or {}
    month_key = (data.get("month") or db.current_month_key()).strip()
    selected_chat_ids = {str(chat_id) for chat_id in (data.get("chat_ids") or []) if str(chat_id).strip()}
    recipients = db.get_communication_recipients()
    if selected_chat_ids:
        recipients = [item for item in recipients if str(item.get("chat_id", "")) in selected_chat_ids]

    sent = 0
    skipped = 0
    errors = []
    already_sent_chat_ids = set(db.get_communication_sent_chat_ids(month_key))
    processed_chat_ids = set()

    for recipient in recipients:
        chat_id = str(recipient.get("chat_id", ""))
        if not chat_id:
            continue
        if chat_id in processed_chat_ids or chat_id in already_sent_chat_ids:
            skipped += 1
            continue
        try:
            send_communication_survey(recipient, month_key)
            sent += 1
            already_sent_chat_ids.add(chat_id)
            processed_chat_ids.add(chat_id)
        except Exception as exc:
            errors.append(
                {
                    "client": recipient.get("client_name", ""),
                    "chat_id": chat_id,
                    "error": str(exc),
                }
            )

    return jsonify(
        {
            "ok": True,
            "month_key": month_key,
            "sent": sent,
            "skipped": skipped,
            "total_recipients": len(recipients),
            "errors": errors,
        }
    )


@app.route("/api/communication-rate/delete", methods=["POST"])
@editor_required
def api_delete_communication_rate_entry():
    data = request.json or {}
    event_id = data.get("event_id")
    dispatch_id = data.get("dispatch_id")

    if dispatch_id:
        dispatch = db.get_communication_survey_dispatch(dispatch_id)
        if dispatch:
            message_id = str(dispatch.get("message_id") or "").strip()
            chat_id = dispatch.get("chat_id")
            if chat_id and message_id:
                try:
                    telegram_delete_message(chat_id, int(message_id))
                except Exception:
                    app.logger.exception(
                        "Failed to delete communication survey message chat_id=%s message_id=%s",
                        chat_id,
                        message_id,
                    )
        db.delete_communication_survey_dispatch(dispatch_id)
        return jsonify({"ok": True, "deleted": "dispatch"})
    if event_id:
        dispatch = db.get_communication_survey_dispatch_for_event(event_id)
        if dispatch:
            message_id = str(dispatch.get("message_id") or "").strip()
            chat_id = dispatch.get("chat_id")
            if chat_id and message_id:
                try:
                    telegram_delete_message(chat_id, int(message_id))
                except Exception:
                    app.logger.exception(
                        "Failed to delete communication survey message via event chat_id=%s message_id=%s",
                        chat_id,
                        message_id,
                    )
            db.delete_communication_survey_dispatch(dispatch.get("id"))
            return jsonify({"ok": True, "deleted": "dispatch"})
        db.delete_communication_rating_event(event_id)
        return jsonify({"ok": True, "deleted": "event"})
    return jsonify({"error": "Не указано, что удалять"}), 400


@app.route("/api/announcements")
@login_required
def api_announcements():
    attachment = db.get_announcement_attachment()
    recipients = db.get_announcement_recipients()
    return jsonify(
        {
            "content": db.get_announcement_template(),
            "attachment": {
                "filename": attachment.get("filename", ""),
                "kind": attachment.get("kind", ""),
            } if attachment else {},
            "recipients": recipients,
            # In-transit batches with linked groups — the "по партиям" filter
            "batches": db.get_announcement_batches(),
            "send_options": db.get_announcement_send_options(),
            "summary": {
                "groups": len(recipients),
                "has_attachment": bool(attachment),
                "last_sent_at": db.get_announcement_last_sent_at(),
            },
        }
    )


@app.route("/api/announcements/template", methods=["POST"])
@editor_required
def api_save_announcement_template():
    data = request.json or {}
    content = (data.get("content") or "").strip()
    db.save_announcement_template(content)
    return jsonify({"ok": True})


@app.route("/api/announcements/attachment", methods=["POST"])
@editor_required
def api_upload_announcement_attachment():
    if "file" not in request.files:
        return jsonify({"error": "Файл не выбран"}), 400

    uploaded_file = request.files["file"]
    if not uploaded_file or not uploaded_file.filename:
        return jsonify({"error": "Выбери файл или фото для загрузки"}), 400

    ext = uploaded_file.filename.rsplit(".", 1)[-1].lower() if "." in uploaded_file.filename else ""
    if ext not in ANNOUNCEMENT_ALLOWED_EXT:
        return jsonify({"error": f"Тип файла .{ext} не разрешён"}), 400

    original_filename = (uploaded_file.filename or "").strip()
    storage_name = secure_filename(original_filename) or f"announcement_{secrets.token_hex(4)}.{ext or 'bin'}"
    filename = original_filename or storage_name
    unique = f"announcement_{secrets.token_hex(4)}_{storage_name}"
    file_path = os.path.join(UPLOAD_FOLDER, unique)
    uploaded_file.save(file_path)
    if ext == "gif":
        kind = "animation"
    elif ext in {"mp4", "mov", "m4v"}:
        kind = "video"
    elif ext in {"png", "jpg", "jpeg", "webp"}:
        kind = "photo"
    else:
        kind = "document"
    db.save_announcement_attachment(filename, file_path, kind)
    return jsonify({"ok": True, "attachment": {"filename": filename, "kind": kind}})


@app.route("/api/announcements/send-options", methods=["POST"])
@editor_required
def api_save_announcement_send_options():
    data = request.json or {}
    db.save_announcement_send_options(
        data.get("delivery_mode") or "",
        data.get("media_order") or "",
    )
    return jsonify({"ok": True, "send_options": db.get_announcement_send_options()})


@app.route("/api/announcements/attachment", methods=["DELETE"])
@editor_required
def api_delete_announcement_attachment():
    db.clear_announcement_attachment()
    return jsonify({"ok": True})


@app.route("/api/announcements/send", methods=["POST"])
@editor_required
def api_send_announcements():
    if not BOT_TOKEN:
        return jsonify({"error": "BOT_TOKEN не настроен в .env"}), 500

    data = request.json or {}
    # Empty text is OK as long as there's an attachment — the file gets
    # sent on its own (no caption, no follow-up message). If both are
    # empty we have nothing to send and reject early.
    content = (data.get("content") or "").strip()
    if not content:
        content = db.get_announcement_template().strip()
    attachment = db.get_announcement_attachment() or {}
    has_attachment = bool((attachment.get("file_path") or "").strip())
    if not content and not has_attachment:
        return jsonify({"error": "Нужно либо текст, либо вложение"}), 400

    selected_chat_ids = [str(chat_id).strip() for chat_id in (data.get("chat_ids") or []) if str(chat_id).strip()]
    if not selected_chat_ids:
        return jsonify({"error": "Выбери хотя бы одну группу"}), 400

    recipients_map = {
        str(item.get("chat_id") or "").strip(): item
        for item in db.get_announcement_recipients()
        if str(item.get("chat_id") or "").strip()
    }
    send_options = db.get_announcement_send_options()
    sent = 0
    skipped = 0
    errors = []

    for chat_id in selected_chat_ids:
        recipient = recipients_map.get(chat_id)
        if not recipient:
            skipped += 1
            errors.append({"chat_id": chat_id, "error": "Группа не найдена или неактивна"})
            continue
        try:
            send_announcement_broadcast(
                chat_id,
                content,
                attachment,
                delivery_mode=send_options["delivery_mode"],
                media_order=send_options["media_order"],
            )
            sent += 1
        except Exception as exc:
            errors.append(
                {
                    "chat_id": chat_id,
                    "title": recipient.get("title", ""),
                    "error": str(exc),
                }
            )

    if sent:
        db.mark_announcement_last_sent()

    return jsonify(
        {
            "ok": True,
            "sent": sent,
            "skipped": skipped,
            "total_recipients": len(selected_chat_ids),
            "errors": errors,
        }
    )


# ─────────────────────────────────────────────────────────────────────────
# Scheduled announcements — server-side queue, background scheduler.
# ─────────────────────────────────────────────────────────────────────────
_TASHKENT_TZ_APP = db.TASHKENT_TZ


def _parse_scheduled_at(raw: str) -> str:
    """Normalize user input into 'YYYY-MM-DD HH:MM:SS' Tashkent local.

    Accepts:
      - 'YYYY-MM-DDTHH:MM'        (HTML datetime-local default)
      - 'YYYY-MM-DDTHH:MM:SS'
      - 'YYYY-MM-DD HH:MM'
      - 'YYYY-MM-DD HH:MM:SS'

    Treats the input as TASHKENT LOCAL TIME — never converts to UTC.
    Raises ValueError on bad input or past timestamps (>1 min in the past).
    """
    s = str(raw or "").strip().replace("T", " ")
    if not s:
        raise ValueError("Не указано время отправки")
    fmts = ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M")
    parsed = None
    for f in fmts:
        try:
            parsed = datetime.strptime(s, f)
            break
        except ValueError:
            continue
    if parsed is None:
        raise ValueError("Неверный формат времени (нужен YYYY-MM-DD HH:MM)")
    parsed_with_tz = parsed.replace(tzinfo=_TASHKENT_TZ_APP)
    now = datetime.now(_TASHKENT_TZ_APP)
    if (now - parsed_with_tz).total_seconds() > 60:
        raise ValueError("Время уже прошло")
    return parsed_with_tz.strftime("%Y-%m-%d %H:%M:%S")


@app.route("/api/announcements/schedule", methods=["POST"])
@editor_required
def api_schedule_announcement():
    """Queue an announcement to fire at a specific Tashkent-local time.

    Body:
      {
        "content":      "...",            (optional — falls back to template)
        "chat_ids":     [str, ...],
        "scheduled_at": "YYYY-MM-DDTHH:MM" (Tashkent local)
      }
    Returns: { ok: true, schedule: {...} }

    Snapshot semantics — text + chat list + current attachment metadata
    are FROZEN at the moment of scheduling. Changing the global template
    or attachment afterwards does NOT affect a pending job.
    """
    if not BOT_TOKEN:
        return jsonify({"error": "BOT_TOKEN не настроен в .env"}), 500

    data = request.json or {}
    content = (data.get("content") or "").strip()
    if not content:
        content = db.get_announcement_template().strip()
    attachment = db.get_announcement_attachment() or {}
    has_attachment = bool((attachment.get("file_path") or "").strip())
    if not content and not has_attachment:
        return jsonify({"error": "Нужно либо текст, либо вложение"}), 400

    chat_ids = [str(c).strip() for c in (data.get("chat_ids") or []) if str(c).strip()]
    if not chat_ids:
        return jsonify({"error": "Выбери хотя бы одну группу"}), 400

    try:
        scheduled_at = _parse_scheduled_at(data.get("scheduled_at") or "")
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    # Delivery options are frozen into the snapshot too — changing the
    # global settings later must not affect an already-scheduled job.
    send_options = db.get_announcement_send_options()
    snapshot = {
        "filename": attachment.get("filename") or "",
        "file_path": attachment.get("file_path") or "",
        "kind": attachment.get("kind") or "",
        "delivery_mode": send_options["delivery_mode"],
        "media_order": send_options["media_order"],
    } if has_attachment else {}

    sid = db.create_announcement_schedule(
        content=content,
        chat_ids=chat_ids,
        attachment_snapshot=snapshot,
        scheduled_at=scheduled_at,
        created_by=str(session.get("username") or ""),
    )
    return jsonify({
        "ok": True,
        "schedule": {
            "id": sid,
            "scheduled_at": scheduled_at,
            "chat_count": len(chat_ids),
            "has_attachment": has_attachment,
        },
    })


@app.route("/api/announcements/schedules", methods=["GET"])
@login_required
def api_list_announcement_schedules():
    rows = db.list_announcement_schedules(limit=100)
    # Don't leak file paths in the response — frontend doesn't need them.
    for r in rows:
        snap = r.get("attachment_snapshot") or {}
        if isinstance(snap, dict):
            r["attachment_snapshot"] = {
                "filename": snap.get("filename") or "",
                "kind": snap.get("kind") or "",
                "had_attachment": bool(snap.get("file_path")),
            }
    return jsonify({"ok": True, "schedules": rows})


@app.route("/api/announcements/schedules/<int:schedule_id>", methods=["DELETE"])
@editor_required
def api_cancel_announcement_schedule(schedule_id: int):
    ok = db.cancel_announcement_schedule(schedule_id)
    if not ok:
        return jsonify({"error": "Эту запись нельзя отменить (уже отправлена или не найдена)"}), 400
    return jsonify({"ok": True})


# Background scheduler ─────────────────────────────────────────────────
# Wakes up every 30 sec, picks up due rows, fires them. Threading.Event
# is used as a sleep primitive so a future shutdown() can wake the loop
# instantly. Idempotent boot — safe to call twice (the lock makes it a
# no-op on the second call).
_ANNOUNCEMENT_SCHEDULER_LOCK = threading.Lock()
_ANNOUNCEMENT_SCHEDULER_STARTED = False
_ANNOUNCEMENT_SCHEDULER_STOP = threading.Event()


def _fire_scheduled_announcement(row: dict) -> tuple[int, int, str]:
    """Send one scheduled broadcast. Returns (sent_count, failed_count, last_error)."""
    content = (row.get("content") or "").strip()
    chat_ids = [str(c).strip() for c in (row.get("chat_ids") or []) if str(c).strip()]
    snap = row.get("attachment_snapshot") or {}
    if not isinstance(snap, dict):
        snap = {}
    file_path = (snap.get("file_path") or "").strip()
    attachment = None
    if file_path and os.path.exists(file_path):
        attachment = {
            "filename": snap.get("filename") or "",
            "file_path": file_path,
            "kind": snap.get("kind") or "",
        }
    delivery_mode = (snap.get("delivery_mode") or "together").strip() or "together"
    media_order = (snap.get("media_order") or "media_first").strip() or "media_first"
    if not content and not attachment:
        return 0, len(chat_ids), "Ни текста, ни вложения на момент отправки"

    sent = 0
    failed = 0
    last_err = ""
    for chat_id in chat_ids:
        try:
            send_announcement_broadcast(
                chat_id,
                content,
                attachment,
                delivery_mode=delivery_mode,
                media_order=media_order,
            )
            sent += 1
        except Exception as exc:
            failed += 1
            last_err = str(exc)[:300]
            app.logger.exception(
                "Scheduled announcement to %s failed (schedule_id=%s)",
                chat_id, row.get("id"),
            )
    return sent, failed, last_err


def _announcement_scheduler_loop() -> None:
    # Revive any rows left in 'sending' state from a crash before the
    # main loop starts (otherwise they'd stay stuck forever).
    try:
        revived = db.reset_orphaned_sending_announcements()
        if revived:
            app.logger.info("Scheduler revived %d orphaned sending row(s)", revived)
    except Exception:
        app.logger.exception("Failed to revive orphaned schedules at boot")

    while not _ANNOUNCEMENT_SCHEDULER_STOP.is_set():
        try:
            due = db.due_announcement_schedules()
            for row in due:
                sent, failed, last_err = _fire_scheduled_announcement(row)
                status = "sent" if (sent > 0 and failed == 0) else (
                    "failed" if (sent == 0 and failed > 0) else "sent"
                )
                db.mark_announcement_schedule_result(
                    row["id"],
                    status=status,
                    sent=sent,
                    failed=failed,
                    error=last_err,
                )
                if sent > 0:
                    try:
                        db.mark_announcement_last_sent()
                    except Exception:
                        pass
        except Exception:
            # Never let the scheduler thread die from a transient error —
            # log and keep ticking.
            app.logger.exception("Announcement scheduler tick error")

        # Wake up every 30 sec or sooner if shutdown requested.
        _ANNOUNCEMENT_SCHEDULER_STOP.wait(timeout=30.0)


def start_announcement_scheduler_once() -> None:
    """Idempotent boot of the background scheduler thread."""
    global _ANNOUNCEMENT_SCHEDULER_STARTED
    with _ANNOUNCEMENT_SCHEDULER_LOCK:
        if _ANNOUNCEMENT_SCHEDULER_STARTED:
            return
        t = threading.Thread(
            target=_announcement_scheduler_loop,
            name="announcement-scheduler",
            daemon=True,
        )
        t.start()
        _ANNOUNCEMENT_SCHEDULER_STARTED = True
        app.logger.info("Announcement scheduler thread started")


# Boot the scheduler at import time so it runs even when Flask hasn't
# received its first request yet. The thread is a daemon — it dies with
# the process and survives nothing (which is what we want for a single-
# replica deployment).
try:
    start_announcement_scheduler_once()
except Exception:
    # Don't block app boot on a scheduler hiccup.
    pass


@app.route("/analytics/api/overview")
@login_required
def analytics_api_overview():
    return jsonify(analytics_service.get_overview(request.args))


@app.route("/analytics/api/sales-growth")
@login_required
def analytics_api_sales_growth():
    return jsonify(analytics_service.get_sales_growth(request.args))


@app.route("/analytics/api/cashflow")
@login_required
def analytics_api_cashflow():
    return jsonify(analytics_service.get_cashflow(request.args))


@app.route("/analytics/api/managers")
@login_required
def analytics_api_managers():
    return jsonify(analytics_service.get_managers(request.args))


@app.route("/analytics/api/logists")
@login_required
def analytics_api_logists():
    return jsonify(analytics_service.get_logists(request.args))


@app.route("/analytics/api/shipments")
@login_required
def analytics_api_shipments():
    return jsonify(analytics_service.get_shipments(request.args))


@app.route("/analytics/api/debts")
@login_required
def analytics_api_debts():
    return jsonify(analytics_service.get_debts(request.args))


@app.route("/analytics/api/sync/status")
@login_required
def analytics_api_sync_status():
    return jsonify(analytics_service.get_sync_settings_payload())


@app.route("/api/director/fura-plan-xlsx", methods=["POST"])
@login_required
def api_fura_plan_xlsx():
    """Export a truck loading plan (from the fura widget) as xlsx in the
    company's loading-plan format: rows colored by weight category,
    sorted by kg/m³, TOTAL row with truck-limit percentages."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    data = request.json or {}
    rows = data.get("rows") or []
    title = str(data.get("title") or "Fura plan").strip()[:80]
    truck = data.get("truck") or {}
    truck_v = float(truck.get("V") or 0)
    truck_w = float(truck.get("W") or 0)

    wb = Workbook()
    ws = wb.active
    ws.title = "Plan"

    headers = ["№", "Клиент / код", "Кол-во мест", "Вес, кг", "Объём, м³",
               "Кг в 1 м³", "Категория", "Склад", "Дата приёмки"]
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1).font = Font(bold=True, size=13)
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=2, column=col)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    tot_kg = tot_cbm = tot_places = 0.0
    for i, r in enumerate(rows):
        kg = float(r.get("kg") or 0)
        cbm = float(r.get("cbm") or 0)
        tot_kg += kg
        tot_cbm += cbm
        try:
            tot_places += float(str(r.get("places") or "0").replace(",", "."))
        except ValueError:
            pass
        ws.append([
            r.get("n") or (i + 1),
            str(r.get("client") or ""),
            str(r.get("places") or ""),
            round(kg),
            round(cbm, 2),
            round(float(r.get("density") or 0)),
            str(r.get("cat_label") or ""),
            str(r.get("wh") or ""),
            str(r.get("date") or ""),
        ])
        color = str(r.get("color") or "").lstrip("#")
        if len(color) == 6:
            fill = PatternFill(start_color=color.upper(), end_color=color.upper(), fill_type="solid")
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).fill = fill

    avg_d = (tot_kg / tot_cbm) if tot_cbm > 0 else 0
    pct = ""
    if truck_v > 0 and truck_w > 0:
        pct = f"объём {round(tot_cbm / truck_v * 100)}% · вес {round(tot_kg / truck_w * 100)}% от лимитов"
    ws.append(["", "TOTAL", round(tot_places) if tot_places else "",
               round(tot_kg), round(tot_cbm, 2), round(avg_d), pct, "", ""])
    for col in range(1, len(headers) + 1):
        ws.cell(row=ws.max_row, column=col).font = Font(bold=True)

    widths = [5, 24, 12, 11, 11, 10, 13, 13, 13]
    for i, wdt in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = wdt

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    safe_name = re.sub(r"[^\w\-]+", "_", title)[:40] or "fura_plan"
    return send_file(
        bio,
        as_attachment=True,
        download_name=f"{safe_name}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/analytics/api/plans", methods=["GET"])
@login_required
def analytics_api_plans():
    return jsonify({"plans": analytics_service.list_sales_plans()})


@app.route("/analytics/api/plans", methods=["POST"])
@editor_required
def analytics_api_plans_save():
    payload = request.json or {}
    try:
        return jsonify(analytics_service.save_sales_plan(payload))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.route("/analytics/api/plans/<int:plan_id>/activate", methods=["POST"])
@editor_required
def analytics_api_plans_activate(plan_id: int):
    return jsonify(analytics_service.activate_sales_plan(plan_id))


@app.route("/analytics/api/plans/<int:plan_id>", methods=["DELETE"])
@editor_required
def analytics_api_plans_delete(plan_id: int):
    return jsonify(analytics_service.delete_sales_plan(plan_id))


@app.route("/analytics/api/sync/config", methods=["GET"])
@login_required
def analytics_api_sync_config():
    return jsonify(analytics_service.get_sync_settings_payload())


@app.route("/analytics/api/sync/config", methods=["POST"])
@editor_required
def analytics_api_sync_config_save():
    data = request.json or {}
    sheet_id = (data.get("sheet_id") or "").strip()
    analytics_importer.set_google_sheet_id(sheet_id)
    return jsonify({"ok": True, "sheet_id": sheet_id, "status": analytics_service.get_sync_settings_payload()})


@app.route("/analytics/api/sync/google", methods=["POST"])
@editor_required
def analytics_api_sync_google():
    data = request.json or {}
    sheet_id = (data.get("sheet_id") or "").strip() or None
    try:
        result = analytics_importer.sync_from_google(sheet_id)
        return jsonify({"ok": True, **result, "status": analytics_service.get_sync_settings_payload()})
    except analytics_importer.AnalyticsImporterError as exc:
        return jsonify({"error": str(exc), "status": analytics_service.get_sync_settings_payload()}), 400
    except Exception as exc:
        app.logger.exception("Google Sheets sync failed")
        return jsonify({"error": f"Google Sheets sync failed: {exc}"}), 500


@app.route("/analytics/api/sync/upload", methods=["POST"])
@editor_required
def analytics_api_sync_upload():
    upload = request.files.get("file")
    if not upload or not upload.filename:
        return jsonify({"error": "CSV/XLSX fayl tanlanmagan"}), 400
    try:
        result = analytics_importer.sync_from_upload(upload)
        return jsonify({"ok": True, **result, "status": analytics_service.get_sync_settings_payload()})
    except analytics_importer.AnalyticsImporterError as exc:
        return jsonify({"error": str(exc), "status": analytics_service.get_sync_settings_payload()}), 400
    except Exception as exc:
        app.logger.exception("Analytics file import failed")
        return jsonify({"error": f"Import failed: {exc}"}), 500


@app.route("/analytics/api/export")
@login_required
def analytics_api_export():
    report_type = (request.args.get("report") or "sales").strip()
    export_format = (request.args.get("format") or "csv").strip().lower()
    try:
        filename_prefix, rows = analytics_service.get_export_dataset(report_type, request.args)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    if export_format == "xlsx":
        try:
            filename, content = report_exporter.export_xlsx(filename_prefix, rows)
        except RuntimeError as exc:
            return jsonify({"error": str(exc)}), 400
        return send_file(
            io.BytesIO(content),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )

    filename, content = report_exporter.export_csv(filename_prefix, rows)
    return Response(
        content,
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.route("/analytics/api/monitor")
@login_required
def analytics_api_monitor():
    return jsonify(monitor_service.get_monitor_payload(request.args))


@app.route("/analytics/api/monitor/month/<ym>")
@login_required
def analytics_api_monitor_month(ym: str):
    """Per-month breakdown for the Oylik dinamika click-popup.

    ym is "YYYY-MM" in Tashkent local. Returns separate SAVDO + LOGISTIKA
    leaderboards for that month — same shape the rotating panels use, so
    the frontend can reuse its existing rendering.

    Wraps every failure mode in JSON so the kiosk page never gets an HTML
    error page (which the JS can't parse and surfaces as
    "Unexpected token '<'").
    """
    try:
        payload = analytics_service.get_monitor_month_breakdown(ym, request.args)
        return jsonify(payload)
    except ValueError as exc:
        return jsonify({"error": str(exc), "kind": "validation"}), 400
    except Exception as exc:
        app.logger.exception("monitor month breakdown failed for ym=%s", ym)
        return jsonify({
            "error": f"Server error: {str(exc)[:300]}",
            "kind": "server",
        }), 500


# ──────────────────────────────────────────────────────────────────────────
# Director dashboard endpoints
# ──────────────────────────────────────────────────────────────────────────
DIRECTOR_DEFAULT_COLUMNS = {
    # SELIY = whole-truck (FTL). Department split is name-based exactly
    # like Sales Monitor: each row's 'sotuvchi' (seller) name is checked
    # against 3 hardcoded LOGIST names (Sayfullayev / O'ktamov /
    # Abdullayev) — match → Logistika, otherwise → Savdo. No explicit
    # department column is needed.
    "savdo_seliy": {
        "date_col":   "A",
        "logist_col": "C",
        "trucks_col": "D",
        "client_col": "E",
        "agent_col":  "F",
    },
    # SBORNIY = consolidated (LTL) — by seller. Defaults match the Ombor
    # sheet shape Sales Monitor uses (cbm V, date Z, seller AG, bl E,
    # header_rows 2). User can override per-sheet.
    # 'fura_*' fields point at the separate 'Fura statuslari' tab in the
    # same spreadsheet — used to build the agent ranking shown under the
    # sellers leaderboard.
    "savdo_sborniy": {
        "date_col":      "Z",
        "seller_col":    "AG",
        "cbm_col":       "V",
        "bl_col":        "E",
        "fura_sheet_name": "Fura statuslari",
        "fura_header_rows": "1",
        "fura_agent_col": "B",
        "fura_date_col":  "",   # optional — blank = auto-detect date column from the header ("sana"/"date"/"дата")
        # Weight-category leaderboard (Eng yengil … Eng og'ir). Sales rows
        # move between 3 tabs during their lifecycle (Ombor → Ortilgan
        # furalar → Yetib keldi), so all 3 are read and merged. Each tab
        # has its own column layout.
        "vazn_sheet_url":    "",   # optional — blank = use the main sheet_url
        # ─── Vazn stage 1: Ombor ───
        "vazn1_sheet_name":  "Ombor",
        "vazn1_seller_col":  "AG",
        "vazn1_weight_col":  "",
        "vazn1_bl_col":      "E",
        "vazn1_date_col":    "Z",
        "vazn1_header_rows": "2",
        # ─── Vazn stage 2: Ortilgan furalar ───
        "vazn2_sheet_name":  "Ortilgan furalar",
        "vazn2_seller_col":  "AF",
        "vazn2_weight_col":  "",
        "vazn2_bl_col":      "D",
        "vazn2_date_col":    "Y",
        "vazn2_header_rows": "1",
        # ─── Vazn stage 3: Yetib keldi ───
        "vazn3_sheet_name":  "Yetib keldi",
        "vazn3_seller_col":  "AE",
        "vazn3_weight_col":  "",
        "vazn3_bl_col":      "C",
        "vazn3_date_col":    "X",
        "vazn3_header_rows": "1",
    },
    # SOTUV BAZASI = the KP / contract-price sheet. Each row is a quote:
    # density (M, kg per 1 m³) + the price the seller agreed (N, USD/m³) +
    # volume (K). The agreed price is compared against the official tariff
    # (DIRECTOR_SOTUV_TARIFF) to compute margin / flag losses.
    "sotuv_bazasi": {
        "status_col":  "A",
        "cpa_col":     "B",
        "seller_col":  "C",
        "client_col":  "E",
        "product_col": "F",
        "volume_col":  "K",
        "density_col": "M",
        "price_col":   "N",
        "date_col":    "AJ",
    },
    # OMBOR = warehouse fill. Aggregates CBM by warehouse name (matched
    # against YIWU/ZHONGSHAN/HORGOS substring). Capacity per warehouse
    # is stored alongside the column letters in columns_json. Also includes
    # 4 sub-metric sources (ortilgan/hajm/yulda/bojxona) each with optional
    # override URL — if blank, uses the main sheet_url.
    "ombor": {
        "date_col":          "Z",
        "cbm_col":           "V",
        "bl_col":            "E",
        "warehouse_col":     "T",
        "capacity_yiwu":      "1000",
        "capacity_zhongshan": "1000",
        "capacity_horgos":    "1000",
        # ─── Vazn kategoriyalari (Ombor'dagi yuklar) ───
        # Weight column on the SAME Ombor tab; snapshot like the gauges —
        # groups current stock into the 5 weight brackets.
        "vazn_weight_col": "",
        # ─── 1) UMUMIY ORTILGAN YUKLAR ───
        "ortilgan_sheet_url":    "",
        "ortilgan_sheet_name":   "Ortilgan furalar",
        "ortilgan_date_col":     "Y",
        "ortilgan_cbm_col":      "U",
        "ortilgan_header_rows":  "1",
        # ─── 2) UMUMIY HAJM ───
        "hajm_sheet_url":    "",
        "hajm_sheet_name":   "Ombor",
        "hajm_date_col":     "Z",
        "hajm_cbm_col":      "V",
        "hajm_header_rows":  "2",
        # ─── 3) YO'LDAGI YUKLAR ───
        # Counts cargos currently in transit: departure_date passed AND
        # arrival_date is empty or in the future. m³ shown as subtitle.
        "yulda_sheet_url":      "",
        "yulda_sheet_name":     "",
        "yulda_departure_col":  "",
        "yulda_arrival_col":    "",
        "yulda_cbm_col":        "",
        "yulda_header_rows":    "1",
        # ─── 4) BOJXONADAGI YUKLAR ───
        # Same transit logic: customs_in date passed AND customs_out empty
        # or in the future.
        "bojxona_sheet_url":     "",
        "bojxona_sheet_name":    "",
        "bojxona_departure_col": "",
        "bojxona_arrival_col":   "",
        "bojxona_cbm_col":       "",
        "bojxona_header_rows":   "1",
    },
    "agentlar": {
        "date_col":   "A",
        "agent_col":  "B",
        "amount_col": "C",
        "deals_col":  "D",
        "client_col": "E",
    },
}


@app.route("/api/director/config", methods=["GET"])
@login_required
def api_director_config_get():
    if session.get("role") == ROLE_KIOSK:
        return jsonify({"error": "Forbidden"}), 403
    try:
        configs = db.get_director_config()
        return jsonify({"sections": configs, "defaults": DIRECTOR_DEFAULT_COLUMNS})
    except Exception as exc:
        app.logger.exception("director config get failed")
        return jsonify({"error": f"Server error: {str(exc)[:300]}"}), 500


@app.route("/api/director/config", methods=["POST"])
@editor_required
def api_director_config_save():
    if session.get("role") == ROLE_KIOSK:
        return jsonify({"error": "Forbidden"}), 403
    data = request.get_json(silent=True) or {}
    section = (data.get("section") or "").strip().lower()
    if section not in db.DIRECTOR_SECTIONS:
        return jsonify({"error": f"Unknown section: {section}"}), 400
    try:
        result = db.save_director_config(
            section=section,
            sheet_url=str(data.get("sheet_url") or ""),
            sheet_name=str(data.get("sheet_name") or ""),
            header_rows=int(data.get("header_rows") or 1),
            columns=data.get("columns") if isinstance(data.get("columns"), dict) else {},
            updated_by=session.get("username") or "",
        )
        return jsonify({"ok": True, "config": result})
    except Exception as exc:
        app.logger.exception("director config save failed")
        return jsonify({"error": f"Server error: {str(exc)[:300]}"}), 500


@app.route("/api/director/<section>/data", methods=["GET"])
@login_required
def api_director_section_data(section: str):
    """Return aggregated KPI + chart series for one section.

    For now returns a placeholder shape so the UI can be built and tested.
    When the director attaches a real Google Sheet, this will fetch the CSV
    and aggregate by the configured columns.
    """
    if session.get("role") == ROLE_KIOSK:
        return jsonify({"error": "Forbidden"}), 403
    if section not in db.DIRECTOR_SECTIONS:
        return jsonify({"error": f"Unknown section: {section}"}), 400
    date_from = (request.args.get("from") or "").strip()
    date_to   = (request.args.get("to") or "").strip()
    try:
        cfg = db.get_director_config(section)
        if not cfg.get("sheet_id"):
            return jsonify({
                "section": section,
                "configured": False,
                "message": "Sheet manbasini sozlang (⚙).",
                "from": date_from,
                "to": date_to,
                "kpis": [],
                "charts": {},
            })
        # SAVDO · SELIY — real aggregation (same logic as Sales Monitor's FTL)
        if section == "savdo_seliy":
            agg = analytics_service.get_director_seliy(cfg, date_from, date_to)
            agg.setdefault("section", section)
            agg.setdefault("from", date_from)
            agg.setdefault("to", date_to)
            agg.setdefault("sheet_id", cfg.get("sheet_id"))
            return jsonify(agg)
        # SAVDO · SBORNIY — LTL aggregation via Ombor (no FTL trucks)
        if section == "savdo_sborniy":
            agg = analytics_service.get_director_sborniy(cfg, date_from, date_to)
            agg.setdefault("section", section)
            agg.setdefault("from", date_from)
            agg.setdefault("to", date_to)
            agg.setdefault("sheet_id", cfg.get("sheet_id"))
            return jsonify(agg)
        # OMBOR — per-warehouse fill (YIWU / ZHONGSHAN / HORGOS) + daily flow
        if section == "ombor":
            agg = analytics_service.get_director_ombor(cfg, date_from, date_to)
            agg.setdefault("section", section)
            agg.setdefault("from", date_from)
            agg.setdefault("to", date_to)
            agg.setdefault("sheet_id", cfg.get("sheet_id"))
            return jsonify(agg)
        # SOTUV BAZASI — KP / contract prices vs the official tariff
        if section == "sotuv_bazasi":
            agg = analytics_service.get_director_sotuv_bazasi(
                cfg, date_from, date_to,
                seller_filter=(request.args.get("seller") or "").strip(),
                exclude_lighting=(request.args.get("no_lighting") or "") == "1",
            )
            agg.setdefault("section", section)
            agg.setdefault("from", date_from)
            agg.setdefault("to", date_to)
            agg.setdefault("sheet_id", cfg.get("sheet_id"))
            return jsonify(agg)
        return jsonify({
            "section": section,
            "configured": True,
            "from": date_from,
            "to": date_to,
            "sheet_id": cfg.get("sheet_id"),
            "kpis": [],
            "charts": {},
            "message": "Manba ulangan. Ko'rsatkichlar ustun mapping tasdiqlangach to'ldiriladi.",
        })
    except Exception as exc:
        app.logger.exception("director section data failed: %s", section)
        return jsonify({"error": f"Server error: {str(exc)[:300]}"}), 500


# ──────────────────────────────────────────────────────────────────────────
# Agent API — read-only JSON endpoints for the Claude-powered Telegram
# agent. Auth: static bearer token from the AGENT_API_TOKEN env var
# (send as "Authorization: Bearer <token>" or "X-API-Key: <token>").
# These endpoints reuse the exact aggregation functions the admin panel
# uses (get_monitor / get_director_*), so the agent's numbers always
# match the dashboards — the panel stays the single source of truth.
# See docs/AGENT_API.md for the Claude tool definitions + system prompt.
# ──────────────────────────────────────────────────────────────────────────
AGENT_API_TOKEN = os.getenv("AGENT_API_TOKEN", "").strip()

# Agent-facing section slugs → internal director section keys.
AGENT_DIRECTOR_SECTIONS = {
    "seliy":   "savdo_seliy",
    "sborniy": "savdo_sborniy",
    "ombor":   "ombor",
    "sotuv":   "sotuv_bazasi",
}


def agent_token_required(func):
    @wraps(func)
    def decorated(*args, **kwargs):
        if not AGENT_API_TOKEN:
            return jsonify({
                "error": "Agent API o'chirilgan: serverda AGENT_API_TOKEN env o'zgaruvchisini o'rnating."
            }), 503
        auth = request.headers.get("Authorization", "")
        if auth.lower().startswith("bearer "):
            supplied = auth[7:].strip()
        else:
            supplied = (request.headers.get("X-API-Key") or "").strip()
        if not supplied or not hmac.compare_digest(supplied, AGENT_API_TOKEN):
            return jsonify({"error": "Invalid or missing API token"}), 401
        return func(*args, **kwargs)

    return decorated


def _agent_scrub_charts(obj):
    """Recursively drop nested Chart.js specs ("chart" keys) — UI-only noise
    for an LLM consumer. Top-level "charts" is converted to plain series by
    the caller before this runs."""
    if isinstance(obj, dict):
        obj.pop("chart", None)
        for v in obj.values():
            _agent_scrub_charts(v)
    elif isinstance(obj, list):
        for v in obj:
            _agent_scrub_charts(v)


def _agent_charts_to_series(payload: dict) -> dict:
    """Replace the top-level Chart.js "charts" object with agent-friendly
    {title: {labels, values}} series (keeps the daily/top-N data, drops
    colors and rendering options)."""
    charts = payload.pop("charts", None)
    if isinstance(charts, dict):
        series = {}
        for key, spec in charts.items():
            if not isinstance(spec, dict):
                continue
            datasets = spec.get("datasets") or []
            first = datasets[0] if datasets and isinstance(datasets[0], dict) else {}
            series[spec.get("title") or key] = {
                "labels": spec.get("labels") or [],
                "values": first.get("data") or [],
            }
        payload["series"] = series
    return payload


@app.route("/api/agent/v1/overview")
@agent_token_required
def agent_api_overview():
    """Entry point for the agent: what data exists and how to ask for it."""
    try:
        plans = analytics_service.list_sales_plans()
    except Exception:
        plans = []
    plans_out = [
        {
            "id": int(p.get("id") or 0),
            "name": p.get("name") or "",
            "period_start": p.get("period_start") or "",
            "period_end": p.get("period_end") or "",
            "target_value": p.get("target_value") or p.get("target_amount_usd") or 0,
            "is_active": bool(int(p.get("is_active") or 0)),
        }
        for p in plans
    ]
    sections = {}
    for slug, internal in AGENT_DIRECTOR_SECTIONS.items():
        try:
            cfg = db.get_director_config(internal)
            sections[slug] = {"configured": bool(cfg.get("sheet_id"))}
        except Exception:
            sections[slug] = {"configured": False}
    return jsonify({
        "server_date": datetime.now().date().isoformat(),
        "sales_plans": plans_out,
        "director_sections": sections,
        "endpoints": {
            "sales_monitor": "/api/agent/v1/sales-monitor?plan_id=<id, optional>",
            "director":      "/api/agent/v1/director/<seliy|sborniy|ombor>?from=YYYY-MM-DD&to=YYYY-MM-DD",
        },
    })


@app.route("/api/agent/v1/sales-monitor")
@agent_token_required
def agent_api_sales_monitor():
    """Sales Monitor payload: plan progress, department leaderboards,
    monthly dynamics. Same function the TV monitor calls."""
    try:
        payload = monitor_service.get_monitor_payload(request.args)
    except Exception as exc:
        app.logger.exception("agent sales-monitor failed")
        return jsonify({"error": f"Server error: {str(exc)[:300]}"}), 500
    if request.args.get("detail") != "full":
        for key in ("diagnostics", "ftl_diagnostics", "history_diagnostics",
                    "ombor_config", "plan_data_status"):
            payload.pop(key, None)
    return jsonify(payload)


@app.route("/api/agent/v1/director/<section>")
@agent_token_required
def agent_api_director(section: str):
    """Director-panel section data (seliy / sborniy / ombor) for a date
    range. Empty from/to = all time. detail=full keeps per-sale lists
    and diagnostics; default response is compact for LLM consumption."""
    internal = AGENT_DIRECTOR_SECTIONS.get((section or "").strip().lower())
    if not internal:
        return jsonify({
            "error": f"Unknown section '{section}'. Use one of: {', '.join(AGENT_DIRECTOR_SECTIONS)}"
        }), 400
    date_from = (request.args.get("from") or "").strip()
    date_to   = (request.args.get("to") or "").strip()
    for label, val in (("from", date_from), ("to", date_to)):
        if val and not re.fullmatch(r"\d{4}-\d{2}-\d{2}", val):
            return jsonify({"error": f"Invalid '{label}' date '{val}': use YYYY-MM-DD"}), 400
    try:
        cfg = db.get_director_config(internal)
        if not cfg.get("sheet_id"):
            return jsonify({
                "section": section,
                "configured": False,
                "message": "Bo'lim manbasi admin panelda sozlanmagan (⚙).",
            })
        if internal == "savdo_seliy":
            payload = analytics_service.get_director_seliy(cfg, date_from, date_to)
        elif internal == "savdo_sborniy":
            payload = analytics_service.get_director_sborniy(cfg, date_from, date_to)
        elif internal == "sotuv_bazasi":
            payload = analytics_service.get_director_sotuv_bazasi(cfg, date_from, date_to)
        else:
            payload = analytics_service.get_director_ombor(cfg, date_from, date_to)
    except Exception as exc:
        app.logger.exception("agent director section failed: %s", section)
        return jsonify({"error": f"Server error: {str(exc)[:300]}"}), 500

    payload = _agent_charts_to_series(payload)
    if request.args.get("detail") != "full":
        payload.pop("diagnostics", None)
        _agent_scrub_charts(payload)
        # Weight-category drill-down lists are large; keep the aggregates.
        wc = payload.get("weight_categories")
        if isinstance(wc, dict):
            for cat in wc.get("categories") or []:
                for s in cat.get("sellers") or []:
                    s.pop("sales", None)
                    s.pop("bl_codes", None)
        # Sotuv bazasi: the full deal list is heavy — keep the 40 worst
        # (deals are pre-sorted by margin ascending, losses first).
        if isinstance(payload.get("deals"), list) and len(payload["deals"]) > 40:
            payload["deals_total"] = len(payload["deals"])
            payload["deals"] = payload["deals"][:40]
    payload["section"] = section
    payload["from"] = date_from
    payload["to"] = date_to
    return jsonify(payload)


@app.route("/analytics/api/plans/<int:plan_id>/ombor-config", methods=["POST"])
@editor_required
def analytics_api_plan_ombor_config(plan_id: int):
    payload = request.json or {}
    payload["id"] = plan_id
    try:
        existing_plans = analytics_service.list_sales_plans()
        plan = next((p for p in existing_plans if int(p.get("id") or 0) == plan_id), None)
        if not plan:
            return jsonify({"error": "Plan topilmadi"}), 404
        merged = {**plan, **payload}
        return jsonify(analytics_service.save_sales_plan(merged))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.route("/api/template")
@login_required
def api_get_template():
    return jsonify(
        {
            "content": db.get_template(),
            "status_details": db.get_status_details(),
            "statuses": db.STATUSES,
        }
    )


@app.route("/api/template", methods=["POST"])
@editor_required
def api_save_template():
    data = request.json or {}
    content = data.get("content", "").strip()
    if not content:
        return jsonify({"error": "Шаблон не может быть пустым"}), 400

    db.save_template(content)

    details = data.get("status_details", {})
    for status_name, detail in details.items():
        if status_name in db.STATUSES:
            db.save_status_detail(status_name, detail)

    return jsonify({"ok": True})


# Хуки прямого исполнения для владельческого режима ассистента
# (ai_assistant не может импортировать app — цикл).
def configure_telegram_menu_button():
    """Синяя кнопка-меню слева от поля ввода (как «Open» у BotFather):
    в личке бота открывает Mini App формы напрямую."""
    if not BOT_TOKEN or not WEBHOOK_BASE_URL:
        return
    try:
        telegram_api("setChatMenuButton", json={
            "menu_button": {
                "type": "web_app",
                "text": "📝 Forma",
                "web_app": {"url": _tgform_url()},
            }
        })
        app.logger.info("Telegram menu button configured")
    except Exception as exc:
        app.logger.warning("setChatMenuButton failed: %s", exc)
    try:
        # зарегистрированные команды включают в ГРУППАХ нативную кнопку «/»
        # у поля ввода — самый близкий аналог кнопки-меню из лички
        telegram_api("setMyCommands", json={
            "commands": [
                {"command": "form", "description": "📝 Treking formasini ochish"},
                {"command": "formon", "description": "Mini App ni guruhda yoqish (admin)"},
                {"command": "formoff", "description": "Mini App ni guruhda o'chirish (admin)"},
            ]
        })
        app.logger.info("Bot commands registered")
    except Exception as exc:
        app.logger.warning("setMyCommands failed: %s", exc)


def _wire_ai_assistant_hooks():
    from services import ai_assistant as _ai

    _ai.direct_send_message = lambda chat_id, text: telegram_send_message(chat_id, text)
    _ai.direct_send_poll = lambda chat_id, question, options, is_anonymous=False: telegram_api(
        "sendPoll",
        json={
            "chat_id": chat_id,
            "question": str(question)[:300],
            "options": json.dumps([str(o)[:100] for o in options][:10], ensure_ascii=False),
            "is_anonymous": bool(is_anonymous),
        },
    )


_wire_ai_assistant_hooks()


if __name__ == "__main__":
    if BOT_TOKEN and WEBHOOK_BASE_URL:
        try:
            configure_telegram_webhook()
            app.logger.info("Telegram webhook configured")
        except Exception as exc:
            app.logger.warning("Failed to configure Telegram webhook: %s", exc)
        configure_telegram_menu_button()
    else:
        app.logger.warning("Telegram webhook is not configured. Set BOT_TOKEN and WEBHOOK_BASE_URL.")

    # Morning packing-list ask in the control group (daily, Tashkent time)
    threading.Thread(target=_packing_reminder_scheduler, daemon=True).start()

    app.run(host="0.0.0.0", port=PORT, debug=False)
