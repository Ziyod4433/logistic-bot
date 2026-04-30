from __future__ import annotations

import csv
import io
import json
import os
import re
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen
from zoneinfo import ZoneInfo

import database as db

GOOGLE_SHEETS_MISSING_MESSAGE = (
    "Google Sheets ulanishi sozlanmagan. .env faylga GOOGLE_SHEET_ID va "
    "GOOGLE_APPLICATION_CREDENTIALS yoki GOOGLE_SERVICE_ACCOUNT_JSON qo‘shing."
)

ANALYTICS_SHEET_ID_KEY = "analytics_google_sheet_id"
ANALYTICS_LAST_SYNC_KEY = "analytics_last_sync_at"
ANALYTICS_SOURCE_NAME_KEY = "analytics_source_name"
GOOGLE_SHEETS_EXPORT_URL = "https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
INVALID_CELL_MARKERS = {"#NAME?", "#REF!", "#N/A", "#DIV/0!", "#VALUE!"}

try:
    TASHKENT_TZ = getattr(db, "TASHKENT_TZ", None) or ZoneInfo("Asia/Tashkent")
except Exception:
    TASHKENT_TZ = timezone(timedelta(hours=5))


class AnalyticsImporterError(Exception):
    pass


def _clean_text(value: Any) -> str:
    text = str(value or "").replace("\u00a0", " ").strip()
    return re.sub(r"\s+", " ", text)


def _normalize_key(value: Any) -> str:
    text = _clean_text(value).lower()
    text = text.replace("ё", "е")
    text = text.replace("ў", "у").replace("қ", "к").replace("ғ", "г").replace("ҳ", "х")
    text = text.replace("/", " ").replace("\\", " ").replace("-", " ").replace("_", " ")
    text = re.sub(r"[^\w\s%№]+", "", text, flags=re.UNICODE)
    return re.sub(r"\s+", " ", text).strip()


def _extract_sheet_id(sheet_id_or_url: str) -> str:
    raw = _clean_text(sheet_id_or_url)
    if not raw:
        return ""
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", raw)
    if match:
        return match.group(1)
    if re.fullmatch(r"[a-zA-Z0-9-_]{20,}", raw):
        return raw
    return raw


def _now_tashkent() -> datetime:
    return datetime.now(TASHKENT_TZ)


def _format_dt(value: datetime | None = None) -> str:
    current = value or _now_tashkent()
    if current.tzinfo is None:
        current = current.replace(tzinfo=TASHKENT_TZ)
    else:
        current = current.astimezone(TASHKENT_TZ)
    return current.strftime("%Y-%m-%d %H:%M:%S")


def _safe_json(data: Any) -> str:
    return json.dumps(data, ensure_ascii=False, default=str)


def _excel_serial_to_date(value: float) -> str:
    try:
        base = datetime(1899, 12, 30)
        converted = base + timedelta(days=float(value))
        return converted.strftime("%d.%m.%Y")
    except Exception:
        return ""


def _parse_date(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, (int, float)):
        if 20000 <= float(value) <= 70000:
            return _excel_serial_to_date(float(value))
        return ""
    text = _clean_text(value)
    if not text or text.upper() in INVALID_CELL_MARKERS:
        return ""
    for fmt in (
        "%d.%m.%Y",
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%Y/%m/%d",
        "%d.%m.%Y %H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
    ):
        try:
            return datetime.strptime(text, fmt).strftime("%d.%m.%Y")
        except ValueError:
            continue
    if re.fullmatch(r"\d+(?:[.,]\d+)?", text):
        return _excel_serial_to_date(float(text.replace(",", ".")))
    match = re.search(r"(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})", text)
    if match:
        day, month, year = match.groups()
        if len(year) == 2:
            year = f"20{year}"
        elif len(year) == 3:
            return ""
        try:
            return datetime(int(year), int(month), int(day)).strftime("%d.%m.%Y")
        except ValueError:
            return ""
    return ""


def _parse_date_obj(value: Any) -> date | None:
    normalized = _parse_date(value)
    if not normalized:
        return None
    try:
        return datetime.strptime(normalized, "%d.%m.%Y").date()
    except ValueError:
        return None


def _clean_number_text(value: Any) -> str:
    return str(value or "").replace("\u00a0", "").replace(" ", "").replace(",", ".")


def _parse_float(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = _clean_number_text(value)
    if not text or text.upper() in INVALID_CELL_MARKERS:
        return 0.0
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return 0.0
    try:
        return float(match.group(0))
    except ValueError:
        return 0.0


def _parse_int(value: Any) -> int:
    return int(round(_parse_float(value)))


def _normalize_currency(value: Any) -> str:
    raw = _normalize_key(value)
    if not raw:
        return ""
    if "usd" in raw or "dollar" in raw:
        return "USD"
    if "yuan" in raw or "rmb" in raw or "cny" in raw:
        return "RMB"
    if "uzs" in raw or "sum" in raw:
        return "UZS"
    return _clean_text(value).upper()


def _read_cell(row: list[Any], index: int) -> Any:
    if index < 0 or index >= len(row):
        return ""
    return row[index]


def _truthy_int(value: Any) -> int:
    text = _clean_text(value).lower()
    if not text:
        return 0
    if text in {"ha", "yes", "true", "1"}:
        return 1
    return 1 if _parse_float(value) > 0 else 0


def _load_workbook_from_bytes(raw: bytes, source_name: str) -> dict[str, list[list[Any]]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise AnalyticsImporterError(f"{source_name} import uchun openpyxl kerak: {exc}")
    workbook = load_workbook(io.BytesIO(raw), data_only=True)
    sheets: dict[str, list[list[Any]]] = {}
    for sheet in workbook.worksheets:
        values: list[list[Any]] = []
        for row in sheet.iter_rows(values_only=True):
            values.append(list(row))
        sheets[sheet.title] = values
    return sheets


def _load_uploaded_workbook(file_storage) -> dict[str, list[list[Any]]]:
    if not file_storage or not getattr(file_storage, "filename", ""):
        raise AnalyticsImporterError("CSV/XLSX fayl tanlanmagan.")
    filename = _clean_text(file_storage.filename)
    ext = filename.rsplit(".", 1)[-1].lower()
    raw = file_storage.read()
    if ext == "csv":
        text = raw.decode("utf-8-sig", errors="replace")
        return {"Imported CSV": list(csv.reader(io.StringIO(text)))}
    if ext not in {"xlsx", "xlsm"}:
        raise AnalyticsImporterError("Faqat CSV yoki XLSX/XLSM fayl yuklash mumkin.")
    return _load_workbook_from_bytes(raw, "XLSX")


def _get_credentials_payload() -> tuple[dict[str, Any] | None, str | None]:
    inline_json = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON", "").strip()
    if inline_json:
        try:
            return json.loads(inline_json), None
        except json.JSONDecodeError as exc:
            raise AnalyticsImporterError(f"GOOGLE_SERVICE_ACCOUNT_JSON noto‘g‘ri: {exc}")
    path = os.getenv("GOOGLE_APPLICATION_CREDENTIALS", "").strip()
    if path:
        payload_path = Path(path)
        if not payload_path.exists():
            raise AnalyticsImporterError("GOOGLE_APPLICATION_CREDENTIALS fayli topilmadi.")
        return None, str(payload_path)
    return None, None


def get_google_sheet_id() -> str:
    return _extract_sheet_id(
        os.getenv("GOOGLE_SHEET_ID", "").strip() or db.get_setting(ANALYTICS_SHEET_ID_KEY, "").strip()
    )


def set_google_sheet_id(sheet_id: str) -> None:
    db.set_setting(ANALYTICS_SHEET_ID_KEY, _extract_sheet_id(sheet_id))


def google_credentials_status() -> dict[str, Any]:
    sheet_id = get_google_sheet_id()
    try:
        payload, path = _get_credentials_payload()
    except AnalyticsImporterError as exc:
        return {"connected": False, "sheet_id": sheet_id, "mode": "invalid", "message": str(exc)}
    if not sheet_id:
        return {
            "connected": False,
            "sheet_id": "",
            "mode": "missing",
            "message": "Google Sheet ID yoki to‘liq havola kiriting.",
        }
    if payload is None and path is None:
        return {
            "connected": True,
            "sheet_id": sheet_id,
            "mode": "public",
            "message": "Google API credentials topilmadi. Public link rejimi ishlaydi, agar jadval havola bo‘yicha ochiq bo‘lsa.",
        }
    return {
        "connected": True,
        "sheet_id": sheet_id,
        "mode": "api",
        "message": "Google Sheets API ulanishi tayyor.",
    }


def _load_google_workbook(sheet_id: str | None = None) -> dict[str, list[list[Any]]]:
    sheet_id = _extract_sheet_id(sheet_id or get_google_sheet_id())
    status = google_credentials_status()
    if not sheet_id or not status["connected"]:
        raise AnalyticsImporterError(status["message"])
    try:
        from google.oauth2.service_account import Credentials
        from googleapiclient.discovery import build
    except Exception as exc:
        raise AnalyticsImporterError(f"Google API kutubxonalari topilmadi: {exc}")

    payload, path = _get_credentials_payload()
    scopes = ["https://www.googleapis.com/auth/spreadsheets.readonly"]
    if payload is not None:
        credentials = Credentials.from_service_account_info(payload, scopes=scopes)
    elif path is not None:
        credentials = Credentials.from_service_account_file(path, scopes=scopes)
    else:
        raise AnalyticsImporterError(GOOGLE_SHEETS_MISSING_MESSAGE)

    service = build("sheets", "v4", credentials=credentials, cache_discovery=False)
    metadata = service.spreadsheets().get(spreadsheetId=sheet_id).execute()
    workbook: dict[str, list[list[Any]]] = {}
    for sheet in metadata.get("sheets", []):
        title = sheet.get("properties", {}).get("title", "")
        if not title:
            continue
        values = (
            service.spreadsheets()
            .values()
            .get(spreadsheetId=sheet_id, range=title, majorDimension="ROWS")
            .execute()
            .get("values", [])
        )
        workbook[title] = values
    return workbook


def _load_public_google_workbook(sheet_id: str | None = None) -> dict[str, list[list[Any]]]:
    sheet_id = _extract_sheet_id(sheet_id or get_google_sheet_id())
    if not sheet_id:
        raise AnalyticsImporterError("Google Sheet ID yoki to‘liq havola kiriting.")
    export_url = GOOGLE_SHEETS_EXPORT_URL.format(sheet_id=sheet_id)
    request = Request(
        export_url,
        headers={
            "User-Agent": "Mozilla/5.0",
            "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
        },
    )
    try:
        with urlopen(request, timeout=30) as response:
            content_type = str(response.headers.get("Content-Type", "")).lower()
            raw = response.read()
    except HTTPError as exc:
        if exc.code in {401, 403}:
            raise AnalyticsImporterError(
                "Google Sheets public export yopiq. Jadvalni havola bo‘yicha ko‘rish uchun oching yoki API credentials qo‘shing."
            ) from exc
        raise AnalyticsImporterError(f"Google Sheets public export xatosi: HTTP {exc.code}") from exc
    except URLError as exc:
        raise AnalyticsImporterError(f"Google Sheets bilan ulanishda xato: {exc.reason}") from exc
    if "text/html" in content_type and b"<html" in raw[:200].lower():
        raise AnalyticsImporterError(
            "Google Sheets public export XLSX qaytarmadi. Jadval havola bo‘yicha ochiq ekanini tekshiring."
        )
    return _load_workbook_from_bytes(raw, "Google Sheets public export")


def _parse_total_sheet(sheet_name: str, matrix: list[list[Any]], source_file_id: str) -> list[dict[str, Any]]:
    if not matrix:
        return []
    header = [_normalize_key(item) for item in matrix[0]]
    if len(header) < 35 or "reys nomer" not in header[:2]:
        return []
    rows: list[dict[str, Any]] = []
    for idx, row in enumerate(matrix[1:], start=2):
        shipping_mark = _clean_text(_read_cell(row, 4)).upper()
        client_name = _clean_text(_read_cell(row, 6))
        sale_date = _parse_date(_read_cell(row, 2)) or _parse_date(_read_cell(row, 3))
        invoice_date = _parse_date(_read_cell(row, 3))
        raw_final = _parse_float(_read_cell(row, 33))
        raw_sale = _parse_float(_read_cell(row, 30))
        final_sale_amount = raw_final if raw_final > 0 else raw_sale
        if not (shipping_mark or client_name or final_sale_amount):
            continue
        rows.append(
            {
                "source_file_id": source_file_id,
                "source_sheet": sheet_name,
                "source_row": idx,
                "reys_number": _clean_text(_read_cell(row, 0)),
                "invoice_date": invoice_date,
                "sale_date": sale_date,
                "shipping_mark": shipping_mark,
                "brand_name": _clean_text(_read_cell(row, 5)),
                "client_name": client_name,
                "phone": _clean_text(_read_cell(row, 7)),
                "client_status": _clean_text(_read_cell(row, 8)),
                "cargo_name": _clean_text(_read_cell(row, 9)),
                "quantity": _parse_float(_read_cell(row, 11)),
                "ctn": _parse_float(_read_cell(row, 12)),
                "cbm": _parse_float(_read_cell(row, 13)),
                "net_weight": _parse_float(_read_cell(row, 14)),
                "gross_weight": _parse_float(_read_cell(row, 15)),
                "customs_payment": _parse_float(_read_cell(row, 24)),
                "company_expense": _parse_float(_read_cell(row, 25)),
                "certificate_expense": _parse_float(_read_cell(row, 26)),
                "client_price": _parse_float(_read_cell(row, 29)),
                "sale_amount": raw_sale,
                "correction_amount": _parse_float(_read_cell(row, 31)),
                "discount_amount": _parse_float(_read_cell(row, 32)),
                "final_sale_amount": final_sale_amount,
                "salesperson": _clean_text(_read_cell(row, 34)),
                "sales_kpi_amount": _parse_float(_read_cell(row, 35)),
                "customs_kpi_amount": _parse_float(_read_cell(row, 36)),
                "raw_data_json": _safe_json({"row": row}),
            }
        )
    return rows


def _parse_cashflow_sheet(sheet_name: str, matrix: list[list[Any]], source_file_id: str) -> list[dict[str, Any]]:
    if _normalize_key(sheet_name) != "cashflow" or not matrix:
        return []
    rows: list[dict[str, Any]] = []
    for idx, row in enumerate(matrix[1:], start=2):
        flow_name = _clean_text(_read_cell(row, 3))
        amount = _parse_float(_read_cell(row, 7))
        if not (flow_name or amount or _clean_text(_read_cell(row, 11))):
            continue
        flow_key = _normalize_key(flow_name)
        if "kirim" in flow_key:
            flow_type = "income"
        elif "chiqim" in flow_key:
            flow_type = "expense"
        else:
            flow_type = "income" if amount >= 0 else "expense"
        rows.append(
            {
                "source_file_id": source_file_id,
                "source_sheet": sheet_name,
                "source_row": idx,
                "created_date": _parse_date(_read_cell(row, 0)),
                "operation_date": _parse_date(_read_cell(row, 1)),
                "wallet": _clean_text(_read_cell(row, 2)),
                "flow_type": flow_type,
                "currency": _normalize_currency(_read_cell(row, 4)) or "USD",
                "comment": _clean_text(_read_cell(row, 5)),
                "category": _clean_text(_read_cell(row, 6)),
                "amount": abs(amount),
                "rate": _parse_float(_read_cell(row, 8)),
                "department": _clean_text(_read_cell(row, 9)),
                "reys_number": _clean_text(_read_cell(row, 10)),
                "counterparty": _clean_text(_read_cell(row, 11)).upper(),
                "auto_confirm": _clean_text(_read_cell(row, 12)),
                "amount_usd": 0.0,
                "raw_data_json": _safe_json({"row": row}),
            }
        )
    return rows


def _parse_kurs_sheet(sheet_name: str, matrix: list[list[Any]], source_file_id: str) -> list[dict[str, Any]]:
    if _normalize_key(sheet_name) != "kurs" or not matrix:
        return []
    rows: list[dict[str, Any]] = []
    for idx, row in enumerate(matrix[1:], start=2):
        rate_date = _parse_date(_read_cell(row, 2))
        currency = _normalize_currency(_read_cell(row, 3))
        rate_to_usd = _parse_float(_read_cell(row, 4))
        if not (currency and rate_to_usd):
            continue
        rows.append(
            {
                "source_file_id": source_file_id,
                "rate_date": rate_date,
                "currency": currency,
                "rate_to_usd": rate_to_usd,
                "raw_data_json": _safe_json({"row": row}),
            }
        )
    return rows


def _parse_kpi_logist_sheet(sheet_name: str, matrix: list[list[Any]], source_file_id: str) -> list[dict[str, Any]]:
    if _normalize_key(sheet_name) != "kpi logist" or not matrix:
        return []
    rows: list[dict[str, Any]] = []
    for idx, row in enumerate(matrix[1:], start=2):
        reys_number = _clean_text(_read_cell(row, 0))
        if not reys_number:
            continue
        warehouse_kpi = _truthy_int(_read_cell(row, 1))
        damage_kpi = _truthy_int(_read_cell(row, 2))
        for position, col_idx in enumerate((3, 4, 5), start=1):
            logist_name = _clean_text(_read_cell(row, col_idx))
            if not logist_name:
                continue
            rows.append(
                {
                    "source_file_id": source_file_id,
                    "reys_number": reys_number,
                    "logist_name": logist_name,
                    "position": position,
                    "warehouse_no_extra_days": warehouse_kpi,
                    "no_damage_or_missing": damage_kpi,
                    "raw_data_json": _safe_json({"row": row}),
                }
            )
    return rows


def _parse_status_sheet(sheet_name: str, matrix: list[list[Any]], source_file_id: str) -> list[dict[str, Any]]:
    if _normalize_key(sheet_name) != "status" or not matrix:
        return []
    rows: list[dict[str, Any]] = []
    for idx, row in enumerate(matrix[1:], start=2):
        reys_number = _clean_text(_read_cell(row, 1))
        if not reys_number:
            continue
        rows.append(
            {
                "source_file_id": source_file_id,
                "agent": _clean_text(_read_cell(row, 0)),
                "reys_number": reys_number,
                "logist_summa": _parse_float(_read_cell(row, 2)),
                "rate": _parse_float(_read_cell(row, 3)),
                "usd": _parse_float(_read_cell(row, 4)),
                "loaded_date": _parse_date(_read_cell(row, 5)),
                "china_truck_number": _clean_text(_read_cell(row, 6)),
                "container_or_truck": _clean_text(_read_cell(row, 7)),
                "container_type": _clean_text(_read_cell(row, 8)),
                "agent_given_date": _parse_date(_read_cell(row, 9)),
                "agent_fact_days": _parse_float(_read_cell(row, 10)),
                "horgos_date": _parse_date(_read_cell(row, 11)),
                "zhongshan_horgos_days": _parse_float(_read_cell(row, 12)),
                "kazakh_truck_date": _parse_date(_read_cell(row, 13)),
                "driver_name": _clean_text(_read_cell(row, 14)),
                "driver_phone": _clean_text(_read_cell(row, 15)),
                "loaded_to_truck_days": _parse_float(_read_cell(row, 16)),
                "kazakh_truck_number": _clean_text(_read_cell(row, 17)),
                "tashkent_date": _parse_date(_read_cell(row, 18)),
                "zhongshan_tashkent_days": _parse_float(_read_cell(row, 19)),
                "customs_date": _parse_date(_read_cell(row, 20)),
                "distributed_date": _parse_date(_read_cell(row, 21)),
                "distribution_days": _parse_float(_read_cell(row, 22)),
                "raw_data_json": _safe_json({"row": row}),
            }
        )
    return rows


def _parse_fura_statusi_sheet(sheet_name: str, matrix: list[list[Any]], source_file_id: str) -> list[dict[str, Any]]:
    if _normalize_key(sheet_name) != "fura statusi" or not matrix:
        return []
    rows: list[dict[str, Any]] = []
    for idx, row in enumerate(matrix[1:], start=2):
        reys_number = _clean_text(_read_cell(row, 1))
        status = _clean_text(_read_cell(row, 2))
        if not (reys_number and status):
            continue
        rows.append(
            {
                "source_file_id": source_file_id,
                "reys_number": reys_number,
                "status": status,
                "status_date": _parse_date(_read_cell(row, 3)),
                "truck_number": _clean_text(_read_cell(row, 4)),
                "driver_name": _clean_text(_read_cell(row, 5)),
                "driver_phone": _clean_text(_read_cell(row, 6)),
                "raw_data_json": _safe_json({"row": row}),
            }
        )
    return rows


def _get_rate_for_date(rates: list[dict[str, Any]], currency: str, on_date: str) -> float | None:
    currency = _normalize_currency(currency)
    if not currency:
        return None
    if currency == "USD":
        return 1.0
    target_dt = _parse_date_obj(on_date)
    candidates: list[tuple[date, float]] = []
    for row in rates:
        if _normalize_currency(row.get("currency")) != currency:
            continue
        rate_dt = _parse_date_obj(row.get("rate_date"))
        if not rate_dt:
            continue
        rate_value = _parse_float(row.get("rate_to_usd"))
        if rate_value <= 0:
            continue
        if target_dt is None or rate_dt <= target_dt:
            candidates.append((rate_dt, rate_value))
    if not candidates:
        return None
    candidates.sort(key=lambda item: item[0], reverse=True)
    return candidates[0][1]


def _apply_cashflow_usd(cashflow_rows: list[dict[str, Any]], rate_rows: list[dict[str, Any]]) -> None:
    for row in cashflow_rows:
        amount = _parse_float(row.get("amount"))
        currency = _normalize_currency(row.get("currency"))
        if currency == "USD":
            row["amount_usd"] = amount
            continue
        rate = _get_rate_for_date(rate_rows, currency, row.get("operation_date") or row.get("created_date"))
        if not rate:
            row["amount_usd"] = 0.0
            continue
        row["amount_usd"] = amount / rate


def _create_sync_log(source_type: str, source_name: str) -> int:
    conn = db.get_conn()
    try:
        cur = conn.execute(
            """
            INSERT INTO analytics_sync_logs(started_at, finished_at, status, rows_imported, rows_skipped, error_message, details_json, source_type, source_name)
            VALUES (?, '', 'running', 0, 0, '', '', ?, ?)
            """,
            (_format_dt(), source_type, source_name),
        )
        conn.commit()
        return int(cur.lastrowid)
    finally:
        conn.close()


def _finish_sync_log(
    log_id: int,
    status: str,
    rows_imported: int,
    rows_skipped: int,
    details: dict[str, Any],
    error_message: str = "",
) -> None:
    conn = db.get_conn()
    try:
        conn.execute(
            """
            UPDATE analytics_sync_logs
            SET finished_at = ?, status = ?, rows_imported = ?, rows_skipped = ?, error_message = ?, details_json = ?
            WHERE id = ?
            """,
            (_format_dt(), status, rows_imported, rows_skipped, error_message, _safe_json(details), log_id),
        )
        conn.commit()
    finally:
        conn.close()


def _clear_analytics_tables(conn) -> None:
    for table_name in (
        "analytics_sales_records",
        "analytics_cashflow_records",
        "analytics_currency_rates",
        "analytics_logist_assignments",
        "analytics_shipment_summary",
        "analytics_shipment_statuses",
    ):
        conn.execute(f"DELETE FROM {table_name}")


def _store_import_result(
    sales_rows: list[dict[str, Any]],
    cashflow_rows: list[dict[str, Any]],
    rate_rows: list[dict[str, Any]],
    logist_rows: list[dict[str, Any]],
    shipment_rows: list[dict[str, Any]],
    status_rows: list[dict[str, Any]],
) -> None:
    conn = db.get_conn()
    try:
        _clear_analytics_tables(conn)

        conn.executemany(
            """
            INSERT INTO analytics_sales_records(
                source_file_id, source_sheet, source_row, reys_number, invoice_date, sale_date, shipping_mark, brand_name,
                client_name, phone, client_status, cargo_name, quantity, ctn, cbm, net_weight, gross_weight, customs_payment,
                company_expense, certificate_expense, client_price, sale_amount, correction_amount, discount_amount,
                final_sale_amount, salesperson, sales_kpi_amount, customs_kpi_amount, raw_data_json, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    row["source_file_id"],
                    row["source_sheet"],
                    row["source_row"],
                    row["reys_number"],
                    row["invoice_date"],
                    row["sale_date"],
                    row["shipping_mark"],
                    row["brand_name"],
                    row["client_name"],
                    row["phone"],
                    row["client_status"],
                    row["cargo_name"],
                    row["quantity"],
                    row["ctn"],
                    row["cbm"],
                    row["net_weight"],
                    row["gross_weight"],
                    row["customs_payment"],
                    row["company_expense"],
                    row["certificate_expense"],
                    row["client_price"],
                    row["sale_amount"],
                    row["correction_amount"],
                    row["discount_amount"],
                    row["final_sale_amount"],
                    row["salesperson"],
                    row["sales_kpi_amount"],
                    row["customs_kpi_amount"],
                    row["raw_data_json"],
                    _format_dt(),
                    _format_dt(),
                )
                for row in sales_rows
            ],
        )

        conn.executemany(
            """
            INSERT INTO analytics_cashflow_records(
                source_file_id, source_sheet, source_row, created_date, operation_date, wallet, flow_type, currency, comment,
                category, amount, rate, department, reys_number, counterparty, auto_confirm, amount_usd, raw_data_json, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    row["source_file_id"],
                    row["source_sheet"],
                    row["source_row"],
                    row["created_date"],
                    row["operation_date"],
                    row["wallet"],
                    row["flow_type"],
                    row["currency"],
                    row["comment"],
                    row["category"],
                    row["amount"],
                    row["rate"],
                    row["department"],
                    row["reys_number"],
                    row["counterparty"],
                    row["auto_confirm"],
                    row["amount_usd"],
                    row["raw_data_json"],
                    _format_dt(),
                    _format_dt(),
                )
                for row in cashflow_rows
            ],
        )

        conn.executemany(
            """
            INSERT INTO analytics_currency_rates(
                source_file_id, rate_date, currency, rate_to_usd, raw_data_json, created_at
            )
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    row["source_file_id"],
                    row["rate_date"],
                    row["currency"],
                    row["rate_to_usd"],
                    row["raw_data_json"],
                    _format_dt(),
                )
                for row in rate_rows
            ],
        )

        conn.executemany(
            """
            INSERT INTO analytics_logist_assignments(
                source_file_id, reys_number, logist_name, position, warehouse_no_extra_days, no_damage_or_missing, raw_data_json, created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    row["source_file_id"],
                    row["reys_number"],
                    row["logist_name"],
                    row["position"],
                    row["warehouse_no_extra_days"],
                    row["no_damage_or_missing"],
                    row["raw_data_json"],
                    _format_dt(),
                )
                for row in logist_rows
            ],
        )

        conn.executemany(
            """
            INSERT INTO analytics_shipment_summary(
                source_file_id, agent, reys_number, logist_summa, rate, usd, loaded_date, china_truck_number, container_or_truck,
                container_type, agent_given_date, agent_fact_days, horgos_date, zhongshan_horgos_days, kazakh_truck_date, driver_name,
                driver_phone, loaded_to_truck_days, kazakh_truck_number, tashkent_date, zhongshan_tashkent_days, customs_date,
                distributed_date, distribution_days, raw_data_json, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    row["source_file_id"],
                    row["agent"],
                    row["reys_number"],
                    row["logist_summa"],
                    row["rate"],
                    row["usd"],
                    row["loaded_date"],
                    row["china_truck_number"],
                    row["container_or_truck"],
                    row["container_type"],
                    row["agent_given_date"],
                    row["agent_fact_days"],
                    row["horgos_date"],
                    row["zhongshan_horgos_days"],
                    row["kazakh_truck_date"],
                    row["driver_name"],
                    row["driver_phone"],
                    row["loaded_to_truck_days"],
                    row["kazakh_truck_number"],
                    row["tashkent_date"],
                    row["zhongshan_tashkent_days"],
                    row["customs_date"],
                    row["distributed_date"],
                    row["distribution_days"],
                    row["raw_data_json"],
                    _format_dt(),
                    _format_dt(),
                )
                for row in shipment_rows
            ],
        )

        conn.executemany(
            """
            INSERT INTO analytics_shipment_statuses(
                source_file_id, reys_number, status, status_date, truck_number, driver_name, driver_phone, raw_data_json, created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    row["source_file_id"],
                    row["reys_number"],
                    row["status"],
                    row["status_date"],
                    row["truck_number"],
                    row["driver_name"],
                    row["driver_phone"],
                    row["raw_data_json"],
                    _format_dt(),
                )
                for row in status_rows
            ],
        )

        conn.commit()
    finally:
        conn.close()


def _sync_plan_defaults() -> None:
    conn = db.get_conn()
    try:
        active_exists = conn.execute(
            "SELECT 1 FROM analytics_sales_plans WHERE is_active = 1 LIMIT 1"
        ).fetchone()
        if active_exists:
            return
        today = _now_tashkent().date()
        start = today.replace(day=1)
        if today.month == 12:
            next_month = date(today.year + 1, 1, 1)
        else:
            next_month = date(today.year, today.month + 1, 1)
        end = next_month - timedelta(days=1)
        conn.execute(
            """
            INSERT INTO analytics_sales_plans(name, period_start, period_end, target_amount_usd, target_metric, target_value, is_active, created_at, updated_at)
            VALUES (?, ?, ?, 0, 'amount_usd', 0, 1, ?, ?)
            """,
            (
                f"{start.strftime('%m.%Y')} Sales plan",
                start.strftime("%d.%m.%Y"),
                end.strftime("%d.%m.%Y"),
                _format_dt(),
                _format_dt(),
            ),
        )
        conn.commit()
    finally:
        conn.close()


def sync_workbook(
    sheets: dict[str, list[list[Any]]],
    source_type: str,
    source_name: str,
    source_file_id: str,
) -> dict[str, Any]:
    log_id = _create_sync_log(source_type, source_name)
    sales_rows: list[dict[str, Any]] = []
    cashflow_rows: list[dict[str, Any]] = []
    rate_rows: list[dict[str, Any]] = []
    logist_rows: list[dict[str, Any]] = []
    shipment_rows: list[dict[str, Any]] = []
    status_rows: list[dict[str, Any]] = []
    skipped_rows = 0
    try:
        for sheet_name, matrix in sheets.items():
            if not matrix:
                continue
            sales_rows.extend(_parse_total_sheet(sheet_name, matrix, source_file_id))
            cashflow_rows.extend(_parse_cashflow_sheet(sheet_name, matrix, source_file_id))
            rate_rows.extend(_parse_kurs_sheet(sheet_name, matrix, source_file_id))
            logist_rows.extend(_parse_kpi_logist_sheet(sheet_name, matrix, source_file_id))
            shipment_rows.extend(_parse_status_sheet(sheet_name, matrix, source_file_id))
            status_rows.extend(_parse_fura_statusi_sheet(sheet_name, matrix, source_file_id))

        _apply_cashflow_usd(cashflow_rows, rate_rows)
        imported_rows = (
            len(sales_rows)
            + len(cashflow_rows)
            + len(rate_rows)
            + len(logist_rows)
            + len(shipment_rows)
            + len(status_rows)
        )
        _store_import_result(sales_rows, cashflow_rows, rate_rows, logist_rows, shipment_rows, status_rows)
        _sync_plan_defaults()
        db.set_setting(ANALYTICS_LAST_SYNC_KEY, _format_dt())
        db.set_setting(ANALYTICS_SOURCE_NAME_KEY, source_name)
        details = {
            "source_type": source_type,
            "source_name": source_name,
            "source_file_id": source_file_id,
            "sheets": list(sheets.keys()),
            "sales_rows": len(sales_rows),
            "cashflow_rows": len(cashflow_rows),
            "rate_rows": len(rate_rows),
            "logist_rows": len(logist_rows),
            "shipment_rows": len(shipment_rows),
            "status_rows": len(status_rows),
        }
        _finish_sync_log(log_id, "success", imported_rows, skipped_rows, details)
        return {
            "ok": True,
            "rows_imported": imported_rows,
            "rows_skipped": skipped_rows,
            "details": details,
            "last_sync_at": db.get_setting(ANALYTICS_LAST_SYNC_KEY, ""),
        }
    except Exception as exc:
        _finish_sync_log(
            log_id,
            "failed",
            0,
            skipped_rows,
            {
                "source_type": source_type,
                "source_name": source_name,
                "source_file_id": source_file_id,
                "sheets": list(sheets.keys()),
            },
            str(exc),
        )
        raise


def sync_from_google(sheet_id: str | None = None) -> dict[str, Any]:
    target_sheet_id = _extract_sheet_id(sheet_id or get_google_sheet_id())
    status = google_credentials_status()
    if status.get("mode") == "api":
        workbook = _load_google_workbook(target_sheet_id)
        result = sync_workbook(workbook, "google_api", f"Google Sheets {target_sheet_id}", target_sheet_id)
        set_google_sheet_id(target_sheet_id)
        return result
    workbook = _load_public_google_workbook(target_sheet_id)
    result = sync_workbook(workbook, "google_public", f"Google Sheets {target_sheet_id}", target_sheet_id)
    set_google_sheet_id(target_sheet_id)
    return result


def sync_from_upload(file_storage) -> dict[str, Any]:
    workbook = _load_uploaded_workbook(file_storage)
    filename = _clean_text(getattr(file_storage, "filename", "")) or "upload.xlsx"
    source_file_id = f"upload:{_now_tashkent().strftime('%Y%m%d%H%M%S')}"
    return sync_workbook(workbook, "upload", filename, source_file_id)


def get_sync_status() -> dict[str, Any]:
    conn = db.get_conn()
    try:
        logs = conn.execute(
            """
            SELECT *
            FROM analytics_sync_logs
            ORDER BY started_at DESC, id DESC
            LIMIT 20
            """
        ).fetchall()
        plans = conn.execute(
            """
            SELECT id, name, period_start, period_end, target_amount_usd, target_metric, target_value, is_active
            FROM analytics_sales_plans
            ORDER BY is_active DESC, period_start DESC, id DESC
            """
        ).fetchall()
    finally:
        conn.close()
    return {
        "sheet_id": get_google_sheet_id(),
        "last_sync_at": db.get_setting(ANALYTICS_LAST_SYNC_KEY, ""),
        "source_name": db.get_setting(ANALYTICS_SOURCE_NAME_KEY, ""),
        "connection": google_credentials_status(),
        "logs": [dict(row) for row in logs],
        "plans": [dict(row) for row in plans],
    }


GOOGLE_SHEETS_MISSING_MESSAGE = (
    "Google Sheets ulanishi sozlanmagan. .env faylga GOOGLE_SHEET_ID va "
    "GOOGLE_APPLICATION_CREDENTIALS yoki GOOGLE_SERVICE_ACCOUNT_JSON qo‘shing."
)


def _get_credentials_payload() -> tuple[dict[str, Any] | None, str | None]:
    inline_json = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON", "").strip()
    if inline_json:
        try:
            return json.loads(inline_json), None
        except json.JSONDecodeError as exc:
            raise AnalyticsImporterError(f"GOOGLE_SERVICE_ACCOUNT_JSON noto‘g‘ri: {exc}")
    path = os.getenv("GOOGLE_APPLICATION_CREDENTIALS", "").strip()
    if path:
        payload_path = Path(path)
        if not payload_path.exists():
            raise AnalyticsImporterError("GOOGLE_APPLICATION_CREDENTIALS fayli topilmadi.")
        return None, str(payload_path)
    return None, None


def google_credentials_status() -> dict[str, Any]:
    sheet_id = get_google_sheet_id()
    try:
        payload, path = _get_credentials_payload()
    except AnalyticsImporterError as exc:
        return {"connected": False, "sheet_id": sheet_id, "mode": "invalid", "message": str(exc)}
    if not sheet_id:
        return {
            "connected": False,
            "sheet_id": "",
            "mode": "missing",
            "message": "Google Sheet ID yoki to‘liq havola kiriting.",
        }
    if payload is None and path is None:
        return {
            "connected": True,
            "sheet_id": sheet_id,
            "mode": "public",
            "message": "Google API credentials topilmadi. Public link rejimi ishlaydi, agar jadval havola bo‘yicha ochiq bo‘lsa.",
        }
    return {
        "connected": True,
        "sheet_id": sheet_id,
        "mode": "api",
        "message": "Google Sheets API ulanishi tayyor.",
    }


def _load_public_google_workbook(sheet_id: str | None = None) -> dict[str, list[list[Any]]]:
    sheet_id = _extract_sheet_id(sheet_id or get_google_sheet_id())
    if not sheet_id:
        raise AnalyticsImporterError("Google Sheet ID yoki to‘liq havola kiriting.")
    export_url = GOOGLE_SHEETS_EXPORT_URL.format(sheet_id=sheet_id)
    request = Request(
        export_url,
        headers={
            "User-Agent": "Mozilla/5.0",
            "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
        },
    )
    try:
        with urlopen(request, timeout=30) as response:
            content_type = str(response.headers.get("Content-Type", "")).lower()
            raw = response.read()
    except HTTPError as exc:
        if exc.code in {401, 403}:
            raise AnalyticsImporterError(
                "Google Sheets public export yopiq. Jadvalni havola bo‘yicha ko‘rish uchun oching yoki API credentials qo‘shing."
            ) from exc
        raise AnalyticsImporterError(f"Google Sheets public export xatosi: HTTP {exc.code}") from exc
    except URLError as exc:
        raise AnalyticsImporterError(f"Google Sheets bilan ulanishda xato: {exc.reason}") from exc
    if "text/html" in content_type and b"<html" in raw[:200].lower():
        raise AnalyticsImporterError(
            "Google Sheets public export XLSX qaytarmadi. Jadval havola bo‘yicha ochiq ekanini tekshiring."
        )
    return _load_workbook_from_bytes(raw, "Google Sheets public export")
