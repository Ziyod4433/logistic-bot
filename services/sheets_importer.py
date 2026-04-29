from __future__ import annotations

import csv
import io
import json
import os
import re
import sqlite3
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

import database as db

GOOGLE_SHEETS_MISSING_MESSAGE = (
    "Google Sheets ulanishi sozlanmagan. .env faylga GOOGLE_SHEET_ID va "
    "GOOGLE_APPLICATION_CREDENTIALS yoki GOOGLE_SERVICE_ACCOUNT_JSON qo‘shing."
)

ANALYTICS_SHEET_ID_KEY = "analytics_google_sheet_id"
ANALYTICS_LAST_SYNC_KEY = "analytics_last_sync_at"
ANALYTICS_EXCHANGE_RATES_KEY = "analytics_exchange_rates"
GOOGLE_SHEETS_EXPORT_URL = "https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"


class SheetsImporterError(Exception):
    pass


def _clean_text(value: Any) -> str:
    text = str(value or "").replace("\u00a0", " ").strip()
    return re.sub(r"\s+", " ", text)


def _extract_sheet_id(sheet_id_or_url: str) -> str:
    raw = _clean_text(sheet_id_or_url)
    if not raw:
        return ""
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", raw)
    if match:
        return match.group(1)
    if re.fullmatch(r"[a-zA-Z0-9-_]{20,}", raw):
        return raw
    return raw


def _normalize_key(value: Any) -> str:
    text = _clean_text(value).lower()
    text = text.replace("ё", "е").replace("ў", "у").replace("қ", "к").replace("ғ", "г")
    text = text.replace("/", " ").replace("\\", " ").replace("-", " ").replace("_", " ")
    text = re.sub(r"[^\w\s%№]+", "", text, flags=re.UNICODE)
    return re.sub(r"\s+", " ", text).strip()


def _clean_number_text(value: Any) -> str:
    return str(value or "").replace("\u00a0", "").replace(" ", "").replace(",", ".")


def _parse_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = _clean_number_text(value)
    if not text:
        return 0.0
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return 0.0
    try:
        return float(match.group(0))
    except ValueError:
        return 0.0


def _parse_int(value: Any) -> int:
    return int(round(_parse_float(value)))


def _parse_date(value: Any) -> str:
    text = _clean_text(value)
    if not text:
        return ""
    for pattern, fmt in (
        (r"^\d{2}\.\d{2}\.\d{4}$", "%d.%m.%Y"),
        (r"^\d{4}-\d{2}-\d{2}$", "%Y-%m-%d"),
        (r"^\d{2}/\d{2}/\d{4}$", "%d/%m/%Y"),
        (r"^\d{2}-\d{2}-\d{4}$", "%d-%m-%Y"),
    ):
        if re.match(pattern, text):
            try:
                parsed = datetime.strptime(text, fmt)
                return parsed.strftime("%d.%m.%Y")
            except ValueError:
                return ""
    match = re.search(r"(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{2,4})", text)
    if match:
        day, month, year = match.groups()
        if len(year) == 2:
            year = f"20{year}"
        try:
            parsed = datetime(int(year), int(month), int(day))
            return parsed.strftime("%d.%m.%Y")
        except ValueError:
            return ""
    return ""


def _parse_datetime(value: Any) -> str:
    date_only = _parse_date(value)
    if date_only:
        return f"{date_only} 00:00:00"
    return ""


def _normalize_currency(value: Any) -> str:
    key = _normalize_key(value)
    if not key:
        return ""
    if any(token in key for token in ("usd", "dollar", "доллар")):
        return "USD"
    if any(token in key for token in ("rmb", "yuan", "cny", "юань")):
        return "RMB"
    if any(token in key for token in ("uzs", "sum", "сум")):
        return "UZS"
    return _clean_text(value).upper()


def _format_dt(dt: datetime | None = None) -> str:
    return (dt or datetime.now()).strftime("%Y-%m-%d %H:%M:%S")


def _safe_json(data: Any) -> str:
    return json.dumps(data, ensure_ascii=False, default=str)


def _bool_env(value: str | None) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "yes", "on"}


SHEET_ALIASES = {
    "date": {
        "date",
        "sana",
        "дата",
        "payment date",
        "yuklangan sana",
        "loaded date",
        "yetib kelgan sana",
        "arrived date",
        "date of arrive",
        "arrival date",
    },
    "bl_code": {
        "bl",
        "bl kodi",
        "bl kodi",
        "bl code",
        "shipping mark",
        "shippingmark",
    },
    "client_name": {
        "mijoz",
        "client",
        "client name",
        "company",
        "company name",
        "klient",
        "клиент",
    },
    "manager_name": {
        "manager",
        "manager name",
        "sales manager",
        "sales cordinator",
        "sales coordinator",
        "sales menejerlar",
        "collector",
        "координатор",
    },
    "logist_name": {"logist", "logist name", "logistic", "логист"},
    "service_type": {"service", "service type", "xizmat", "тип услуги"},
    "amount": {"amount", "summa", "сумма", "savdo", "sale amount", "jami summa", "total amount"},
    "cost": {"cost", "expense", "harajat", "chiqim", "себестоимость", "расход"},
    "profit": {"profit", "foyda", "прибыль"},
    "paid_amount": {"paid", "paid amount", "tolangan", "tolangan", "to'langan", "оплачено"},
    "debt_amount": {"debt", "qarz", "qoldiq qarz", "debitorka", "долг", "остаток"},
    "payment_status": {"payment status", "status оплаты", "payment", "оплата"},
    "currency": {"currency", "valyuta", "валюта"},
    "category": {"category", "kategoriya", "категория"},
    "bank_or_cash": {"bank", "kassa", "cash", "bank/kassa", "bank or cash"},
    "contractor": {"contractor", "counterparty", "supplier", "podryadchik", "контрагент"},
    "comment": {"comment", "izoh", "комментарий", "примечание"},
    "reys_number": {"reys", "рейс", "reys nomeri", "reys number"},
    "fura_number": {"fura", "fura number", "truck", "car", "машина"},
    "container_type": {"container", "container type", "тип контейнера"},
    "station": {"station", "stansiya", "станция"},
    "agent": {"agent", "агент"},
    "sales_manager_name": {"sales manager", "sales manager name", "sales menejerlar", "manager"},
    "warehouse": {"warehouse", "sklad", "ombor", "склад"},
    "status": {"status", "holat", "статус"},
    "loaded_date": {"loaded date", "yuklangan sana", "дата загрузки"},
    "arrived_date": {"arrived date", "yetib kelgan sana", "date of arrive", "дата прибытия"},
    "expected_date": {"expected date", "taxminiy sana", "ожидаемая дата"},
    "cargo_type": {"cargo type", "tovar turi", "вид товара", "товар"},
    "weight_kg": {"kg", "weight", "weight kg", "ogirligi", "og'irligi"},
    "volume_m3": {"cbm", "t cbm", "t/cbm", "volume", "hajmi", "m3", "м3"},
    "places": {"ctn 件数", "ctn", "件数", "places", "quantity places", "joy soni", "количество мест"},
}


def _detect_header_row(matrix: list[list[Any]], wanted_fields: set[str], min_matches: int = 2) -> tuple[int, dict[int, str]] | tuple[None, None]:
    best_row_idx = None
    best_mapping: dict[int, str] = {}
    for row_idx, row in enumerate(matrix[:50]):
        mapping: dict[int, str] = {}
        matches = 0
        for col_idx, raw_value in enumerate(row):
            normalized = _normalize_key(raw_value)
            if not normalized:
                continue
            for field_name in wanted_fields:
                aliases = SHEET_ALIASES.get(field_name, set())
                if normalized in aliases:
                    if field_name not in mapping.values():
                        mapping[col_idx] = field_name
                        matches += 1
                    break
        if matches >= min_matches and matches > len(best_mapping):
            best_row_idx = row_idx
            best_mapping = mapping
    if best_row_idx is None:
        return None, None
    return best_row_idx, best_mapping


def _iter_table_rows(matrix: list[list[Any]], header_idx: int, mapping: dict[int, str]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    blank_streak = 0
    for row in matrix[header_idx + 1 :]:
        values = {field: _clean_text(row[col_idx] if col_idx < len(row) else "") for col_idx, field in mapping.items()}
        populated = sum(1 for value in values.values() if value)
        if populated == 0:
            blank_streak += 1
            if blank_streak >= 3:
                break
            continue
        blank_streak = 0
        marker = _normalize_key(next(iter(values.values()), ""))
        if marker in {"total", "итого"}:
            continue
        rows.append(values)
    return rows


def _extract_rates_from_sheet(sheet_name: str, matrix: list[list[Any]]) -> dict[str, float]:
    if "kurs" not in _normalize_key(sheet_name):
        return {}
    rates: dict[str, float] = {}
    for row in matrix:
        if len(row) < 2:
            continue
        left = _normalize_currency(row[0])
        right = _parse_float(row[1])
        if left and right > 0:
            rates[left] = right
    if "UZS" not in rates:
        rates["UZS"] = 1.0
    return rates


def _extract_sheet_date(value: Any) -> str:
    return _parse_date(value)


def _parse_shipping_blocks(sheet_name: str, matrix: list[list[Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    aggregated: dict[tuple[str, str], dict[str, Any]] = {}

    def cell(r_idx: int, c_idx: int) -> str:
        if r_idx < 0 or r_idx >= len(matrix):
            return ""
        row = matrix[r_idx]
        if c_idx < 0 or c_idx >= len(row):
            return ""
        return _clean_text(row[c_idx])

    for header_row_idx, row in enumerate(matrix):
        for start_col_idx, value in enumerate(row):
            if _normalize_key(value) != "shipping mark":
                continue

            sheet_date = ""
            reys_name = ""
            for up_idx in range(header_row_idx - 1, -1, -1):
                search_row = matrix[up_idx]
                left = max(0, start_col_idx - 2)
                right = min(len(search_row), start_col_idx + 6)
                for candidate in search_row[left:right]:
                    if not sheet_date:
                        sheet_date = _extract_sheet_date(candidate)
                    elif not reys_name and _clean_text(candidate) and not _extract_sheet_date(candidate):
                        reys_name = _clean_text(candidate)
                if sheet_date and reys_name:
                    break

            blank_streak = 0
            data_row_idx = header_row_idx + 1
            while data_row_idx < len(matrix):
                code = cell(data_row_idx, start_col_idx)
                relevant = [cell(data_row_idx, start_col_idx + offset) for offset in range(5)]
                normalized_code = _normalize_key(code)
                if normalized_code == "shipping mark":
                    break
                if normalized_code in {"total", "итого"}:
                    break
                if not any(relevant):
                    blank_streak += 1
                    if blank_streak >= 3:
                        break
                    data_row_idx += 1
                    continue
                blank_streak = 0
                if code:
                    code_text = _clean_text(code).upper()
                    quantity_places = _parse_int(cell(data_row_idx, start_col_idx + 1))
                    volume_cbm = _parse_float(cell(data_row_idx, start_col_idx + 2))
                    weight_kg = _parse_float(cell(data_row_idx, start_col_idx + 3))
                    arrived_date = _parse_date(cell(data_row_idx, start_col_idx + 5))
                    quantity_piece = str(quantity_places) if quantity_places else ""
                    aggregate_key = (sheet_date or "", code_text)
                    existing = aggregated.get(aggregate_key)
                    if existing:
                        existing["weight_kg"] += weight_kg
                        existing["volume_m3"] += volume_cbm
                        if quantity_piece:
                            existing["places_items"].append(quantity_piece)
                        existing["source_rows"].append(data_row_idx + 1)
                    else:
                        aggregated[aggregate_key] = {
                            "bl_code": code_text,
                            "client_name": "",
                            "reys_number": reys_name,
                            "fura_number": "",
                            "container_type": "",
                            "station": "",
                            "agent": "",
                            "logist_name": "",
                            "sales_manager_name": "",
                            "warehouse": "",
                            "status": "",
                            "loaded_date": sheet_date,
                            "arrived_date": arrived_date,
                            "expected_date": "",
                            "cargo_type": "",
                            "weight_kg": weight_kg,
                            "volume_m3": volume_cbm,
                            "places": quantity_places,
                            "places_items": [quantity_piece] if quantity_piece else [],
                            "description": "",
                            "source_sheet": sheet_name,
                            "raw_data_json": "",
                            "source_rows": [data_row_idx + 1],
                        }
                data_row_idx += 1

    for row in aggregated.values():
        row["places_breakdown"] = " + ".join(item for item in row.pop("places_items", []) if item)
        row["places"] = db._sum_quantity_breakdown(row["places_breakdown"], row.get("places", 0))
        row["raw_data_json"] = _safe_json(
            {"source_rows": row.pop("source_rows", []), "source_sheet": row["source_sheet"]}
        )
        rows.append(row)
    return rows


def _parse_cashflow_sheet(sheet_name: str, matrix: list[list[Any]]) -> list[dict[str, Any]]:
    if "cashflow" not in _normalize_key(sheet_name):
        return []
    wanted = {"date", "type", "category", "amount", "currency", "bank_or_cash", "contractor", "bl_code", "reys_number", "comment"}
    header_idx, mapping = _detect_header_row(matrix, wanted, min_matches=3)
    if header_idx is None or mapping is None:
        return []
    records = []
    for raw in _iter_table_rows(matrix, header_idx, mapping):
        amount = _parse_float(raw.get("amount"))
        if not amount:
            continue
        txn_type = _normalize_key(raw.get("type"))
        if txn_type not in {"income", "expense"}:
            if amount < 0:
                txn_type = "expense"
            else:
                txn_type = "income"
        records.append(
            {
                "date": _parse_date(raw.get("date")),
                "type": txn_type,
                "category": _clean_text(raw.get("category")),
                "amount": abs(amount),
                "currency": _normalize_currency(raw.get("currency")) or "UZS",
                "bank_or_cash": _clean_text(raw.get("bank_or_cash")),
                "contractor": _clean_text(raw.get("contractor")),
                "bl_code": _clean_text(raw.get("bl_code")).upper(),
                "reys_number": _clean_text(raw.get("reys_number")),
                "comment": _clean_text(raw.get("comment")),
                "source_sheet": sheet_name,
                "raw_data_json": _safe_json(raw),
            }
        )
    return records


def _parse_sales_sheet(sheet_name: str, matrix: list[list[Any]]) -> list[dict[str, Any]]:
    title_key = _normalize_key(sheet_name)
    if "cashflow" in title_key or "kurs" in title_key:
        return []
    wanted = {
        "date",
        "bl_code",
        "client_name",
        "manager_name",
        "service_type",
        "amount",
        "cost",
        "profit",
        "currency",
        "paid_amount",
        "debt_amount",
        "payment_status",
    }
    header_idx, mapping = _detect_header_row(matrix, wanted, min_matches=3)
    if header_idx is None or mapping is None:
        return []
    rows = []
    for raw in _iter_table_rows(matrix, header_idx, mapping):
        amount = _parse_float(raw.get("amount"))
        cost = _parse_float(raw.get("cost"))
        profit = _parse_float(raw.get("profit"))
        paid_amount = _parse_float(raw.get("paid_amount"))
        debt_amount = _parse_float(raw.get("debt_amount"))
        if not any((amount, cost, profit, paid_amount, debt_amount, _clean_text(raw.get("client_name")), _clean_text(raw.get("bl_code")))):
            continue
        if not profit and (amount or cost):
            profit = amount - cost
        if not debt_amount and amount:
            debt_amount = max(amount - paid_amount, 0)
        rows.append(
            {
                "date": _parse_date(raw.get("date")),
                "bl_code": _clean_text(raw.get("bl_code")).upper(),
                "client_name": _clean_text(raw.get("client_name")),
                "manager_name": _clean_text(raw.get("manager_name")),
                "service_type": _clean_text(raw.get("service_type")),
                "amount": amount,
                "cost": cost,
                "profit": profit,
                "currency": _normalize_currency(raw.get("currency")) or "UZS",
                "paid_amount": paid_amount,
                "debt_amount": debt_amount,
                "payment_status": _clean_text(raw.get("payment_status")),
                "source": "sheet",
                "source_sheet": sheet_name,
                "raw_data_json": _safe_json(raw),
            }
        )
    return rows


def _merge_shipments(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: dict[tuple[str, str], dict[str, Any]] = {}
    for row in rows:
        key = ((_clean_text(row.get("loaded_date")) or ""), (_clean_text(row.get("bl_code")) or "").upper())
        if not key[1]:
            continue
        existing = merged.get(key)
        if existing:
            existing["weight_kg"] += float(row.get("weight_kg") or 0)
            existing["volume_m3"] += float(row.get("volume_m3") or 0)
            breakdown_parts = [part for part in [existing.get("places_breakdown", ""), row.get("places_breakdown", "")] if part]
            existing["places_breakdown"] = " + ".join(breakdown_parts)
            existing["places"] = db._sum_quantity_breakdown(existing["places_breakdown"], existing.get("places", 0))
            if not existing.get("arrived_date"):
                existing["arrived_date"] = row.get("arrived_date", "")
            if not existing.get("status"):
                existing["status"] = row.get("status", "")
        else:
            row = dict(row)
            row["weight_kg"] = float(row.get("weight_kg") or 0)
            row["volume_m3"] = float(row.get("volume_m3") or 0)
            row["places_breakdown"] = _clean_text(row.get("places_breakdown"))
            row["places"] = db._sum_quantity_breakdown(row["places_breakdown"], row.get("places", 0))
            merged[key] = row
    return list(merged.values())


def _get_credentials_payload() -> tuple[dict[str, Any] | None, str | None]:
    inline_json = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON", "").strip()
    if inline_json:
        try:
            return json.loads(inline_json), None
        except json.JSONDecodeError as exc:
            raise SheetsImporterError(f"GOOGLE_SERVICE_ACCOUNT_JSON noto‘g‘ri: {exc}")
    path = os.getenv("GOOGLE_APPLICATION_CREDENTIALS", "").strip()
    if path:
        payload_path = Path(path)
        if not payload_path.exists():
            raise SheetsImporterError("GOOGLE_APPLICATION_CREDENTIALS fayli topilmadi.")
        return None, str(payload_path)
    return None, None


def google_credentials_status() -> dict[str, Any]:
    sheet_id = get_google_sheet_id()
    try:
        payload, path = _get_credentials_payload()
    except SheetsImporterError as exc:
        return {"connected": False, "sheet_id": sheet_id, "mode": "invalid", "message": str(exc)}
    if not sheet_id:
        return {
            "connected": False,
            "sheet_id": sheet_id,
            "mode": "missing",
            "message": "Google Sheets uchun Sheet ID yoki to‘liq havola kiriting.",
        }
    if payload is None and path is None:
        return {
            "connected": True,
            "sheet_id": sheet_id,
            "mode": "public",
            "message": (
                "Google API credentials topilmadi. Public link rejimi ishlaydi, "
                "agar jadval havola bo‘yicha ochiq bo‘lsa."
            ),
        }
    return {
        "connected": True,
        "sheet_id": sheet_id,
        "mode": "api",
        "message": "Google Sheets API ulanishi tayyor.",
    }


def get_google_sheet_id() -> str:
    return _extract_sheet_id(
        os.getenv("GOOGLE_SHEET_ID", "").strip() or db.get_setting(ANALYTICS_SHEET_ID_KEY, "").strip()
    )


def set_google_sheet_id(sheet_id: str) -> None:
    db.set_setting(ANALYTICS_SHEET_ID_KEY, _extract_sheet_id(sheet_id))


def _load_google_workbook(sheet_id: str | None = None) -> dict[str, list[list[Any]]]:
    sheet_id = _extract_sheet_id(sheet_id or get_google_sheet_id())
    status = google_credentials_status()
    if not sheet_id or not status["connected"]:
        raise SheetsImporterError(status["message"])
    try:
        from google.oauth2.service_account import Credentials
        from googleapiclient.discovery import build
    except Exception as exc:  # pragma: no cover
        raise SheetsImporterError(f"Google API kutubxonalari topilmadi: {exc}")

    payload, path = _get_credentials_payload()
    scopes = ["https://www.googleapis.com/auth/spreadsheets.readonly"]
    if payload is not None:
        credentials = Credentials.from_service_account_info(payload, scopes=scopes)
    elif path is not None:
        credentials = Credentials.from_service_account_file(path, scopes=scopes)
    else:  # pragma: no cover
        raise SheetsImporterError(GOOGLE_SHEETS_MISSING_MESSAGE)

    service = build("sheets", "v4", credentials=credentials, cache_discovery=False)
    metadata = service.spreadsheets().get(spreadsheetId=sheet_id).execute()
    workbook: dict[str, list[list[Any]]] = {}
    for sheet in metadata.get("sheets", []):
        title = sheet.get("properties", {}).get("title", "")
        if not title:
            continue
        values = service.spreadsheets().values().get(
            spreadsheetId=sheet_id,
            range=title,
            majorDimension="ROWS",
        ).execute().get("values", [])
        workbook[title] = values
    return workbook


def _load_workbook_from_bytes(raw: bytes, source_name: str) -> dict[str, list[list[Any]]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover
        raise SheetsImporterError(f"{source_name} import uchun openpyxl kerak: {exc}")

    workbook = load_workbook(io.BytesIO(raw), data_only=True)
    result: dict[str, list[list[Any]]] = {}
    for sheet in workbook.worksheets:
        values: list[list[Any]] = []
        for row in sheet.iter_rows(values_only=True):
            values.append([cell if cell is not None else "" for cell in row])
        result[sheet.title] = values
    return result


def _load_public_google_workbook(sheet_id: str | None = None) -> dict[str, list[list[Any]]]:
    sheet_id = _extract_sheet_id(sheet_id or get_google_sheet_id())
    if not sheet_id:
        raise SheetsImporterError("Google Sheets uchun Sheet ID yoki to‘liq havola kiriting.")
    export_url = GOOGLE_SHEETS_EXPORT_URL.format(sheet_id=sheet_id)
    request = Request(
        export_url,
        headers={
            "User-Agent": "Mozilla/5.0",
            "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
        },
    )
    try:
        with urlopen(request, timeout=30) as response:
            content_type = str(response.headers.get("Content-Type", "")).lower()
            raw = response.read()
    except HTTPError as exc:
        if exc.code in {401, 403}:
            raise SheetsImporterError(
                "Google Sheets public export yopiq. Jadvalni havola bo‘yicha ko‘rish uchun oching yoki API credentials qo‘shing."
            ) from exc
        raise SheetsImporterError(f"Google Sheets public export xatosi: HTTP {exc.code}") from exc
    except URLError as exc:
        raise SheetsImporterError(f"Google Sheets bilan ulanishda xato: {exc.reason}") from exc
    if "text/html" in content_type and b"<html" in raw[:200].lower():
        raise SheetsImporterError(
            "Google Sheets public export XLSX qaytarmadi. Jadval havola bo‘yicha ochiq ekanini tekshiring."
        )
    return _load_workbook_from_bytes(raw, "Google Sheets public export")


def _load_uploaded_workbook(file_storage) -> dict[str, list[list[Any]]]:
    if not file_storage:
        raise SheetsImporterError("CSV/XLSX fayl tanlanmagan.")
    filename = _clean_text(getattr(file_storage, "filename", ""))
    if not filename:
        raise SheetsImporterError("CSV/XLSX fayl tanlanmagan.")
    ext = filename.rsplit(".", 1)[-1].lower()
    raw = file_storage.read()
    if ext == "csv":
        text = raw.decode("utf-8-sig", errors="replace")
        return {"Imported CSV": list(csv.reader(io.StringIO(text)))}
    if ext not in {"xlsx", "xlsm"}:
        raise SheetsImporterError("Faqat CSV yoki XLSX/XLSM fayl yuklash mumkin.")
    return _load_workbook_from_bytes(raw, "XLSX")


def _create_sync_log(source_name: str) -> int:
    conn = db.get_conn()
    try:
        cur = conn.execute(
            """
            INSERT INTO sheet_sync_logs(started_at, finished_at, status, rows_imported, rows_skipped, error_message, details_json)
            VALUES (?, '', 'running', 0, 0, '', ?)
            """,
            (_format_dt(), _safe_json({"source": source_name})),
        )
        conn.commit()
        return int(cur.lastrowid)
    finally:
        conn.close()


def _finish_sync_log(log_id: int, status: str, rows_imported: int, rows_skipped: int, details: dict[str, Any], error_message: str = "") -> None:
    conn = db.get_conn()
    try:
        conn.execute(
            """
            UPDATE sheet_sync_logs
            SET finished_at = ?, status = ?, rows_imported = ?, rows_skipped = ?, error_message = ?, details_json = ?
            WHERE id = ?
            """,
            (_format_dt(), status, rows_imported, rows_skipped, error_message, _safe_json(details), log_id),
        )
        conn.commit()
    finally:
        conn.close()


def _store_import_result(
    sales_rows: list[dict[str, Any]],
    cashflow_rows: list[dict[str, Any]],
    shipment_rows: list[dict[str, Any]],
    rates: dict[str, float],
) -> None:
    conn = db.get_conn()
    try:
        conn.execute("DELETE FROM sales_transactions")
        conn.execute("DELETE FROM cashflow_transactions")
        conn.execute("DELETE FROM shipments")
        conn.execute("DELETE FROM analytics_snapshots")

        conn.executemany(
            """
            INSERT INTO sales_transactions(
                date, bl_code, client_name, manager_name, service_type, amount, cost, profit,
                currency, paid_amount, debt_amount, payment_status, source, source_sheet, raw_data_json,
                created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    row.get("date", ""),
                    row.get("bl_code", ""),
                    row.get("client_name", ""),
                    row.get("manager_name", ""),
                    row.get("service_type", ""),
                    float(row.get("amount") or 0),
                    float(row.get("cost") or 0),
                    float(row.get("profit") or 0),
                    row.get("currency", ""),
                    float(row.get("paid_amount") or 0),
                    float(row.get("debt_amount") or 0),
                    row.get("payment_status", ""),
                    row.get("source", "sheet"),
                    row.get("source_sheet", ""),
                    row.get("raw_data_json", ""),
                    _format_dt(),
                    _format_dt(),
                )
                for row in sales_rows
            ],
        )

        conn.executemany(
            """
            INSERT INTO cashflow_transactions(
                date, type, category, amount, currency, bank_or_cash, contractor, bl_code,
                reys_number, comment, source_sheet, raw_data_json, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    row.get("date", ""),
                    row.get("type", ""),
                    row.get("category", ""),
                    float(row.get("amount") or 0),
                    row.get("currency", ""),
                    row.get("bank_or_cash", ""),
                    row.get("contractor", ""),
                    row.get("bl_code", ""),
                    row.get("reys_number", ""),
                    row.get("comment", ""),
                    row.get("source_sheet", ""),
                    row.get("raw_data_json", ""),
                    _format_dt(),
                    _format_dt(),
                )
                for row in cashflow_rows
            ],
        )

        conn.executemany(
            """
            INSERT INTO shipments(
                bl_code, client_name, reys_number, fura_number, container_type, station, agent,
                logist_name, sales_manager_name, warehouse, status, loaded_date, arrived_date,
                expected_date, cargo_type, weight_kg, volume_m3, places, places_breakdown, description,
                source_sheet, raw_data_json, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    row.get("bl_code", ""),
                    row.get("client_name", ""),
                    row.get("reys_number", ""),
                    row.get("fura_number", ""),
                    row.get("container_type", ""),
                    row.get("station", ""),
                    row.get("agent", ""),
                    row.get("logist_name", ""),
                    row.get("sales_manager_name", ""),
                    row.get("warehouse", ""),
                    row.get("status", ""),
                    row.get("loaded_date", ""),
                    row.get("arrived_date", ""),
                    row.get("expected_date", ""),
                    row.get("cargo_type", ""),
                    float(row.get("weight_kg") or 0),
                    float(row.get("volume_m3") or 0),
                    int(row.get("places") or 0),
                    row.get("places_breakdown", ""),
                    row.get("description", ""),
                    row.get("source_sheet", ""),
                    row.get("raw_data_json", ""),
                    _format_dt(),
                    _format_dt(),
                )
                for row in shipment_rows
            ],
        )

        conn.execute(
            """
            INSERT INTO analytics_snapshots(
                date, total_sales, total_income, total_expense, profit, debt,
                active_bl_count, arrived_shipments_count, delayed_shipments_count, created_at
            )
            VALUES (?, 0, 0, 0, 0, 0, 0, 0, 0, ?)
            """,
            (datetime.now().strftime("%d.%m.%Y"), _format_dt()),
        )

        conn.commit()
    finally:
        conn.close()
    db.set_setting(ANALYTICS_LAST_SYNC_KEY, _format_dt())
    db.set_setting(ANALYTICS_EXCHANGE_RATES_KEY, _safe_json(rates))


def sync_workbook(sheets: dict[str, list[list[Any]]], source_name: str = "google", sheet_id: str | None = None) -> dict[str, Any]:
    log_id = _create_sync_log(source_name)
    sales_rows: list[dict[str, Any]] = []
    cashflow_rows: list[dict[str, Any]] = []
    shipment_rows: list[dict[str, Any]] = []
    skipped_rows = 0
    rates: dict[str, float] = {"UZS": 1.0}
    try:
        for sheet_name, matrix in sheets.items():
            if not matrix:
                continue
            rates.update(_extract_rates_from_sheet(sheet_name, matrix))
            shipping_rows = _parse_shipping_blocks(sheet_name, matrix)
            if shipping_rows:
                shipment_rows.extend(shipping_rows)
            sales_rows.extend(_parse_sales_sheet(sheet_name, matrix))
            cashflow_rows.extend(_parse_cashflow_sheet(sheet_name, matrix))

        shipment_rows = _merge_shipments(shipment_rows)
        imported_rows = len(sales_rows) + len(cashflow_rows) + len(shipment_rows)
        _store_import_result(sales_rows, cashflow_rows, shipment_rows, rates)
        if sheet_id:
            set_google_sheet_id(sheet_id)
        details = {
            "source": source_name,
            "sheet_id": sheet_id or "",
            "sheets": list(sheets.keys()),
            "sales_rows": len(sales_rows),
            "cashflow_rows": len(cashflow_rows),
            "shipment_rows": len(shipment_rows),
            "rates": rates,
        }
        _finish_sync_log(log_id, "success", imported_rows, skipped_rows, details)
        return {
            "ok": True,
            "rows_imported": imported_rows,
            "rows_skipped": skipped_rows,
            "details": details,
            "last_sync_at": db.get_setting(ANALYTICS_LAST_SYNC_KEY, ""),
        }
    except Exception as exc:
        _finish_sync_log(
            log_id,
            "failed",
            0,
            skipped_rows,
            {"source": source_name, "sheet_id": sheet_id or "", "sheets": list(sheets.keys())},
            str(exc),
        )
        raise


def sync_from_google(sheet_id: str | None = None) -> dict[str, Any]:
    target_sheet_id = _extract_sheet_id(sheet_id or get_google_sheet_id())
    status = google_credentials_status()
    if status.get("mode") == "api":
        workbook = _load_google_workbook(target_sheet_id)
        return sync_workbook(workbook, source_name="google_api", sheet_id=target_sheet_id)
    workbook = _load_public_google_workbook(target_sheet_id)
    return sync_workbook(workbook, source_name="google_public", sheet_id=target_sheet_id)


def sync_from_upload(file_storage) -> dict[str, Any]:
    workbook = _load_uploaded_workbook(file_storage)
    return sync_workbook(workbook, source_name="upload")


def get_sync_status() -> dict[str, Any]:
    conn = db.get_conn()
    try:
        last_log = conn.execute(
            """
            SELECT *
            FROM sheet_sync_logs
            ORDER BY started_at DESC, id DESC
            LIMIT 10
            """
        ).fetchall()
    finally:
        conn.close()
    connection = google_credentials_status()
    return {
        "sheet_id": get_google_sheet_id(),
        "connection": connection,
        "last_sync_at": db.get_setting(ANALYTICS_LAST_SYNC_KEY, ""),
        "exchange_rates": json.loads(db.get_setting(ANALYTICS_EXCHANGE_RATES_KEY, '{"UZS":1}')),
        "logs": [dict(row) for row in last_log],
    }
