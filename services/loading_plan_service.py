# -*- coding: utf-8 -*-
"""Loading-plan spreadsheet (партии / фуры).

The batches in the panel are OPENED and ADJUSTED from this Google
Sheets file. Monthly tabs (e.g. "AVGUST 2026", "Iyul2026") hold plan
blocks laid out in column groups:

    <date of departure>          ← e.g. 14.08.2026
    <plan title>                 ← e.g. "YIWU TO HORGOS - YARGXOL"
    SHIPPING MARK | CTN | T/CBM | KG | KG IN 1 CBM | DATE OF ARRIVE
    BL-309        | 2   | 1.53  | ...
    ...
    TOTAL         | ... |

Plan kinds (by title):
  - CHINA truck  ("YIWU TO HORGOS…", "ZHONGSHAN TO HORGOS…", legacy
    "YIWU MUHAMMAD" / "ZHONGSHAN YARGXOL"): loading from a China
    warehouse toward Horgos. A batch is opened from this plan.
  - KAZAKH truck ("HORGOS TO TASHKENT YIWU + ZH …"): loading at Horgos
    toward Tashkent; usually merges the BLs of the YIWU and ZHONGSHAN
    China trucks. When a batch reaches "Horgos (Qozoq)" its BL list is
    re-synced from this plan. The kazakh block sometimes lacks the
    SHIPPING MARK header row, so parsing anchors on date+title instead.
"""

import io
import os
import re
import threading
import time
import warnings
from datetime import datetime, date

import requests

LOADING_PLAN_SHEET_ID = (
    os.environ.get("LOADING_PLAN_SHEET_ID", "").strip()
    or "10xia-LGEIwi6zufRvfxP3vzwxEix3sf7sRP71M6hjPw"
)

CACHE_TTL_SECONDS = 120
_lock = threading.Lock()
_cache: dict = {}

_TITLE_KEYWORDS = ("YARGXOL", "YARGOL", "MUHAMMAD", "YIWU", "ZHONGSHAN", "HORGOS", "FURA")

# месячные вкладки: разные написания месяца по-узбекски/русски
_MONTH_TOKENS = {
    1: ("yanvar", "январ"), 2: ("fevral", "феврал"), 3: ("mart", "март"),
    4: ("aprel", "апрел"), 5: ("may", "май"), 6: ("iyun", "июн"),
    7: ("iyul", "июл"), 8: ("avgust", "август"), 9: ("sentabr", "сентяб"),
    10: ("oktabr", "октяб"), 11: ("noyabr", "нояб"), 12: ("dekabr", "декаб"),
}


def _download_workbook():
    from openpyxl import load_workbook

    url = f"https://docs.google.com/spreadsheets/d/{LOADING_PLAN_SHEET_ID}/export?format=xlsx"
    response = requests.get(url, timeout=90)
    response.raise_for_status()
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        return load_workbook(io.BytesIO(response.content), read_only=True, data_only=True)


def list_tabs(force: bool = False) -> list:
    data = _get_parsed(force=force)
    return sorted(data.keys())


def default_tab(tabs: list, today: date | None = None) -> str:
    today = today or datetime.now().date()
    tokens = _MONTH_TOKENS[today.month]
    year = str(today.year)
    best = ""
    for tab in tabs:
        low = tab.lower().replace(" ", "")
        if any(t in low for t in tokens) and year in low:
            best = tab
    return best or (tabs[-1] if tabs else "")


def _is_title(value) -> bool:
    if not isinstance(value, str):
        return False
    text = value.strip().upper()
    if not text or len(text) < 4:
        return False
    # отбрасываем текстовые даты вида "10-Aug" из колонок DATE OF ARRIVE
    if re.match(r"^\d{1,2}[-./]", text):
        return False
    return any(k in text for k in _TITLE_KEYWORDS)


def _fmt_date(value) -> str:
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    return str(value or "").strip()


def _num(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace("\xa0", "").replace(",", ".").strip() or 0)
    except ValueError:
        return 0.0


def _parse_tab(grid: list) -> list:
    """Extract every plan block from one tab's cell grid."""
    blocks = []
    n_rows = len(grid)
    for i in range(n_rows - 2):
        row = grid[i]
        for j, cell in enumerate(row):
            if not isinstance(cell, (datetime, date)):
                continue
            title_cell = grid[i + 1][j] if j < len(grid[i + 1]) else None
            if not _is_title(title_cell):
                continue
            title = str(title_cell).strip()
            # данные начинаются под заголовком; строку 'SHIPPING MARK'
            # пропускаем (в казахских блоках её может не быть)
            start = i + 2
            first = grid[start][j] if start < n_rows and j < len(grid[start]) else None
            if isinstance(first, str) and first.strip().upper().startswith("SHIPPING MARK"):
                start += 1
            items = []
            empty_streak = 0
            k = start
            while k < n_rows and empty_streak <= 2:
                mark_cell = grid[k][j] if j < len(grid[k]) else None
                mark = str(mark_cell).strip() if mark_cell is not None else ""
                if not mark:
                    empty_streak += 1
                    k += 1
                    continue
                empty_streak = 0
                if mark.upper() == "TOTAL":
                    break
                if isinstance(mark_cell, (datetime, date)):
                    # наткнулись на дату следующего блока в этой же колонке
                    break
                def cval(off):
                    idx = j + off
                    return grid[k][idx] if idx < len(grid[k]) else None
                arrive = ""
                for off in (5, 4, 6):
                    v = cval(off)
                    if isinstance(v, (datetime, date)):
                        arrive = _fmt_date(v)
                        break
                    if isinstance(v, str) and re.match(r"^\d{1,2}[-./]", v.strip()):
                        arrive = v.strip()
                        break
                items.append({
                    "mark": mark,
                    "ctn": _num(cval(1)),
                    "cbm": _num(cval(2)),
                    "kg": _num(cval(3)),
                    "arrive": arrive,
                })
                k += 1
            if not items:
                continue
            upper_title = title.upper()
            # казахская фура: «HORGOS TO TASHKENT…», «HORGOS - TASHKENT…»,
            # «HORGOS YARGXOL» — всё, что начинается с HORGOS или содержит
            # HORGOS не в виде «TO HORGOS» (китайская: «YIWU TO HORGOS…»)
            compact = re.sub(r"\s+", " ", upper_title).strip()
            kind = "china"
            if compact.startswith("HORGOS") or ("HORGOS" in compact and "TO HORGOS" not in compact):
                kind = "kazakh"
            warehouses = [w for w in ("YIWU", "ZHONGSHAN") if w in upper_title]
            if kind == "kazakh" and ("ZH" in upper_title and "ZHONGSHAN" not in warehouses):
                warehouses.append("ZHONGSHAN")
            blocks.append({
                "date": _fmt_date(cell),
                "title": title,
                "kind": kind,          # china (склад→Horgos) | kazakh (Horgos→Tashkent)
                "warehouses": warehouses,
                "items": items,
                "total_ctn": round(sum(x["ctn"] for x in items), 2),
                "total_cbm": round(sum(x["cbm"] for x in items), 3),
                "total_kg": round(sum(x["kg"] for x in items), 2),
            })
    return blocks


def _get_parsed(force: bool = False) -> dict:
    """{tab_name: [blocks]} for the whole workbook, cached."""
    now = time.monotonic()
    with _lock:
        cached = _cache.get("parsed")
        if not force and cached and cached["expires_at"] > now:
            return cached["data"]
    wb = _download_workbook()
    data = {}
    for ws in wb.worksheets:
        grid = [list(r) for r in ws.iter_rows(max_col=80, values_only=True)]
        try:
            blocks = _parse_tab(grid)
        except Exception:
            blocks = []
        if blocks:
            data[ws.title] = blocks
    with _lock:
        _cache["parsed"] = {"data": data, "expires_at": time.monotonic() + CACHE_TTL_SECONDS}
    return data


def get_loading_plans(tab: str = "", force: bool = False) -> dict:
    """Plan blocks of one tab (default: the current month's tab)."""
    data = _get_parsed(force=force)
    tabs = sorted(data.keys())
    chosen = ""
    if tab:
        wanted = tab.strip().lower().replace(" ", "")
        for name in tabs:
            if name.lower().replace(" ", "") == wanted:
                chosen = name
                break
        if not chosen:
            for name in tabs:
                if wanted in name.lower().replace(" ", ""):
                    chosen = name
                    break
        if not chosen:
            return {"ok": False, "error": f"Вкладка «{tab}» не найдена", "tabs": tabs}
    else:
        chosen = default_tab(tabs)
    if not chosen:
        return {"ok": False, "error": "В шитсе не найдено ни одного блока планов", "tabs": tabs}
    return {"ok": True, "tab": chosen, "tabs": tabs, "plans": data[chosen]}


def find_plan(tab: str, title: str, plan_date: str = "", force: bool = False) -> dict | None:
    """Locate ONE plan block by tab + (partial) title + optional date."""
    result = get_loading_plans(tab, force=force)
    if not result.get("ok"):
        return None
    wanted_title = (title or "").strip().upper()
    wanted_date = (plan_date or "").strip()
    candidates = []
    for block in result["plans"]:
        if wanted_title and wanted_title not in block["title"].upper():
            continue
        if wanted_date and block["date"] != wanted_date:
            continue
        candidates.append(block)
    if len(candidates) == 1:
        found = dict(candidates[0])
        found["tab"] = result["tab"]
        return found
    return None
